import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import urllib.request
import json, re, sqlite3, os
from datetime import datetime

st.set_page_config(page_title="Micki Analítica", page_icon="🏀", layout="wide", initial_sidebar_state="expanded")

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "historic.db")
API_BASE = "https://msstats.optimalwayconsulting.com/v1/fcbq/getJsonWithMatchMoves/{match_id}?currentSeason=true"
WEB_BASE = "https://www.basquetcatala.cat/estadistiques/2025/{match_id}"
COLOR_A, COLOR_B = "#185FA5", "#993C1D"

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');

/* ── Base ── */
html,body,[class*="css"]{font-family:'Inter',sans-serif;color:#e8eaf0!important;}
.stApp{background:#EBF4FC!important;color:#1a2744!important;}
.main .block-container{padding-top:1.5rem!important;}
p, span, div, label, h1, h2, h3, h4{color:#1a2744;}

/* ── Sidebar ── */
div[data-testid="stSidebar"]{background:#D6E8F7!important;border-right:1px solid #B5D4F4!important;}

/* ── Pestanyes grans estil NBA ── */
.stTabs [data-baseweb="tab-list"]{
    background:#D6E8F7!important;
    border-radius:12px!important;
    padding:6px!important;
    gap:4px!important;
    border:1px solid #B5D4F4!important;
}
.stTabs [data-baseweb="tab"]{
    border-radius:8px!important;
    font-size:14px!important;
    font-weight:600!important;
    color:#185FA5!important;
    padding:10px 18px!important;
    letter-spacing:0.02em!important;
    transition:all 0.2s!important;
}
.stTabs [data-baseweb="tab"]:hover{
    background:#B5D4F4!important;
    color:#0C447C!important;
}
.stTabs [aria-selected="true"]{
    background:#185FA5!important;
    color:#ffffff!important;
    box-shadow:0 2px 8px rgba(24,95,165,0.4)!important;
}

/* ── Botons ── */
.stButton button{
    background:#85B7EB!important;
    color:#042C53!important;
    border:none!important;
    border-radius:8px!important;
    font-weight:700!important;
    font-size:13px!important;
    padding:8px 20px!important;
    transition:all 0.2s!important;
}
.stButton button:hover{
    background:#378ADD!important;
    color:#ffffff!important;
}
/* Download button */
a[data-testid="stDownloadButton"] button,
div[data-testid="stDownloadButton"] button{
    background:#85B7EB!important;
    color:#042C53!important;
    font-weight:700!important;
    border:none!important;
    border-radius:8px!important;
}
a[data-testid="stDownloadButton"] button:hover,
div[data-testid="stDownloadButton"] button:hover{
    background:#378ADD!important;
    color:#ffffff!important;
}

/* ── DataFrames ── */
div[data-testid="stDataFrame"]{
    border-radius:10px!important;
    overflow:hidden!important;
    border:1px solid #1f2330!important;
}
div[data-testid="stDataFrame"] th{
    background:#D6E8F7!important;
    color:#185FA5!important;
}
div[data-testid="stDataFrame"] td{
    background:#EBF4FC!important;
    color:#1a2744!important;
}

/* ── Inputs i selects ── */
.stSelectbox>div>div,
.stMultiSelect>div>div,
.stTextInput>div>div,
.stTextInput input,
.stSelectbox input,
input, textarea, select{
    border-radius:8px!important;
    background:#232840!important;
    border-color:#2d3450!important;
    color:#e8eaf0!important;
}
/* Text dins dels camps de selecció */
[data-baseweb="select"] [data-baseweb="tag"],
[data-baseweb="select"] input,
[data-baseweb="select"] div,
[data-baseweb="input"] input{
    color:#e8eaf0!important;
    background:#D6E8F7!important;
}
/* Placeholder text */
input::placeholder, textarea::placeholder{color:#378ADD!important;}
/* Dropdown options */
[data-baseweb="popover"] li,
[data-baseweb="menu"] li{
    background:#232840!important;
    color:#e8eaf0!important;
}
[data-baseweb="popover"] li:hover,
[data-baseweb="menu"] li:hover{
    background:#B5D4F4!important;
}

/* ── Expanders ── */
.stExpander{
    border:1px solid #1f2330!important;
    border-radius:10px!important;
    background:#13161e!important;
}

/* ── Mètriques ── */
div[data-testid="stMetric"]{
    background:#ffffff!important;
    border-radius:10px!important;
    padding:12px!important;
    border:1px solid #1f2330!important;
}

/* ── Info/Warning boxes ── */
div[data-testid="stAlert"]{
    border-radius:8px!important;
    border:1px solid #1f2330!important;
}

/* ── Plotly charts fons transparent ── */
.js-plotly-plot .plotly .modebar{background:transparent!important;}

/* ── Captions i text secundari ── */
.stCaption, small{color:#378ADD!important;}

/* ── Separadors ── */
hr{border-color:#B5D4F4!important;}

/* ── Fix sidebar i expanders ── */
div[data-testid="stSidebar"],
div[data-testid="stSidebar"] *,
div[data-testid="stSidebar"] label,
div[data-testid="stSidebar"] p,
div[data-testid="stSidebar"] span,
div[data-testid="stSidebar"] div{
    color:#1a2744!important;
}
div[data-testid="stSidebar"] input,
div[data-testid="stSidebar"] textarea,
div[data-testid="stSidebar"] [data-baseweb="input"] div,
div[data-testid="stSidebar"] [data-baseweb="input"] input,
div[data-testid="stSidebar"] div[class*="Input"],
div[data-testid="stSidebar"] div[class*="input"]{
    background:#ffffff!important;
    color:#1a2744!important;
    border:1px solid #B5D4F4!important;
}

/* Expanders — tot el contingut interior visible */
div[data-testid="stExpander"]{
    background:#ffffff!important;
    border:1px solid #B5D4F4!important;
    border-radius:10px!important;
}
div[data-testid="stExpander"] *,
div[data-testid="stExpander"] p,
div[data-testid="stExpander"] span,
div[data-testid="stExpander"] div,
div[data-testid="stExpander"] td,
div[data-testid="stExpander"] th,
div[data-testid="stExpander"] label{
    color:#1a2744!important;
    background:#ffffff!important;
}
div[data-testid="stExpander"] summary{
    color:#1a2744!important;
    background:#EBF4FC!important;
}
/* Força fons blanc pur al contingut de l'expander */
div[data-testid="stExpander"] > div:last-child{
    background:#ffffff!important;
}
div[data-testid="stExpander"] summary:hover{
    background:#D6E8F7!important;
}
/* DataFrames dins expanders i globals */
div[data-testid="stExpander"] div[data-testid="stDataFrame"] *,
div[data-testid="stDataFrame"] *,
div[data-testid="stDataFrame"] td,
div[data-testid="stDataFrame"] th,
div[data-testid="stDataFrame"] span,
div[data-testid="stDataFrame"] div,
.dvn-scroller *,
.dvn-cell *,
[role="gridcell"] *,
[role="columnheader"] *{
    color:#1a2744!important;
}
/* Headers del dataframe */
div[data-testid="stDataFrame"] th,
div[data-testid="stDataFrame"] [role="columnheader"]{
    background:#D6E8F7!important;
    color:#0C447C!important;
}
/* Files alternatives */
div[data-testid="stDataFrame"] [role="gridcell"]{
    background:#ffffff!important;
    color:#1a2744!important;
}

/* ── Fix desplegables BaseWeb (selectbox, multiselect) ── */
/* Fons i text del camp tancat */
div[data-baseweb="select"] > div:first-child{
    background:#ffffff!important;
    border-color:#B5D4F4!important;
}
/* Text valor seleccionat i placeholder */
div[data-baseweb="select"] span,
div[data-baseweb="select"] div,
div[data-baseweb="select"] p,
div[data-baseweb="select"] input{
    color:#1a2744!important;
    background:transparent!important;
}
/* Llista desplegada - fons de cada opció */
ul[data-baseweb="menu"],
ul[data-baseweb="menu"] li,
ul[data-baseweb="menu"] li *,
div[data-baseweb="popover"],
div[data-baseweb="popover"] *,
li[role="option"],
li[role="option"] *{
    background:#ffffff!important;
    color:#1a2744!important;
}
/* Opció hover */
li[role="option"]:hover,
li[role="option"]:hover *{
    background:#D6E8F7!important;
    color:#0C447C!important;
}
/* Opció seleccionada */
li[aria-selected="true"],
li[aria-selected="true"] *{
    background:#B5D4F4!important;
    color:#0C447C!important;
    font-weight:600!important;
}
/* Tags del multiselect */
span[data-baseweb="tag"],
span[data-baseweb="tag"] *{
    background:#185FA5!important;
    color:#ffffff!important;
}
</style>
""", unsafe_allow_html=True)

def card(label, value, sub="", color="#185FA5"):
    return f"""<div style="background:#fff;border:0.5px solid #e2e4e8;border-radius:10px;padding:14px 16px;text-align:center;margin-bottom:8px">
    <div style="font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:#9ca3af;margin-bottom:4px">{label}</div>
    <div style="font-size:28px;font-weight:600;color:{color};line-height:1.1">{value}</div>
    <div style="font-size:11px;color:#9ca3af;margin-top:3px">{sub}</div></div>"""

def sec(s):
    return f'<div style="font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.08em;color:#6b7280;border-left:3px solid #185FA5;padding-left:8px;margin:24px 0 10px;border-radius:0">{s}</div>'

def badge(text, color, bg):
    return f'<span style="background:{bg};color:{color};font-size:10px;font-weight:600;padding:2px 7px;border-radius:20px;letter-spacing:.04em">{text}</span>'

def chart_style(fig, h=280, title=""):
    if title:
        fig.update_layout(title=dict(text=title, font=dict(color="#374151", size=13, family="Inter"), x=0))
    fig.update_layout(
        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
        font=dict(color="#374151", family="Inter", size=12),
        legend=dict(bgcolor="#ffffff", bordercolor="#e2e4e8", borderwidth=1, title="",
                    orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        xaxis=dict(showgrid=False, color="#9ca3af", linecolor="#e2e4e8"),
        yaxis=dict(showgrid=True, gridcolor="#f3f4f6", color="#9ca3af", linecolor="#e2e4e8"),
        margin=dict(l=0, r=0, t=40 if title else 10, b=0), height=h)
    return fig

def eff_color(p):
    if p >= 55: return "#16a34a"
    if p >= 35: return "#d97706"
    return "#dc2626"

def shot_map_svg(zones, width=300, height=280):
    pts = [
        {"val":1, "label":"1pt",  "cy":248, "made":zones[0][0], "miss":zones[0][1]},
        {"val":2, "label":"2pts", "cy":175, "made":zones[1][0], "miss":zones[1][1]},
        {"val":3, "label":"3pts", "cy":76,  "made":zones[2][0], "miss":zones[2][1]},
    ]
    max_t = max((z["made"]+z["miss"] for z in pts), default=1) or 1
    cx = width // 2
    L = [f'<svg viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg" style="width:100%;display:block">']
    L.append(f'<rect x="0" y="0" width="{width}" height="{height}" fill="#f9fafb" rx="8"/>')
    L.append(f'<rect x="20" y="15" width="{width-40}" height="{height-15}" fill="none" stroke="#e5e7eb" stroke-width="1"/>')
    L.append(f'<rect x="{cx-52}" y="170" width="104" height="90" fill="none" stroke="#d1d5db" stroke-width="1"/>')
    L.append(f'<path d="M 28 170 A {cx-28} {cx-28} 0 0 1 {width-28} 170" fill="none" stroke="#d1d5db" stroke-width="1" stroke-dasharray="5,3"/>')
    L.append(f'<line x1="28" y1="115" x2="28" y2="{height}" stroke="#d1d5db" stroke-width="1" stroke-dasharray="5,3"/>')
    L.append(f'<line x1="{width-28}" y1="115" x2="{width-28}" y2="{height}" stroke="#d1d5db" stroke-width="1" stroke-dasharray="5,3"/>')
    L.append(f'<path d="M {cx-28} 260 A 28 28 0 0 1 {cx+28} 260" fill="none" stroke="#d1d5db" stroke-width="1"/>')
    L.append(f'<circle cx="{cx}" cy="242" r="5" fill="none" stroke="#9ca3af" stroke-width="1.5"/>')
    L.append(f'<line x1="20" y1="125" x2="{width-20}" y2="125" stroke="#eff0f1" stroke-width="0.5" stroke-dasharray="4,4"/>')
    L.append(f'<line x1="20" y1="215" x2="{width-20}" y2="215" stroke="#eff0f1" stroke-width="0.5" stroke-dasharray="4,4"/>')
    for z in pts:
        t = z["made"] + z["miss"]
        if t == 0:
            L.append(f'<circle cx="{cx}" cy="{z["cy"]}" r="20" fill="#f3f4f6"/>')
            L.append(f'<text x="{cx}" y="{z["cy"]}" text-anchor="middle" dominant-baseline="middle" font-size="9" fill="#9ca3af">—</text>')
            continue
        p = round(z["made"]/t*100)
        col = eff_color(p)
        r_out = 18 + round((t/max_t)*30)
        r_in  = round(r_out * 0.55)
        L.append(f'<circle cx="{cx}" cy="{z["cy"]}" r="{r_out}" fill="{col}" fill-opacity="0.12" stroke="{col}" stroke-width="1.5" stroke-opacity="0.35"/>')
        L.append(f'<circle cx="{cx}" cy="{z["cy"]}" r="{r_in}" fill="{col}" fill-opacity="0.88"/>')
        L.append(f'<text x="{cx}" y="{z["cy"]}" text-anchor="middle" dominant-baseline="middle" font-size="11" font-weight="600" fill="white">{p}%</text>')
        L.append(f'<text x="{cx}" y="{z["cy"]+r_out+11}" text-anchor="middle" font-size="9" fill="{col}" font-weight="500">{z["made"]}/{t}</text>')
        L.append(f'<text x="{cx+r_out+7}" y="{z["cy"]}" dominant-baseline="middle" font-size="9" fill="#9ca3af">{z["label"]}</text>')
    L.append('</svg>')
    return "\n".join(L)

# ══════════════════════════════════════════════════
# BASE DE DADES
# ══════════════════════════════════════════════════
# Migració automàtica: afegir columnes noves si no existeixen
def migrate_db():
    con = sqlite3.connect(DB_PATH)
    try:
        con.execute("ALTER TABLE stats_jugador ADD COLUMN minuts REAL DEFAULT 0")
        con.commit()
    except:
        pass  # La columna ja existeix
    con.close()

def init_db():
    con = sqlite3.connect(DB_PATH)
    con.executescript("""
    CREATE TABLE IF NOT EXISTS partits (
        match_id TEXT PRIMARY KEY, data_consulta TEXT,
        nom_a TEXT, nom_b TEXT, id_equip_a TEXT, id_equip_b TEXT,
        score_a INTEGER, score_b INTEGER, total_jugades INTEGER
    );
    CREATE TABLE IF NOT EXISTS jugades (
        id INTEGER PRIMARY KEY AUTOINCREMENT, match_id TEXT,
        num INTEGER, quart INTEGER, min_num REAL, temps TEXT,
        id_equip TEXT, dorsal TEXT, jugador TEXT, accio TEXT,
        marcador TEXT, punts INTEGER, team_action INTEGER
    );
    CREATE TABLE IF NOT EXISTS stats_jugador (
        id INTEGER PRIMARY KEY AUTOINCREMENT, match_id TEXT, data_consulta TEXT,
        jugador TEXT, equip_nom TEXT, punts INTEGER,
        cistelles_2 INTEGER, cistelles_3 INTEGER, tirs_lliures INTEGER,
        faltes INTEGER, accions INTEGER, impacte INTEGER, pts_per_min REAL,
        minuts REAL DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS shots_zones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        match_id TEXT, data_consulta TEXT, equip_nom TEXT, jugador TEXT,
        val1_made INTEGER, val1_miss INTEGER,
        val2_made INTEGER, val2_miss INTEGER,
        val3_made INTEGER, val3_miss INTEGER
    );
    CREATE TABLE IF NOT EXISTS timeouts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        match_id TEXT, data_consulta TEXT,
        equip_nom TEXT, quart INTEGER,
        min_timeout REAL, min_cistella REAL,
        segons_resposta REAL,
        jugadora TEXT, accio TEXT,
        va_anotar INTEGER, dins_24s INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS equips (
        id_equip TEXT PRIMARY KEY, nom TEXT
    );
    """)
    con.commit(); con.close()

def save_nom_equip(id_equip, nom):
    con = sqlite3.connect(DB_PATH)
    con.execute("INSERT OR REPLACE INTO equips (id_equip, nom) VALUES (?,?)", (str(id_equip), nom))
    con.commit(); con.close()

def load_noms_equips():
    con = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql("SELECT id_equip, nom FROM equips", con)
        con.close(); return dict(zip(df["id_equip"], df["nom"]))
    except:
        con.close(); return {}

def partit_exists(match_id):
    con = sqlite3.connect(DB_PATH)
    r = con.execute("SELECT 1 FROM partits WHERE match_id=?", (match_id,)).fetchone()
    con.close(); return r is not None

def save_partit(match_id, df, nom_a, nom_b, id_a, id_b, score_a, score_b):
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM partits WHERE match_id=?", (match_id,))
    con.execute("DELETE FROM jugades WHERE match_id=?", (match_id,))
    con.execute("DELETE FROM stats_jugador WHERE match_id=?", (match_id,))
    con.execute("DELETE FROM shots_zones WHERE match_id=?", (match_id,))
    con.execute("INSERT INTO partits VALUES (?,?,?,?,?,?,?,?,?)",
        (match_id, datetime.now().strftime("%Y-%m-%d %H:%M"),
         nom_a, nom_b, str(id_a), str(id_b), score_a, score_b, len(df)))
    rows = [(match_id, int(r["num"]), int(r["quart"]) if r["quart"]!="" else 0,
             float(r["min_num"]), str(r["temps"]), str(r["idEquip"]), str(r["dorsal"]),
             str(r["jugador"]), str(r["accio"]), str(r["marcador"]), int(r["punts"]),
             int(r.get("teamAction",0) or 0)) for _,r in df.iterrows()]
    con.executemany("INSERT INTO jugades (match_id,num,quart,min_num,temps,id_equip,dorsal,jugador,accio,marcador,punts,team_action) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", rows)
    con.commit(); con.close()

MINS_PER_QUART = 10

def calc_minuts_reals(df):
    """Calcula els minuts reals jugats per cada jugador usant Entra/Surt del camp."""
    players = {}
    for _, m in df.iterrows():
        nom = m.get("jugador","")
        if not nom or str(nom) in ("","nan"): continue
        idmove_str = str(m.get("accio",""))
        min_ = float(m.get("min_num", 0))
        # Reconstruïm minut absolut: min_num ja és minut dins del quart
        quart = int(m.get("quart", 1)) if m.get("quart","") != "" else 1
        t = (quart - 1) * MINS_PER_QUART + min_

        if nom not in players:
            players[nom] = {"intervals": [], "entrada": None}

        if "Entra al camp" in idmove_str:
            players[nom]["entrada"] = t
        elif "Surt del camp" in idmove_str:
            if players[nom]["entrada"] is not None:
                players[nom]["intervals"].append((players[nom]["entrada"], t))
                players[nom]["entrada"] = None
            else:
                inici_quart = (quart - 1) * MINS_PER_QUART
                players[nom]["intervals"].append((inici_quart, t))

    # Tanca intervals oberts
    for nom, p in players.items():
        if p["entrada"] is not None:
            p["intervals"].append((p["entrada"], p["entrada"] + MINS_PER_QUART))

    minuts = {}
    for nom, p in players.items():
        total = sum(fi - ini for ini, fi in p["intervals"])
        minuts[nom] = round(total, 1)
    return minuts

def calc_usage_rate(df_jug, df_equip_on_pista):
    """Calcula l'Usage Rate d'una jugadora.
    df_jug = accions de la jugadora
    df_equip_on_pista = accions de l'equip mentre la jugadora és a pista
    """
    tc_jug = int(df_jug["accio"].str.contains(
        "Cistella de 2|Cistella de 3|Intent fallat de 2|Intent fallat de 3|fallat de 2|fallat de 3",
        case=False, na=False).sum())
    tl_jug = int(df_jug["accio"].str.contains(
        "Cistella de 1|Intent fallat de 1", case=False, na=False).sum())

    tc_eq = int(df_equip_on_pista["accio"].str.contains(
        "Cistella de 2|Cistella de 3|Intent fallat de 2|Intent fallat de 3|fallat de 2|fallat de 3",
        case=False, na=False).sum())
    tl_eq = int(df_equip_on_pista["accio"].str.contains(
        "Cistella de 1|Intent fallat de 1", case=False, na=False).sum())

    num_jug = tc_jug + 0.44 * tl_jug
    den_eq  = tc_eq  + 0.44 * tl_eq
    if den_eq == 0: return 0
    return round(num_jug / den_eq * 100, 1)

def save_stats_jugador(match_id, data_consulta, df, teams, team_names):
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM stats_jugador WHERE match_id=?", (match_id,))
    minuts_reals = calc_minuts_reals(df)
    # Detecta nom de columna de jugador
    col_jug_fn = "jugador" if "jugador" in df.columns else "jugadora"
    # Unifica columna per facilitar el codi
    df = df.copy()
    df["jugador"] = df[col_jug_fn].fillna("")
    rows = []
    for jug in df["jugador"].unique():
        if not jug or str(jug) in ("","nan"): continue
        dj = df[df["jugador"]==jug]
        eq_id = dj["idEquip"].iloc[0]
        eq_nom = team_names.get(str(eq_id),"?")
        punts   = int(dj["punts"].sum())
        cist2   = int(dj["accio"].str.contains("Cistella de 2",case=False,na=False).sum())
        cist3   = int(dj["accio"].str.contains("Cistella de 3",case=False,na=False).sum())
        tl      = int(dj["accio"].str.contains("Cistella de 1",case=False,na=False).sum())
        faltes  = int(dj["accio"].str.contains("falta",case=False,na=False).sum())
        accions = len(dj)
        rival = [t for t in teams if t!=eq_id]
        rival_id = rival[0] if rival else None

        # Determina el nom de la columna de jugador
        col_jug = "jugador" if "jugador" in df.columns else "jugadora"

        # Usa intervals reals (Entra/Surt) per calcular el +/-
        MINS_Q = 10
        intervals_imp = []; en_pista_imp = {}
        # Detecta si la jugadora comença a pista (primer event és Surt sense Entra previ)
        primer_event = df[df["jugador"]==jug].sort_values("num").iloc[0] if not df[df["jugador"]==jug].empty else None
        if primer_event is not None:
            primer_accio = str(primer_event.get("accio",""))
            primer_quart = int(primer_event.get("quart",1))
            if "Surt" in primer_accio and "camp" in primer_accio:
                # Comença a pista des de l'inici del quart
                en_pista_imp[jug] = (primer_quart-1)*MINS_Q
        for _, row_imp in df.sort_values("num").iterrows():
            jug_imp = str(row_imp.get(col_jug, ""))
            if jug_imp != str(jug): continue
            accio_imp = str(row_imp.get("accio",""))
            quart_imp = int(row_imp.get("quart",1))
            min_imp = float(row_imp.get("min_num",0))
            t_imp = (quart_imp-1)*MINS_Q + (MINS_Q - min_imp if min_imp <= MINS_Q else min_imp)
            t_imp = max(0, min(t_imp, quart_imp*MINS_Q))
            if "Entra" in accio_imp and "camp" in accio_imp:
                en_pista_imp[jug] = t_imp
            elif "Surt" in accio_imp and "camp" in accio_imp:
                ti = en_pista_imp.pop(jug, (quart_imp-1)*MINS_Q)
                if t_imp > ti: intervals_imp.append((ti, t_imp))
            elif "Final de període" in accio_imp or "Final període" in accio_imp:
                fi_imp = quart_imp * MINS_Q
                if jug in en_pista_imp:
                    ti = en_pista_imp.pop(jug)
                    if fi_imp > ti: intervals_imp.append((ti, fi_imp))
        for ti_o in en_pista_imp.values():
            fi_o = df["quart"].max() * MINS_Q if not df.empty else 40
            if fi_o > ti_o: intervals_imp.append((ti_o, fi_o))
        # Normalitza intervals (elimina np.int64)
        intervals_imp = [(float(ti), float(tf)) for ti, tf in intervals_imp]

        if not intervals_imp:
            # Fallback: rang de num
            n_min,n_max = dj["num"].min(),dj["num"].max()
            dr = df[(df["num"]>=n_min)&(df["num"]<=n_max)]
            pf = int(dr[dr["idEquip"]==eq_id]["punts"].sum())
            pc = int(dr[dr["idEquip"]==rival_id]["punts"].sum()) if rival_id else 0
            import sys; print(f"FALLBACK {jug}: pf={pf} pc={pc}", file=sys.stderr)
        else:
            import sys; print(f"INTERVALS {jug}: {intervals_imp}", file=sys.stderr)
            df_t = df.copy()
            df_t["t_abs"] = df_t.apply(
                lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
                if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)
            pf = pc = 0
            for ti_r, tf_r in intervals_imp:
                df_i = df_t[(df_t["t_abs"]>=ti_r)&(df_t["t_abs"]<=tf_r)]
                pf_i = int(df_i[df_i["idEquip"]==eq_id]["punts"].sum())
                pc_i = int(df_i[df_i["idEquip"]==rival_id]["punts"].sum()) if rival_id else 0
                print(f"  tram {ti_r:.1f}-{tf_r:.1f}: pf={pf_i} pc={pc_i}", file=sys.stderr)
                pf += pf_i
                pc += pc_i
            print(f"  TOTAL {jug}: pf={pf} pc={pc} impacte={pf-pc}", file=sys.stderr)
            df_t["t_abs"] = df_t.apply(
                lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
                if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)
            pf = pc = 0
            for ti_r,tf_r in intervals_imp:
                df_i = df_t[(df_t["t_abs"]>=ti_r)&(df_t["t_abs"]<=tf_r)]
                pf += int(df_i[df_i["idEquip"]==eq_id]["punts"].sum())
                if rival_id:
                    pc += int(df_i[df_i["idEquip"]==rival_id]["punts"].sum())
        impacte = pf - pc
        min_jug = minuts_reals.get(jug, 0)
        pts_min = round(punts/min_jug, 2) if min_jug > 0 else 0.0
        # Usage Rate
        usage = 0.0
        if intervals_imp:
            df_t_us = df.copy()
            df_t_us["t_abs"] = df_t_us.apply(
                lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
                if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)
            mask_on_us = df_t_us["t_abs"].apply(
                lambda t: any(ti<=t<=tf for ti,tf in intervals_imp))
            df_eq_on = df_t_us[mask_on_us & (df_t_us["idEquip"]==eq_id)]
            usage = calc_usage_rate(dj, df_eq_on)
        else:
            usage = calc_usage_rate(dj, df[df["idEquip"]==eq_id])
        rows.append((match_id,data_consulta,jug,eq_nom,punts,cist2,cist3,tl,faltes,accions,impacte,pts_min,round(usage,1)))
    # Afegir minuts a cada row
    rows_amb_min = [r + (minuts_reals.get(r[2], 0),) for r in rows]
    con.executemany("INSERT INTO stats_jugador (match_id,data_consulta,jugador,equip_nom,punts,cistelles_2,cistelles_3,tirs_lliures,faltes,accions,impacte,pts_per_min,minuts,usage_rate) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", rows_amb_min)
    con.commit(); con.close()

def save_shots_zones(match_id, data_consulta, df, team_names):
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM shots_zones WHERE match_id=?", (match_id,))
    players_list = [("__equip__", eq) for eq in df["idEquip"].unique() if eq and str(eq)!="0"]
    for jug in df["jugador"].unique():
        if jug and str(jug) not in ("","nan"):
            eq_id = df[df["jugador"]==jug]["idEquip"].iloc[0]
            players_list.append((jug, eq_id))
    for jug, eq_id in players_list:
        dj = df[df["idEquip"]==eq_id] if jug=="__equip__" else df[df["jugador"]==jug]
        eq_nom = team_names.get(str(eq_id),"?")
        v1m = int(dj["accio"].str.contains("Cistella de 1|Tir lliure convertit",case=False,na=False).sum())
        v1x = int(dj["accio"].str.contains("Intent fallat de 1",case=False,na=False).sum())
        v2m = int(dj["accio"].str.contains("Cistella de 2",case=False,na=False).sum())
        v2x = int(dj["accio"].str.contains("Intent fallat de 2|fall.*2|2.*fall",case=False,na=False).sum())
        v3m = int(dj["accio"].str.contains("Cistella de 3",case=False,na=False).sum())
        v3x = int(dj["accio"].str.contains("Intent fallat de 3|fall.*3|3.*fall",case=False,na=False).sum())
        con.execute("INSERT INTO shots_zones (match_id,data_consulta,equip_nom,jugador,val1_made,val1_miss,val2_made,val2_miss,val3_made,val3_miss) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (match_id,data_consulta,eq_nom,jug,v1m,v1x,v2m,v2x,v3m,v3x))
    con.commit(); con.close()

def load_jugades_db(match_id):
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT * FROM jugades WHERE match_id=? ORDER BY num", con, params=(match_id,))
    con.close()
    return df.rename(columns={"id_equip":"idEquip","team_action":"teamAction"})

def load_partits_db():
    con = sqlite3.connect(DB_PATH)
    try: df = pd.read_sql("SELECT * FROM partits ORDER BY data_consulta DESC", con)
    except: df = pd.DataFrame()
    con.close(); return df

def load_stats_jugador_db():
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT * FROM stats_jugador ORDER BY data_consulta", con)
    con.close(); return df

def load_shots_zones_db():
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT * FROM shots_zones ORDER BY data_consulta", con)
    con.close(); return df

def delete_partit_db(match_id):
    con = sqlite3.connect(DB_PATH)
    for tbl in ["partits","jugades","stats_jugador","shots_zones","timeouts"]:
        con.execute(f"DELETE FROM {tbl} WHERE match_id=?", (match_id,))
    con.commit(); con.close()

init_db()

# ── Funcions de temps morts ────────────────────────────────────────────────
MINS_PER_QUART = 10

def analyze_timeouts(df, team_names):
    results = []
    moves = df.to_dict('records')
    for i, m in enumerate(moves):
        if 'Temps mort' not in str(m.get('accio','')):
            continue
        eq_id  = m.get('idEquip','')
        quart  = int(m.get('quart',1)) if m.get('quart','') != '' else 1
        min_to = float(m.get('min_num',0))
        min_abs_to = (quart-1)*MINS_PER_QUART + min_to
        eq_nom = team_names.get(str(eq_id),'?')
        va_anotar = 0; jugadora = ''; accio_cist = ''; mins_cist = None
        for j in range(i+1, len(moves)):
            nm = moves[j]
            q_next = int(nm.get('quart',1)) if nm.get('quart','') != '' else 1
            if q_next != quart: break
            move_str = str(nm.get('accio',''))
            if nm.get('idEquip','') == eq_id and any(c in move_str for c in ['Cistella de 1','Cistella de 2','Cistella de 3']):
                min_cist = float(nm.get('min_num',0))
                mins_cist = (q_next-1)*MINS_PER_QUART + min_cist
                jugadora = str(nm.get('jugador',''))
                accio_cist = move_str
                va_anotar = 1
                break
        # Calcula diferència en MINUTS de joc (no de vídeo)
        # min_abs_to i mins_cist són en minuts absoluts de joc
        if mins_cist is not None:
            diff_min = mins_cist - min_abs_to  # positiu = cistella DESPRÉS del TM
            segons = round(diff_min * 60, 1)
            dins_24s = (diff_min >= 0 and diff_min * 60 <= 24)
        else:
            segons = None
            dins_24s = False
        results.append({'equip_nom':eq_nom,'quart':quart,'min_timeout':round(min_abs_to,2),
            'min_cistella':round(mins_cist,2) if mins_cist else None,
            'segons_resposta':segons,'jugadora':jugadora,'accio':accio_cist,
            'va_anotar':va_anotar,'dins_24s':dins_24s})
    return results

def save_timeouts(match_id, data_consulta, df, team_names):
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM timeouts WHERE match_id=?", (match_id,))
    results = analyze_timeouts(df, team_names)
    for r in results:
        con.execute(
            "INSERT INTO timeouts (match_id,data_consulta,equip_nom,quart,min_timeout,min_cistella,segons_resposta,jugadora,accio,va_anotar,dins_24s) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (match_id,data_consulta,r['equip_nom'],r['quart'],r['min_timeout'],
             r['min_cistella'],r['segons_resposta'],r['jugadora'],r['accio'],
             r['va_anotar'],r.get('dins_24s',0)))
    con.commit(); con.close()

def load_timeouts_db():
    con = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql("SELECT * FROM timeouts ORDER BY data_consulta, min_timeout", con)
    except:
        df = pd.DataFrame()
    con.close(); return df

# Migració automàtica: afegir columnes noves si no existeixen a BD antigues
def migrate_db():
    con = sqlite3.connect(DB_PATH)
    try:
        con.execute("ALTER TABLE stats_jugador ADD COLUMN minuts REAL DEFAULT 0")
        con.commit()
    except Exception:
        pass
    try:
        con.execute("ALTER TABLE stats_jugador ADD COLUMN usage_rate REAL DEFAULT 0")
        con.commit()
    except Exception:
        pass
    try:
        con.execute("ALTER TABLE timeouts ADD COLUMN dins_24s INTEGER DEFAULT 0")
        con.commit()
    except Exception:
        pass
    try:
        con.executescript("""
        CREATE TABLE IF NOT EXISTS timeouts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id TEXT, data_consulta TEXT,
            equip_nom TEXT, quart INTEGER,
            min_timeout REAL, min_cistella REAL,
            segons_resposta REAL,
            jugadora TEXT, accio TEXT,
            va_anotar INTEGER
        );""")
        con.commit()
    except Exception:
        pass
    con.close()

migrate_db()

# ══════════════════════════════════════════════════
# FETCH
# ══════════════════════════════════════════════════
def extract_match_id(text):
    m = re.search(r"/([a-f0-9]{24})(?:\?|$)", text)
    if m: return m.group(1)
    if re.match(r"^[a-f0-9]{24}$", text.strip()): return text.strip()
    return None

def fetch_and_parse(match_id):
    url = API_BASE.format(match_id=match_id)
    req = urllib.request.Request(url, headers={
        "User-Agent":"Mozilla/5.0","Accept":"application/json",
        "Referer":WEB_BASE.format(match_id=match_id)})
    with urllib.request.urlopen(req, timeout=15) as resp:
        data = json.loads(resp.read())
    if isinstance(data, list): data = {"moves": data}
    raw = data.get("moves") or data.get("matchMoves") or data.get("playByPlay") or []
    if not raw:
        for v in data.values():
            if isinstance(v,list) and len(v)>3: raw=v; break
    rows = []
    for i,play in enumerate(raw):
        if not isinstance(play,dict): continue
        mn=play.get("min",""); sc=play.get("sec","")
        temps=f"{int(mn):02d}:{int(sc):02d}" if mn!="" and sc!="" else str(mn)
        move=play.get("move","")
        punts=3 if "Cistella de 3" in move else (2 if "Cistella de 2" in move else (1 if ("Cistella de 1" in move or "Tir lliure convertit" in move) else 0))
        rows.append({"num":i+1,"quart":play.get("period",""),"temps":temps,
            "min_num":float(mn)+float(sc)/60 if mn!="" else 0,
            "idEquip":str(play.get("idTeam","")),"dorsal":play.get("actorShirtNumber",""),
            "jugador":play.get("actorName",""),"accio":move,
            "marcador":play.get("score",""),"punts":punts,
            "teamAction":play.get("teamAction",False)})
    return pd.DataFrame(rows)

# ══════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════
def get_teams(df):
    return [t for t in df["idEquip"].unique() if t and t!="0"]

def get_teams_ordered(df):
    """Retorna [equip_local, equip_visitant] usant el marcador per detectar qui és qui.
    El primer número del marcador API és sempre el local."""
    teams = get_teams(df)
    if len(teams) < 2:
        return teams

    # Agafa l'últim marcador vàlid
    last_score = None
    for _, r in df.iloc[::-1].iterrows():
        marc = str(r.get("marcador",""))
        if "-" in marc:
            try:
                sa, sb = int(marc.split("-")[0]), int(marc.split("-")[1])
                last_score = (sa, sb)
                break
            except: pass

    if last_score is None:
        return teams

    # Calcula punts totals per equip
    pts = {}
    for tid in teams:
        pts[tid] = int(df[df["idEquip"]==tid]["punts"].sum())

    sa, sb = last_score
    # L'equip local és el que té pts més propers a sa
    t0, t1 = teams[0], teams[1]
    if abs(pts.get(t0,0) - sa) <= abs(pts.get(t1,0) - sa):
        return [t0, t1]  # t0 és local
    else:
        return [t1, t0]  # t1 és local

def score_evo(df):
    rows=[]
    for _,r in df.iterrows():
        marc=str(r["marcador"])
        if "-" in marc:
            try:
                sa,sb=int(marc.split("-")[0]),int(marc.split("-")[1])
                rows.append({"num":r["num"],"quart":r["quart"],"temps":r["temps"],"scoreA":sa,"scoreB":sb,"diff":sa-sb})
            except: pass
    return pd.DataFrame(rows)

def final_score(sdf):
    if sdf.empty: return 0, 0
    last=sdf.iloc[-1]
    try: return int(last["scoreA"]), int(last["scoreB"])
    except: return 0, 0

def estat_marc(row, teams):
    marc=str(row.get("marcador",""))
    if "-" not in marc: return "desconegut"
    try:
        sa,sb=int(marc.split("-")[0]),int(marc.split("-")[1])
        diff=(sa-sb) if row["idEquip"]==teams[0] else (sb-sa)
        return "Guanyant" if diff>0 else ("Empatat" if diff==0 else "Perdent")
    except: return "desconegut"

def get_shot_counts(df_sub):
    v1m=int(df_sub["accio"].str.contains("Cistella de 1|Tir lliure convertit",case=False,na=False).sum())
    v1x=int(df_sub["accio"].str.contains("Intent fallat de 1",case=False,na=False).sum())
    v2m=int(df_sub["accio"].str.contains("Cistella de 2",case=False,na=False).sum())
    v2x=int(df_sub["accio"].str.contains("Intent fallat de 2|fall.*2|2.*fall",case=False,na=False).sum())
    v3m=int(df_sub["accio"].str.contains("Cistella de 3",case=False,na=False).sum())
    v3x=int(df_sub["accio"].str.contains("Intent fallat de 3|fall.*3|3.*fall",case=False,na=False).sum())
    return v1m,v1x,v2m,v2x,v3m,v3x

# ══════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""<div style="display:flex;align-items:center;gap:10px;padding-bottom:14px;border-bottom:0.5px solid #e2e4e8;margin-bottom:14px">
        <div style="width:34px;height:34px;background:#E6F1FB;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:18px">🏀</div>
        <div><div style="font-size:13px;font-weight:600;color:#1a1c22">Micki Analítica</div>
        <div style="font-size:11px;color:#9ca3af">Analítica de Bàsquet</div></div></div>""", unsafe_allow_html=True)

    st.markdown('<div style="font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.08em;color:#9ca3af;margin-bottom:6px">Partit</div>', unsafe_allow_html=True)
    url_input = st.text_input("", placeholder="URL o ID del partit", label_visibility="collapsed")
    nom_equip_1 = st.text_input("Equip local", placeholder="Ex: Manresa", key="nom_eq1")
    nom_equip_2 = st.text_input("Equip visitant", placeholder="Ex: Girona", key="nom_eq2")
    carregar = st.button("⬇ Carregar partit", use_container_width=True)

    # ── Càrrega múltiple ─────────────────────────────────────────────────
    with st.expander("📋 Carregar múltiples partits", expanded=False):
        st.caption("Enganxa una URL o ID per línia. S'intentaran carregar tots.")
        urls_multi = st.text_area("URLs / IDs (un per línia)", height=120,
                                   placeholder="69ec95d4339c3d0001f523a1\n6a1c25041cc34c000132763e\nhttps://www.basquetcatala.cat/...")
        carregar_multi = st.button("⬇ Carregar tots", use_container_width=True, key="btn_multi")
        if carregar_multi and urls_multi.strip():
            linies = [l.strip() for l in urls_multi.strip().split("\n") if l.strip()]
            ok = 0; errors = []
            progress = st.progress(0)
            for i, linia in enumerate(linies):
                # Extreu l'ID de la URL si cal
                mid_multi = linia.split("/")[-1].strip() if "/" in linia else linia.strip()
                mid_multi = mid_multi.split("?")[0].strip()
                try:
                    df_m, teams_m, team_names_m = fetch_and_parse(mid_multi)
                    if df_m is not None and not df_m.empty:
                        ts_m = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        save_jugades(mid_multi, ts_m, df_m)
                        save_stats_jugador(mid_multi, ts_m, df_m, teams_m, team_names_m)
                        save_shots_zones(mid_multi, ts_m, df_m, teams_m)
                        save_timeouts(mid_multi, ts_m, df_m, teams_m)
                        n_a = team_names_m.get(teams_m[0],"?") if teams_m else "?"
                        n_b = team_names_m.get(teams_m[1],"?") if len(teams_m)>1 else "?"
                        sa = int(df_m[df_m["idEquip"]==teams_m[0]]["punts"].sum()) if teams_m else 0
                        sb = int(df_m[df_m["idEquip"]==teams_m[1]]["punts"].sum()) if len(teams_m)>1 else 0
                        save_partit(mid_multi, ts_m, n_a, n_b, sa, sb)
                        ok += 1
                    else:
                        errors.append(f"No trobat: {mid_multi[:20]}")
                except Exception as ex:
                    errors.append(f"Error {mid_multi[:20]}: {str(ex)[:30]}")
                progress.progress((i+1)/len(linies))
            if ok > 0:
                st.success(f"✅ {ok}/{len(linies)} partits carregats!")
            if errors:
                st.warning("\n".join(errors))
    st.markdown("---")
    st.markdown('<div style="font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.08em;color:#9ca3af;margin-bottom:6px">Filtres play-by-play</div>', unsafe_allow_html=True)
    quart_sel = st.multiselect("Quart", options=[1,2,3,4], default=[1,2,3,4])
    accio_cerca = st.text_input("Acció", placeholder="Cistella, falta...")
    jugador_cerca = st.text_input("Jugadora", placeholder="Nom...")
    st.markdown("---")
    st.caption("Micki Analítica")

# ── Sessió ─────────────────────────────────────────────────────────────────────
for k,v in [("df",None),("match_id",None),("team_names",{}),("score_a",0),("score_b",0)]:
    if k not in st.session_state: st.session_state[k]=v

# ── Càrrega ────────────────────────────────────────────────────────────────────
if carregar and url_input:
    mid = extract_match_id(url_input)
    if not mid:
        st.error("ID no vàlid.")
    elif partit_exists(mid):
        st.session_state.df = load_jugades_db(mid)
        st.session_state.match_id = mid
        df_part = load_partits_db()
        row = df_part[df_part["match_id"]==mid]
        if not row.empty:
            t_list = get_teams(st.session_state.df)
            noms = {}
            if len(t_list)>=1: noms[t_list[0]] = row.iloc[0]["nom_a"]
            if len(t_list)>=2: noms[t_list[1]] = row.iloc[0]["nom_b"]
            st.session_state.team_names = noms
            # Guardem els scores de la BD per mostrar-los correctament
            st.session_state.score_a = int(row.iloc[0]["score_a"])
            st.session_state.score_b = int(row.iloc[0]["score_b"])
        st.success("Carregat des de l'històric ⚡")
    else:
        with st.spinner("Descarregant de l'API..."):
            try:
                df = fetch_and_parse(mid)
                teams_tmp = get_teams_ordered(df)
                noms_guardats = load_noms_equips()
                def get_nom(i, tid, input_nom):
                    if input_nom and input_nom.strip():
                        save_nom_equip(tid, input_nom.strip()); return input_nom.strip()
                    if tid in noms_guardats: return noms_guardats[tid]
                    return f"Equip {chr(65+i)}"
                noms = {}
                inputs = [nom_equip_1, nom_equip_2]
                for i,tid in enumerate(teams_tmp[:2]):
                    noms[tid] = get_nom(i, tid, inputs[i] if i<len(inputs) else "")
                id_a = teams_tmp[0] if teams_tmp else ""
                id_b = teams_tmp[1] if len(teams_tmp)>1 else ""
                sdf = score_evo(df); fa,fb = final_score(sdf)
                ts = datetime.now().strftime("%Y-%m-%d %H:%M")
                save_partit(mid, df, noms.get(id_a,"A"), noms.get(id_b,"B"), id_a, id_b, fa, fb)
                save_stats_jugador(mid, ts, df, teams_tmp, noms)
                save_shots_zones(mid, ts, df, noms)
                save_timeouts(mid, ts, df, noms)
                st.session_state.df = df
                st.session_state.match_id = mid
                st.session_state.team_names = noms
                st.session_state.score_a = fa
                st.session_state.score_b = fb
                st.success("Carregat i desat ✅")
            except Exception as e:
                st.error(f"Error: {e}")

# ── Pantalla inicial ────────────────────────────────────────────────────────────
if st.session_state.df is None:
    st.markdown("""<div style="text-align:center;padding:80px 0">
        <div style="font-size:64px">🏀</div>
        <h1 style="font-size:38px;font-weight:600;color:#1a1c22;margin:16px 0 8px">Micki Analítica</h1>
        <p style="color:#6b7280;font-size:15px">Enganxa la URL o l'ID d'un partit al panell esquerre i prem Carregar.</p>
        <p style="color:#d1d5db;font-size:12px;margin-top:32px">Exemple: 69ec95d4339c3d0001f523a1</p>
    </div>""", unsafe_allow_html=True)
    st.stop()

# ── Variables globals del partit carregat ─────────────────────────────────────
df_orig = st.session_state.df.copy()
match_id = st.session_state.match_id
teams = get_teams_ordered(df_orig)
team_names = st.session_state.team_names
nom_a = team_names.get(teams[0],"Equip A") if teams else "Equip A"
nom_b = team_names.get(teams[1],"Equip B") if len(teams)>1 else "Equip B"
color_map_eq = {nom_a:COLOR_A, nom_b:COLOR_B}
df_orig["equip_nom"] = df_orig["idEquip"].map(team_names).fillna("?")
score_df = score_evo(df_orig)
# Usa els scores guardats si existeixen, sinó calcula del marcador
if st.session_state.score_a or st.session_state.score_b:
    fa, fb = st.session_state.score_a, st.session_state.score_b
else:
    fa, fb = final_score(score_df)

# ══════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════
def get_intervals_jugadores_global(df):
    """Retorna dict jugadora -> [(t_ini, t_fi, equip_id)] en minuts absoluts de partit."""
    MINS_Q = 10
    intervals = {}
    en_pista  = {}
    col_j = "jugador" if "jugador" in df.columns else "jugadora"

    for jug_x in df[col_j].unique():
        if not jug_x or str(jug_x) in ("","nan"): continue
        df_jx = df[df[col_j]==jug_x].sort_values("num")
        if df_jx.empty: continue
        primer_acc = str(df_jx.iloc[0].get("accio",""))
        primer_q   = int(df_jx.iloc[0].get("quart",1))
        primer_ei  = str(df_jx.iloc[0].get("idEquip",""))
        if "Surt" in primer_acc and "camp" in primer_acc:
            en_pista[jug_x] = ((primer_q-1)*MINS_Q, primer_ei)

    for _, row in df.sort_values("num").iterrows():
        jug = row.get(col_j, "")
        if not jug or str(jug) in ("","nan"): continue
        accio = str(row.get("accio",""))
        quart = int(row.get("quart",1)) if row.get("quart","") != "" else 1
        min_dins = float(row.get("min_num", 0))
        if min_dins > MINS_Q:
            t_min = min_dins
        else:
            t_min = (quart-1)*MINS_Q + (MINS_Q - min_dins)
        t_min = max(0, min(t_min, quart * MINS_Q))
        eq_id = str(row.get("idEquip",""))
        if "Entra al camp" in accio:
            en_pista[jug] = (t_min, eq_id)
        elif "Surt del camp" in accio:
            ini_t = en_pista.pop(jug, ((quart-1)*MINS_Q, eq_id))
            t_ini_real = ini_t[0]; eq_real = ini_t[1]
            if t_min > t_ini_real:
                intervals.setdefault(jug, []).append((t_ini_real, t_min, eq_real))
        elif "Final de període" in accio:
            fi = quart * MINS_Q
            for j, (ti, ei) in list(en_pista.items()):
                if fi > ti:
                    intervals.setdefault(j, []).append((ti, fi, ei))
            en_pista = {}
    for j, (ti, ei) in en_pista.items():
        fi = df["quart"].max() * MINS_Q if not df.empty else 40
        if fi > ti:
            intervals.setdefault(j, []).append((ti, fi, ei))
    return intervals


def calc_pm_combinacions(df_orig, mode="quintets"):
    """Calcula +/- per combinacions de jugadores (quintets o parelles).
    mode: 'quintets' o 'parelles'
    Retorna llista de dicts: {combinacio, equip, minuts, pf, pc, pm, pm_min}
    """
    import itertools
    MINS_Q = 10
    col_j = "jugador" if "jugador" in df_orig.columns else "jugadora"
    df_orig = df_orig.copy()
    df_orig["jugador"] = df_orig[col_j].fillna("")
    intervals_jug = get_intervals_jugadores_global(df_orig)

    # Construeix línia temporal d'esdeveniments (canvis de pista) per equip
    teams = get_teams_ordered(df_orig)
    if len(teams) < 2: return []

    # Per cada equip, construeix una llista de "punts de canvi" (timestamps)
    results = {}  # key: (equip, frozenset(jugadores)) -> [minuts, pf, pc]

    df_t = df_orig.copy()
    df_t["t_abs"] = df_t.apply(
        lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
        if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)

    for eq_id in teams:
        rival_id = [t for t in teams if t != eq_id][0]
        # Jugadores d'aquest equip amb intervals
        jugs_eq = [j for j,ivs in intervals_jug.items()
                   if ivs and str(ivs[0][2]) == str(eq_id)]
        if not jugs_eq: continue

        # Recull tots els punts de canvi (inicis i finals d'intervals)
        canvis = set([0.0])
        max_t = float(df_orig["quart"].max() * MINS_Q) if not df_orig.empty else 40.0
        canvis.add(max_t)
        for j in jugs_eq:
            for ti,tf,_ in intervals_jug[j]:
                canvis.add(round(ti,2)); canvis.add(round(tf,2))
        canvis = sorted(canvis)

        # Per cada microinterval entre canvis consecutius, determina qui és a pista
        for i in range(len(canvis)-1):
            t0, t1 = canvis[i], canvis[i+1]
            if t1 - t0 < 0.01: continue
            tm = (t0+t1)/2
            en_pista_ara = []
            for j in jugs_eq:
                for ti,tf,_ in intervals_jug[j]:
                    if ti <= tm < tf:
                        en_pista_ara.append(j)
                        break

            if mode == "quintets":
                if len(en_pista_ara) != 5: continue
                combos = [tuple(sorted(en_pista_ara))]
            else:  # parelles
                if len(en_pista_ara) < 2: continue
                combos = [tuple(sorted(c)) for c in itertools.combinations(en_pista_ara, 2)]

            # Punts en aquest microinterval
            df_i = df_t[(df_t["t_abs"]>=t0)&(df_t["t_abs"]<t1)]
            pf_i = int(df_i[df_i["idEquip"]==eq_id]["punts"].sum())
            pc_i = int(df_i[df_i["idEquip"]==rival_id]["punts"].sum())
            durada = t1-t0

            for combo in combos:
                key = (str(eq_id), combo)
                if key not in results:
                    results[key] = [0.0, 0, 0]
                results[key][0] += durada
                results[key][1] += pf_i
                results[key][2] += pc_i

    rows = []
    for (eq_id, combo), (mins_c, pf_c, pc_c) in results.items():
        if mins_c < 0.5: continue
        rows.append({
            "equip": eq_id,
            "combinacio": combo,
            "minuts": round(mins_c,1),
            "pf": pf_c,
            "pc": pc_c,
            "pm": pf_c - pc_c,
            "pm_min": round((pf_c-pc_c)/mins_c, 3) if mins_c>0 else 0
        })
    return rows


def calc_possessions(df_equip):
    """Calcula les possessions estimades d'un equip."""
    tc_int = int(df_equip["accio"].str.contains(
        "Cistella de 2|Cistella de 3|Intent fallat de 2|Intent fallat de 3|"
        "Tir de 2|Tir de 3|fallat de 2|fallat de 3",
        case=False, na=False).sum())
    tl_int = int(df_equip["accio"].str.contains(
        "Cistella de 1|Intent fallat de 1", case=False, na=False).sum())
    return tc_int + 0.44 * tl_int

def calc_eficiencies(df_orig, teams, team_names):
    """Calcula eficiència ofensiva i defensiva per equip."""
    result = {}
    for i, tid in enumerate(teams[:2]):
        rival_id = teams[1-i] if len(teams) > 1 else None
        df_eq  = df_orig[df_orig["idEquip"] == tid]
        df_riv = df_orig[df_orig["idEquip"] == rival_id] if rival_id else pd.DataFrame()
        pts_of = int(df_eq["punts"].sum())
        pts_def = int(df_riv["punts"].sum()) if not df_riv.empty else 0
        poss_of  = calc_possessions(df_eq)
        poss_def = calc_possessions(df_riv) if not df_riv.empty else 1
        off_rtg = round(pts_of  / poss_of  * 100, 1) if poss_of  > 0 else 0
        def_rtg = round(pts_def / poss_def * 100, 1) if poss_def > 0 else 0
        net_rtg = round(off_rtg - def_rtg, 1)
        result[tid] = {
            "nom": team_names.get(tid, "?"),
            "pts_of": pts_of, "pts_def": pts_def,
            "poss_of": round(poss_of, 1), "poss_def": round(poss_def, 1),
            "off_rtg": off_rtg, "def_rtg": def_rtg, "net_rtg": net_rtg
        }
    return result

def calc_onoff(df_orig, jugadora, equip_id, teams):
    """Calcula On/Off Rating d'una jugadora usant intervals reals."""
    rival_id = next((t for t in teams if t != equip_id), None)
    if rival_id is None: return None

    MINS_Q = 10
    MIN_POSS = 4  # mínim de possessions per considerar el rating vàlid

    col_j = "jugador" if "jugador" in df_orig.columns else "jugadora"

    # Calcula intervals reals de la jugadora
    intervals_on = []; en_pista = {}
    df_jug_rows = df_orig[df_orig[col_j]==jugadora]
    if df_jug_rows.empty: return None

    # Detecta si comença a pista
    primer = df_jug_rows.sort_values("num").iloc[0]
    if "Surt" in str(primer.get("accio","")) and "camp" in str(primer.get("accio","")):
        en_pista[jugadora] = (int(primer.get("quart",1))-1)*MINS_Q

    for _,row in df_orig.sort_values("num").iterrows():
        if str(row.get(col_j,"")) != str(jugadora): continue
        acc = str(row.get("accio",""))
        q = int(row.get("quart",1))
        m = float(row.get("min_num",0))
        t = (q-1)*MINS_Q + (MINS_Q-m if m<=MINS_Q else m)
        t = max(0, min(t, q*MINS_Q))
        if "Entra" in acc and "camp" in acc:
            en_pista[jugadora] = t
        elif "Surt" in acc and "camp" in acc:
            ti = en_pista.pop(jugadora, (q-1)*MINS_Q)
            if t > ti: intervals_on.append((float(ti), float(t)))
        elif "Final de període" in acc:
            if jugadora in en_pista:
                ti = en_pista.pop(jugadora)
                fi = float(q*MINS_Q)
                if fi > ti: intervals_on.append((float(ti), fi))
    for ti_o in en_pista.values():
        fi_o = float(df_orig["quart"].max()*MINS_Q)
        if fi_o > ti_o: intervals_on.append((float(ti_o), fi_o))

    if not intervals_on: return None

    # Precalcula t_abs
    df_t = df_orig.copy()
    df_t["t_abs"] = df_t.apply(
        lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
        if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)

    # Construeix màscara ON (intervals reals)
    mask_on = df_t["t_abs"].apply(
        lambda t: any(ti<=t<=tf for ti,tf in intervals_on))

    df_on_eq  = df_t[mask_on  & (df_t["idEquip"]==equip_id)]
    df_on_riv = df_t[mask_on  & (df_t["idEquip"]==rival_id)]
    df_off_eq = df_t[~mask_on & (df_t["idEquip"]==equip_id)]
    df_off_riv= df_t[~mask_on & (df_t["idEquip"]==rival_id)]

    def rtg(df_e, df_r):
        pts  = int(df_e["punts"].sum())
        poss = calc_possessions(df_e)
        pts_r  = int(df_r["punts"].sum())
        poss_r = calc_possessions(df_r)
        # Si poques possessions, rating no fiable
        if poss < MIN_POSS: return None, None, None
        off  = round(pts/poss*100, 1)
        deff = round(pts_r/poss_r*100, 1) if poss_r >= MIN_POSS else None
        net  = round(off-deff, 1) if deff is not None else None
        return off, deff, net

    on_off,  on_def,  on_net  = rtg(df_on_eq,  df_on_riv)
    off_off, off_def, off_net = rtg(df_off_eq, df_off_riv)

    if on_net is None or off_net is None:
        diff = None
    else:
        diff = round(on_net - off_net, 1)
        # Cap de valors poc raonables (>50 punts/100 poss és sospitós)
        if abs(diff) > 50 and not (
            (oo2.get("on_poss",0) if False else calc_possessions(df_on_eq)) >= 8 and
            calc_possessions(df_off_eq) >= 8):
            diff = None  # marca com no fiable

    return {
        "on_off_rtg":  on_off,  "on_def_rtg":  on_def,  "on_net_rtg":  on_net,
        "off_off_rtg": off_off, "off_def_rtg": off_def, "off_net_rtg": off_net,
        "diff": diff,
        "on_poss":  round(calc_possessions(df_on_eq), 1),
        "off_poss": round(calc_possessions(df_off_eq), 1),
    }

def calc_metriques_partit(df_jug, match_id, nom_equip, nom_rival):
    """Calcula totes les mètriques avançades d'un equip en un partit."""
    pts_tot = int(df_jug["punts"].sum())
    pts_2   = int(df_jug["accio"].str.contains("Cistella de 2",case=False,na=False).sum()) * 2
    pts_3   = int(df_jug["accio"].str.contains("Cistella de 3",case=False,na=False).sum()) * 3
    pts_tl  = int(df_jug["accio"].str.contains("Cistella de 1",case=False,na=False).sum())
    tc_conv = int(df_jug["accio"].str.contains("Cistella de 2|Cistella de 3",case=False,na=False).sum())
    tc_fall = int(df_jug["accio"].str.contains("Intent fallat de 2|Intent fallat de 3|fallat de 2|fallat de 3",case=False,na=False).sum())
    tc_int  = tc_conv + tc_fall
    tl_conv = pts_tl
    tl_int  = tl_conv + int(df_jug["accio"].str.contains("Intent fallat de 1",case=False,na=False).sum())
    c3_conv = int(df_jug["accio"].str.contains("Cistella de 3",case=False,na=False).sum())
    c3_int  = c3_conv + int(df_jug["accio"].str.contains("Intent fallat de 3|fallat de 3",case=False,na=False).sum())
    c2_conv = int(df_jug["accio"].str.contains("Cistella de 2",case=False,na=False).sum())
    c2_int  = c2_conv + int(df_jug["accio"].str.contains("Intent fallat de 2|fallat de 2",case=False,na=False).sum())

    poss     = calc_possessions(df_jug)
    ts_denom = 2 * (tc_int + 0.44 * tl_int)
    return {
        "Equip":        nom_equip,
        "Rival":        nom_rival,
        "Pts":          pts_tot,
        "Pts 2pts":     pts_2,
        "Pts 3pts":     pts_3,
        "Pts TL":       pts_tl,
        "Possessions":  round(poss, 1),
        "Pts/Poss":     round(pts_tot / poss, 3) if poss > 0 else 0,
        "Off Rtg":      round(pts_tot / poss * 100, 1) if poss > 0 else 0,
        "TS%":          round(pts_tot / ts_denom * 100, 1) if ts_denom > 0 else 0,
        "TC%":          round(tc_conv / tc_int * 100, 1) if tc_int > 0 else 0,
        "2pts%":        round(c2_conv / c2_int * 100, 1) if c2_int > 0 else 0,
        "3pts%":        round(c3_conv / c3_int * 100, 1) if c3_int > 0 else 0,
        "TL%":          round(tl_conv / tl_int * 100, 1) if tl_int > 0 else 0,
        "%Pts 2pts":    round(pts_2 / pts_tot * 100, 1) if pts_tot > 0 else 0,
        "%Pts 3pts":    round(pts_3 / pts_tot * 100, 1) if pts_tot > 0 else 0,
        "%Pts TL":      round(pts_tl / pts_tot * 100, 1) if pts_tot > 0 else 0,
        "2pts conv/int": f"{c2_conv}/{c2_int}",
        "3pts conv/int": f"{c3_conv}/{c3_int}",
        "TL conv/int":   f"{tl_conv}/{tl_int}",
        # Detall complet per cistella
        "1pt conv":     tl_conv,
        "1pt int":      tl_int,
        "1pt%":         round(tl_conv / tl_int * 100, 1) if tl_int > 0 else 0,
        "2pts conv":    c2_conv,
        "2pts int":     c2_int,
        "2pts% ef":     round(c2_conv / c2_int * 100, 1) if c2_int > 0 else 0,
        "3pts conv":    c3_conv,
        "3pts int":     c3_int,
        "3pts% ef":     round(c3_conv / c3_int * 100, 1) if c3_int > 0 else 0,
    }

def genera_excel_analisi():
    """Genera Excel amb mètriques avançades de tots els partits de la BD."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    BLAU_FOSC='0C447C'; BLAU_MIG='185FA5'; BLAU_CLAR='EBF4FC'
    VERD='D5F5E3'; VERMELL='FADBD8'; GROC='FFF3CD'
    GRIS='F2F4F6'; BLANC='FFFFFF'

    def fons(c): return PatternFill('solid', fgColor=c)
    def vora():
        s=Side(style='thin',color='CCCCCC')
        return Border(top=s,bottom=s,left=s,right=s)
    def fc(ws,row,col,value,bold=False,bg=None,fg='000000',align='center',size=10,num_fmt=None):
        c=ws.cell(row=row,column=col,value=value)
        c.font=Font(name='Arial',bold=bold,color=fg,size=size)
        if bg: c.fill=fons(bg)
        c.alignment=Alignment(horizontal=align,vertical='center')
        c.border=vora()
        if num_fmt: c.number_format=num_fmt
        return c

    df_p = load_partits_db()
    if df_p.empty:
        return None

    wb = Workbook(); wb.remove(wb.active)

    # ── PESTANYA 1: TOTS ELS PARTITS ──────────────────────────────────
    ws1 = wb.create_sheet("📊 Partits")
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 2

    ws1.merge_cells('B1:Z1')
    c=ws1['B1']; c.value='🏀  MICKI ANALÍTICA — ANÀLISI DE PARTITS'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=14)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws1.row_dimensions[1].height=36

    ws1.merge_cells('B2:Z2')
    c=ws1['B2']; c.value=f"Generat: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  {len(df_p)} partits"
    c.font=Font(name='Arial',color=BLANC,size=10); c.fill=fons(BLAU_MIG)
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws1.row_dimensions[2].height=20; ws1.row_dimensions[3].height=6

    # Capçaleres per grups
    grups = [
        ("IDENTIFICACIÓ",2,4,'0C447C'),
        ("RESULTAT",5,6,'185FA5'),
        ("POSSESSIONS",7,9,'0F6E56'),
        ("EFICIÈNCIA DE TIR",10,14,'3B6D11'),
        ("DISTRIBUCIÓ PUNTS",15,17,'854F0B'),
        ("DETALL TIRS",18,20,'533800' if False else '993C1D'),
    ]
    row=4
    for grup,c_ini,c_fi,color in grups:
        ws1.merge_cells(f'{get_column_letter(c_ini)}{row}:{get_column_letter(c_fi)}{row}')
        c=ws1.cell(row=row,column=c_ini,value=grup)
        c.font=Font(name='Arial',bold=True,color=BLANC,size=9)
        c.fill=fons(color); c.alignment=Alignment(horizontal='center',vertical='center')
        c.border=vora()
    ws1.row_dimensions[row].height=16; row+=1

    caps=[('Data',11),('Equip',20),('Rival',20),
          ('Pts',8),('Pts rival',9),
          ('Poss.',9),('Pts/Poss',10),('Off Rtg',9),
          ('TS%',8),('TC%',8),('2pts%',8),('3pts%',8),('TL%',8),
          ('%Pts 2',9),('%Pts 3',9),('%Pts TL',9),
          ('1pt conv',9),('1pt int',9),('1pt%',8),
          ('2pts conv',9),('2pts int',9),('2pts%',8),
          ('3pts conv',9),('3pts int',9),('3pts%',8)]
    for ci,(cap,w) in enumerate(caps,2):
        fc(ws1,row,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,size=9)
        ws1.column_dimensions[get_column_letter(ci)].width=w
    ws1.row_dimensions[row].height=20; row+=1

    # Dades de tots els partits
    all_rows = []
    for _,p in df_p.iterrows():
        df_jug = load_jugades_db(p['match_id'])
        if df_jug.empty: continue
        teams_p = get_teams_ordered(df_jug)
        tn_p = {}
        if len(teams_p)>=1: tn_p[teams_p[0]]=p['nom_a']
        if len(teams_p)>=2: tn_p[teams_p[1]]=p['nom_b']

        for i,tid in enumerate(teams_p[:2]):
            rival_id = teams_p[1-i] if len(teams_p)>1 else None
            df_eq = df_jug[df_jug['idEquip']==tid]
            nom_eq = tn_p.get(tid,'?')
            nom_riv = tn_p.get(rival_id,'?') if rival_id else '?'
            pts_rival = int(df_jug[df_jug['idEquip']==rival_id]['punts'].sum()) if rival_id else 0
            met = calc_metriques_partit(df_eq, p['match_id'], nom_eq, nom_riv)
            met['Data'] = str(p['data_consulta'])[:10]
            met['Pts rival'] = pts_rival
            all_rows.append(met)

    for i,r in enumerate(all_rows):
        bg = BLAU_CLAR if i%2==0 else BLANC
        data_vals = [
            r['Data'], r['Equip'], r['Rival'],
            r['Pts'], r['Pts rival'],
            r['Possessions'], r['Pts/Poss'], r['Off Rtg'],
            r['TS%'], r['TC%'], r['2pts%'], r['3pts%'], r['TL%'],
            r['%Pts 2pts'], r['%Pts 3pts'], r['%Pts TL'],
            r['1pt conv'], r['1pt int'], r['1pt%'],
            r['2pts conv'], r['2pts int'], r['2pts% ef'],
            r['3pts conv'], r['3pts int'], r['3pts% ef'],
        ]
        num_fmts = [None,None,None,None,None,
                    '0.0','0.000','0.0',
                    '0.0%' if False else '0.0','0.0','0.0','0.0','0.0',
                    '0.0','0.0','0.0',
                    None,None,None]
        for ci,val in enumerate(data_vals,2):
            fc(ws1,row,ci,val,bg=bg,
               align='left' if ci<=4 else 'center',
               size=9)
        ws1.row_dimensions[row].height=17; row+=1

    # Fila de mitjanes
    row+=1
    fc(ws1,row,2,'MITJANA',bold=True,bg=GROC,fg=BLAU_FOSC,align='left',size=10)
    fc(ws1,row,3,'',bg=GROC); fc(ws1,row,4,'',bg=GROC)
    data_ini = row - len(all_rows) - 1
    data_fi  = row - 2
    for ci in range(5,21):
        col_l = get_column_letter(ci)
        c2=ws1.cell(row=row,column=ci)
        c2.value=f'=IFERROR(AVERAGE({col_l}{data_ini}:{col_l}{data_fi}),"")'
        c2.font=Font(name='Arial',bold=True,color=BLAU_FOSC,size=10)
        c2.fill=fons(GROC); c2.alignment=Alignment(horizontal='center',vertical='center')
        c2.border=vora(); c2.number_format='0.0' if ci<=17 else '0.0'
    ws1.row_dimensions[row].height=22

    # ── PESTANYA 2: JUGADORES ─────────────────────────────────────────
    ws2 = wb.create_sheet("👤 Jugadores")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions['A'].width=2
    ws2.merge_cells('B1:N1')
    c=ws2['B1']; c.value='🏀  MICKI ANALÍTICA — RÀNQUING DE JUGADORES'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=14)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws2.row_dimensions[1].height=36
    ws2.row_dimensions[2].height=8
    row2=3
    for ci,cap,w in zip(range(2,16),
        ['#','Jugadora','Equip','Part.','Pts','Pts/P','Min','Min/P','C2','C3','TL','Faltes','Impacte','Usage%'],
        [5,24,20,8,9,9,9,9,7,7,7,9,12,10]):
        fc(ws2,row2,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,size=10)
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.row_dimensions[row2].height=22; row2+=1

    df_s=load_stats_jugador_db()
    if not df_s.empty:
        agg_dict2=dict(p=('match_id','nunique'),pts=('punts','sum'),
            c2=('cistelles_2','sum'),c3=('cistelles_3','sum'),
            tl=('tirs_lliures','sum'),f=('faltes','sum'),imp=('impacte','sum'))
        if 'minuts' in df_s.columns: agg_dict2['mn']=('minuts','sum')
        else: agg_dict2['mn']=('punts','count')
        if 'usage_rate' in df_s.columns: agg_dict2['usage']=('usage_rate','mean')
        agg=df_s.groupby(['jugador','equip_nom']).agg(**agg_dict2).reset_index().sort_values('pts',ascending=False)
        for rank,(_,r) in enumerate(agg.iterrows(),1):
            bg=BLAU_CLAR if rank%2==0 else BLANC
            ppp=round(r['pts']/r['p'],1) if r['p']>0 else 0
            mpp=round(r['mn']/r['p'],1) if r['p']>0 else 0
            iv=f"+{int(r['imp'])}" if r['imp']>=0 else str(int(r['imp']))
            ic='0F6E56' if r['imp']>=0 else '993C1D'
            fc(ws2,row2,2,rank,align='center',bg=bg,bold=True,fg=BLAU_FOSC)
            fc(ws2,row2,3,r['jugador'],bold=True,bg=bg,fg=BLAU_FOSC,align='left')
            fc(ws2,row2,4,r['equip_nom'],bg=bg,align='left')
            fc(ws2,row2,5,int(r['p']),bg=bg)
            fc(ws2,row2,6,int(r['pts']),bold=True,bg=bg,fg=BLAU_FOSC)
            fc(ws2,row2,7,ppp,bg=bg)
            fc(ws2,row2,8,round(r['mn'],1),bg=bg)
            fc(ws2,row2,9,mpp,bg=bg)
            fc(ws2,row2,10,int(r['c2']),bg=bg)
            fc(ws2,row2,11,int(r['c3']),bg=bg)
            fc(ws2,row2,12,int(r['tl']),bg=bg)
            fc(ws2,row2,13,int(r['f']),bg=bg)
            fc(ws2,row2,14,iv,bold=True,bg=bg,fg=ic)
            # Usage% com a número (0.75 = 75%) per poder calcular mitjanes a Excel
            usage_num = round(r.get('usage',0)/100, 4) if r.get('usage',0) > 0 else None
            uc2 = '0F6E56' if (usage_num or 0)>=0.25 else ('854F0B' if (usage_num or 0)>=0.15 else '374151')
            c_us = ws2.cell(row=row2, column=15, value=usage_num)
            c_us.font = Font(name='Arial', bold=True, color=uc2, size=10)
            c_us.fill = fons(bg)
            c_us.alignment = Alignment(horizontal='center', vertical='center')
            c_us.border = vora()
            c_us.number_format = '0%'
            ws2.row_dimensions[row2].height=18; row2+=1

    # ── PESTANYA 3: EFICIÈNCIA DE TIR ─────────────────────────────────
    ws3=wb.create_sheet("🎯 Eficiència tir")
    ws3.sheet_view.showGridLines=False; ws3.column_dimensions['A'].width=2
    ws3.merge_cells('B1:L1')
    c=ws3['B1']; c.value='🏀  MICKI ANALÍTICA — EFICIÈNCIA DE TIR'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=14)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws3.row_dimensions[1].height=36; ws3.row_dimensions[2].height=8
    row3=3
    for ci,cap,w in zip(range(2,13),
        ['Jugadora','Equip','TL conv/int','TL%','2pts conv/int','2pts%','3pts conv/int','3pts%','Total conv/int','Total%'],
        [24,20,14,9,16,9,16,9,16,9]):
        fc(ws3,row3,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,size=10)
        ws3.column_dimensions[get_column_letter(ci)].width=w
    ws3.row_dimensions[row3].height=22; row3+=1

    df_sh=load_shots_zones_db()
    if not df_sh.empty:
        def ef_bg(m,t):
            if t==0: return BLANC
            p=m/t
            return 'D5F5E3' if p>=0.55 else (GROC if p>=0.35 else 'FADBD8')
        at=df_sh[df_sh['jugador']!='__equip__'].groupby(['jugador','equip_nom']).agg(
            v1m=('val1_made','sum'),v1x=('val1_miss','sum'),
            v2m=('val2_made','sum'),v2x=('val2_miss','sum'),
            v3m=('val3_made','sum'),v3x=('val3_miss','sum'),
        ).reset_index()
        at['tm']=at['v1m']+at['v2m']+at['v3m']
        at=at.sort_values('tm',ascending=False)
        for _,r in at.iterrows():
            def ratio(m,x): return f"{int(m)}/{int(m+x)}" if (m+x)>0 else "—"
            def pct(m,x): return f"{round(m/(m+x)*100)}%" if (m+x)>0 else "—"
            fc(ws3,row3,2,r['jugador'],bold=True,fg=BLAU_FOSC,align='left')
            fc(ws3,row3,3,r['equip_nom'],align='left')
            fc(ws3,row3,4,ratio(r['v1m'],r['v1x']))
            fc(ws3,row3,5,pct(r['v1m'],r['v1x']),bold=True,bg=ef_bg(r['v1m'],r['v1m']+r['v1x']))
            fc(ws3,row3,6,ratio(r['v2m'],r['v2x']))
            fc(ws3,row3,7,pct(r['v2m'],r['v2x']),bold=True,bg=ef_bg(r['v2m'],r['v2m']+r['v2x']))
            fc(ws3,row3,8,ratio(r['v3m'],r['v3x']))
            fc(ws3,row3,9,pct(r['v3m'],r['v3x']),bold=True,bg=ef_bg(r['v3m'],r['v3m']+r['v3x']))
            tt=r['v1m']+r['v1x']+r['v2m']+r['v2x']+r['v3m']+r['v3x']
            fc(ws3,row3,10,ratio(r['tm'],tt-r['tm']),bold=True)
            fc(ws3,row3,11,pct(r['tm'],tt-r['tm']),bold=True,
               bg=ef_bg(r['tm'],tt),fg=BLAU_FOSC)
            ws3.row_dimensions[row3].height=18; row3+=1

    # ── PESTANYA 4: ROTACIONS I +/- ──────────────────────────────────────
    ws4 = wb.create_sheet("Rotacions")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['A'].width = 2

    ws4.merge_cells('B1:L1')
    c=ws4['B1']; c.value='🏀  MICKI ANALÍTICA — ROTACIONS I +/-'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=14)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws4.row_dimensions[1].height=36

    ws4.merge_cells('B2:L2')
    c=ws4['B2']; c.value=f"Generat: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  {len(df_p)} partits"
    c.font=Font(name='Arial',color=BLANC,size=10); c.fill=fons(BLAU_MIG)
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws4.row_dimensions[2].height=20; ws4.row_dimensions[3].height=8

    # Secció 1: +/- individual per jugadora per partit
    row4 = 4
    ws4.merge_cells(f'B{row4}:L{row4}')
    c=ws4[f'B{row4}']; c.value='SECCIÓ 1 — +/- individual per jugadora i partit'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=11)
    c.fill=fons('0F6E56'); c.alignment=Alignment(horizontal='left',vertical='center')
    ws4.row_dimensions[row4].height=20; row4+=1

    for ci,cap,w in zip(range(2,11),
        ['Data','Equip','Jugadora','Minuts','Pts favor','Pts contra','+/-','+/- per min','Rival'],
        [12,20,26,10,11,11,10,12,20]):
        fc(ws4,row4,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,size=10)
        ws4.column_dimensions[get_column_letter(ci)].width=w
    ws4.row_dimensions[row4].height=20; row4+=1

    pm_acumulat = {}  # per calcular mitjanes

    for _,p in df_p.iterrows():
        df_jug_p = load_jugades_db(p['match_id'])
        if df_jug_p.empty: continue
        teams_p = get_teams_ordered(df_jug_p)
        tn_p = {}
        if len(teams_p)>=1: tn_p[teams_p[0]]=p['nom_a']
        if len(teams_p)>=2: tn_p[teams_p[1]]=p['nom_b']

        # Calcula intervals per aquest partit
        MINS_Q = 10
        intervals_p = {}; en_pista_p = {}
        df_jug_p_s = df_jug_p.sort_values("num")
        for _,row in df_jug_p_s.iterrows():
            jug_p = row.get("jugadora", row.get("jugador",""))
            if not jug_p or str(jug_p) in ("","nan"): continue
            accio_p = str(row.get("accio",""))
            quart_p = int(row.get("quart",1))
            min_p = float(row.get("min_num",0))
            t_p = (quart_p-1)*MINS_Q + (MINS_Q - min_p if min_p <= MINS_Q else min_p)
            t_p = max(0, min(t_p, quart_p*MINS_Q))
            eq_p = str(row.get("idEquip",""))
            if "Entra al camp" in accio_p:
                en_pista_p[jug_p]=(t_p,eq_p)
            elif "Surt del camp" in accio_p:
                ini_p=en_pista_p.pop(jug_p,((quart_p-1)*MINS_Q,eq_p))
                if t_p > ini_p[0]:
                    intervals_p.setdefault(jug_p,[]).append((ini_p[0],t_p,ini_p[1]))
            elif "Final de període" in accio_p:
                fi_p=quart_p*MINS_Q
                for jj,(ti_p,ei_p) in list(en_pista_p.items()):
                    if fi_p > ti_p: intervals_p.setdefault(jj,[]).append((ti_p,fi_p,ei_p))
                en_pista_p={}

        # Precalcula t_abs
        df_t = df_jug_p.copy()
        df_t["t_abs"] = df_t.apply(
            lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
            if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)

        for jug_x, ivs_x in intervals_p.items():
            for tid_x in teams_p[:2]:
                ivs_eq_x = [(ti,tf) for ti,tf,ei in ivs_x if ei==tid_x]
                if not ivs_eq_x: continue
                total_min_x = sum(tf-ti for ti,tf in ivs_eq_x)
                if total_min_x < 0.5: continue
                rival_x = [t for t in teams_p if t!=tid_x]
                rival_x = rival_x[0] if rival_x else None
                pf_x=pc_x=0
                for ti_x,tf_x in ivs_eq_x:
                    df_i = df_t[(df_t["t_abs"]>=ti_x)&(df_t["t_abs"]<=tf_x)]
                    pf_x += int(df_i[df_i["idEquip"]==tid_x]["punts"].sum())
                    if rival_x:
                        pc_x += int(df_i[df_i["idEquip"]==rival_x]["punts"].sum())
                pm_x = pf_x - pc_x
                pmm_x = round(pm_x/total_min_x,2) if total_min_x>0 else 0
                eq_nom_x = tn_p.get(tid_x,"?")
                rival_nom_x = tn_p.get(rival_x,"?") if rival_x else "?"

                bg_r = BLAU_CLAR if len(all_rows)%2==0 else BLANC if False else (BLAU_CLAR if row4%2==0 else BLANC)
                iv_color = '0F6E56' if pm_x>=0 else '993C1D'
                iv_bg = 'D5F5E3' if pm_x>0 else ('FADBD8' if pm_x<0 else BLANC)
                fc(ws4,row4,2,str(p['data_consulta'])[:10],align='center')
                fc(ws4,row4,3,eq_nom_x,bold=True,fg=BLAU_FOSC)
                fc(ws4,row4,4,jug_x,fg=BLAU_FOSC)
                fc(ws4,row4,5,round(total_min_x,1),align='center')
                fc(ws4,row4,6,pf_x,align='center')
                fc(ws4,row4,7,pc_x,align='center')
                fc(ws4,row4,8,f"{'+'if pm_x>=0 else ''}{pm_x}",align='center',bold=True,bg=iv_bg,fg=iv_color)
                fc(ws4,row4,9,pmm_x,align='center',num_fmt='+0.00;-0.00;0.00')
                fc(ws4,row4,10,rival_nom_x)
                ws4.row_dimensions[row4].height=17; row4+=1

                # Acumula per mitjanes
                key = (eq_nom_x, jug_x)
                pm_acumulat.setdefault(key, []).append(pm_x)

    # Fila de mitjanes per jugadora
    row4+=1
    ws4.merge_cells(f'B{row4}:L{row4}')
    c=ws4[f'B{row4}']; c.value='MITJANA +/- per jugadora (tots els partits)'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=11)
    c.fill=fons('0F6E56'); c.alignment=Alignment(horizontal='left',vertical='center')
    ws4.row_dimensions[row4].height=20; row4+=1

    for ci,cap in zip(range(2,8),['Equip','Jugadora','Partits','Mitjana +/-','Millor +/-','Pitjor +/-']):
        fc(ws4,row4,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,size=10)
    ws4.row_dimensions[row4].height=20; row4+=1

    for (eq_k,jug_k),vals in sorted(pm_acumulat.items(),key=lambda x:-sum(x[1])/len(x[1])):
        mitj = round(sum(vals)/len(vals),1)
        iv_bg2 = 'D5F5E3' if mitj>0 else ('FADBD8' if mitj<0 else BLANC)
        iv_col2 = '0F6E56' if mitj>=0 else '993C1D'
        fc(ws4,row4,2,eq_k,bold=True,fg=BLAU_FOSC)
        fc(ws4,row4,3,jug_k,fg=BLAU_FOSC)
        fc(ws4,row4,4,len(vals),align='center')
        fc(ws4,row4,5,f"{'+'if mitj>=0 else ''}{mitj}",align='center',bold=True,bg=iv_bg2,fg=iv_col2)
        fc(ws4,row4,6,f"+{max(vals)}",align='center',bg='D5F5E3',fg='0F6E56')
        fc(ws4,row4,7,f"{min(vals)}",align='center',bg='FADBD8',fg='993C1D')
        ws4.row_dimensions[row4].height=18; row4+=1

    # ── PESTANYA 5: ON/OFF RATING ──────────────────────────────────────────
    ws5 = wb.create_sheet("On-Off Rating")
    ws5.sheet_view.showGridLines=False; ws5.column_dimensions['A'].width=2
    ws5.merge_cells('B1:O1')
    c=ws5['B1']; c.value='🏀  MICKI ANALÍTICA — ON/OFF RATING PER JUGADORA'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=14)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws5.row_dimensions[1].height=36
    ws5.merge_cells('B2:O2')
    c=ws5['B2']; c.value=f"Generat: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  {len(df_p)} partits  ·  Min 4 possessions On i Off"
    c.font=Font(name='Arial',color=BLANC,size=10); c.fill=fons(BLAU_MIG)
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws5.row_dimensions[2].height=18; ws5.row_dimensions[3].height=6

    # Amplades
    for ci,w in zip(range(2,16),[24,18,9,9,9,9,9,9,9,9,9,9,10,10]):
        ws5.column_dimensions[get_column_letter(ci)].width=w

    row5=4
    # Capçalera de grups
    ws5.merge_cells(f'F{row5}:H{row5}')
    c=ws5[f'F{row5}']; c.value='ON (jugadora a pista)'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=9)
    c.fill=fons('0F6E56'); c.alignment=Alignment(horizontal='center',vertical='center'); c.border=vora()
    ws5.merge_cells(f'I{row5}:K{row5}')
    c=ws5[f'I{row5}']; c.value='OFF (jugadora fora)'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=9)
    c.fill=fons('993C1D'); c.alignment=Alignment(horizontal='center',vertical='center'); c.border=vora()
    for ci,cap in zip(range(2,16),[
        'Jugadora','Equip','Partits','Poss On','Poss Off',
        'Off Rtg','Def Rtg','Net Rtg',
        'Off Rtg','Def Rtg','Net Rtg',
        'Δ Net Rtg','Fiabilitat','Interpretació'
    ]):
        fc(ws5,row5+1,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,size=9)
        ws5.row_dimensions[row5+1].height=18
    row5+=2

    # Calcula On/Off per cada jugadora acumulat de tots els partits
    all_jugs_oo = {}
    for _,p_oo in df_p.iterrows():
        mid_oo = p_oo['match_id']
        df_oo = load_jugades_db(mid_oo)
        if df_oo.empty: continue
        teams_oo = get_teams_ordered(df_oo)
        col_j_oo = "jugador" if "jugador" in df_oo.columns else "jugadora"
        df_oo["jugador"] = df_oo[col_j_oo].fillna("")
        for jug_oo in df_oo["jugador"].unique():
            if not jug_oo or str(jug_oo) in ("","nan"): continue
            eq_oo = df_oo[df_oo["jugador"]==jug_oo]["idEquip"].iloc[0]
            oo_res = calc_onoff(df_oo, jug_oo, eq_oo, teams_oo)
            if oo_res is None: continue
            key_oo = (str(eq_oo), jug_oo)
            if key_oo not in all_jugs_oo:
                all_jugs_oo[key_oo] = []
            all_jugs_oo[key_oo].append(oo_res)

    # Noms equips
    eq_noms_oo = {}
    for _,p_oo in df_p.iterrows():
        df_oo = load_jugades_db(p_oo['match_id'])
        if df_oo.empty: continue
        teams_oo = get_teams_ordered(df_oo)
        tn_oo = {}
        if len(teams_oo)>=1: tn_oo[str(teams_oo[0])] = p_oo['nom_a']
        if len(teams_oo)>=2: tn_oo[str(teams_oo[1])] = p_oo['nom_b']
        eq_noms_oo.update(tn_oo)

    oo_rows5 = []
    for (eq_k_oo, jug_k_oo), oo_list in all_jugs_oo.items():
        # Promedia els valors de tots els partits
        def mitj_val(key):
            vals = [o[key] for o in oo_list if o.get(key) is not None]
            return round(sum(vals)/len(vals),1) if vals else None

        on_off  = mitj_val('on_off_rtg')
        on_def  = mitj_val('on_def_rtg')
        on_net  = mitj_val('on_net_rtg')
        off_off = mitj_val('off_off_rtg')
        off_def = mitj_val('off_def_rtg')
        off_net = mitj_val('off_net_rtg')
        diff    = mitj_val('diff')
        on_poss = mitj_val('on_poss')
        off_poss= mitj_val('off_poss')
        n_part  = len(oo_list)
        fiable  = (on_poss or 0)>=4 and (off_poss or 0)>=4

        if diff is None: continue
        oo_rows5.append({
            'jug': jug_k_oo, 'eq': eq_k_oo,
            'n': n_part, 'on_poss': on_poss, 'off_poss': off_poss,
            'on_off': on_off, 'on_def': on_def, 'on_net': on_net,
            'off_off': off_off, 'off_def': off_def, 'off_net': off_net,
            'diff': diff, 'fiable': fiable
        })

    oo_rows5.sort(key=lambda x: x['diff'] if x['diff'] else 0, reverse=True)

    for i,r in enumerate(oo_rows5):
        bg = BLAU_CLAR if i%2==0 else BLANC
        diff_v = r['diff'] or 0
        diff_bg = 'D5F5E3' if diff_v>2 else ('FADBD8' if diff_v<-2 else GROC)
        diff_fg = '0F6E56' if diff_v>2 else ('993C1D' if diff_v<-2 else '854F0B')
        fiab_txt = '✅ Fiable' if r['fiable'] else '⚠️ Poques poss'
        fiab_bg = BLANC if r['fiable'] else 'FFF3CD'

        # Interpretació automàtica
        if diff_v > 5: interp = 'Jugadora molt impactant'
        elif diff_v > 2: interp = 'Impacte positiu'
        elif diff_v > -2: interp = 'Impacte neutre'
        elif diff_v > -5: interp = 'Impacte negatiu'
        else: interp = 'Rendiment baix en pista'

        eq_nom_oo = eq_noms_oo.get(r['eq'], r['eq'][:8])
        fc(ws5,row5,2,r['jug'],bold=True,fg=BLAU_FOSC,bg=bg,align='left')
        fc(ws5,row5,3,eq_nom_oo,bg=bg,align='left',size=9)
        fc(ws5,row5,4,r['n'],bg=bg)
        fc(ws5,row5,5,r['on_poss'] or '—',bg=bg)
        fc(ws5,row5,6,r['off_poss'] or '—',bg=bg)
        fc(ws5,row5,7,r['on_off'] if r['on_off'] is not None else '—',bg='E8F5E9',fg='0F6E56',bold=True)
        fc(ws5,row5,8,r['on_def'] if r['on_def'] is not None else '—',bg='E8F5E9')
        fc(ws5,row5,9,r['on_net'] if r['on_net'] is not None else '—',bg='E8F5E9',bold=True)
        fc(ws5,row5,10,r['off_off'] if r['off_off'] is not None else '—',bg='FDE8E8',fg='993C1D',bold=True)
        fc(ws5,row5,11,r['off_def'] if r['off_def'] is not None else '—',bg='FDE8E8')
        fc(ws5,row5,12,r['off_net'] if r['off_net'] is not None else '—',bg='FDE8E8',bold=True)
        fc(ws5,row5,13,f"{'+'if diff_v>=0 else ''}{diff_v}",bold=True,bg=diff_bg,fg=diff_fg)
        fc(ws5,row5,14,fiab_txt,bg=fiab_bg,size=9)
        fc(ws5,row5,15,interp,bg=bg,align='left',size=9,fg='555555')
        ws5.row_dimensions[row5].height=17; row5+=1

    # ── Noms equips per match_id (per pestanyes 6 i 7) ─────────────────────
    eq_noms_per_match = {}
    for _,p_x in df_p.iterrows():
        df_x = load_jugades_db(p_x['match_id'])
        if df_x.empty: continue
        teams_x = get_teams_ordered(df_x)
        tn_x = {}
        if len(teams_x)>=1: tn_x[str(teams_x[0])] = p_x['nom_a']
        if len(teams_x)>=2: tn_x[str(teams_x[1])] = p_x['nom_b']
        eq_noms_per_match[p_x['match_id']] = tn_x

    # ── PESTANYA 6: QUINTETS (+/-) ──────────────────────────────────────────
    ws6 = wb.create_sheet("Quintets")
    ws6.sheet_view.showGridLines=False; ws6.column_dimensions['A'].width=2
    ws6.merge_cells('B1:K1')
    c=ws6['B1']; c.value='🏀  MICKI ANALÍTICA — +/- PER QUINTETS'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=14)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws6.row_dimensions[1].height=36
    ws6.merge_cells('B2:K2')
    c=ws6['B2']; c.value=f"Generat: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Combinacions de 5 jugadores a pista simultàniament"
    c.font=Font(name='Arial',color=BLANC,size=10); c.fill=fons(BLAU_MIG)
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws6.row_dimensions[2].height=18; ws6.row_dimensions[3].height=6

    for ci,w in zip(range(2,12),[18,52,9,9,9,9,11,40]):
        ws6.column_dimensions[get_column_letter(ci)].width=w

    row6=4
    # ── Secció 1: per partit ──
    ws6.merge_cells(f'B{row6}:I{row6}')
    c=ws6[f'B{row6}']; c.value='SECCIÓ 1 — Quintets per partit'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=11)
    c.fill=fons(BLAU_MIG); c.alignment=Alignment(horizontal='left',vertical='center')
    ws6.row_dimensions[row6].height=20; row6+=1

    for ci,cap in zip(range(2,10),['Partit','Quintet','Min','Pts favor','Pts contra','+/-','+/- per min','Equip']):
        fc(ws6,row6,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,size=9)
    ws6.row_dimensions[row6].height=18; row6+=1

    # Acumulat per quintet (clau = jugadores, independent del partit)
    quintets_acum = {}

    for _,p_x in df_p.iterrows():
        mid_x = p_x['match_id']
        df_x = load_jugades_db(mid_x)
        if df_x.empty: continue
        tn_x = eq_noms_per_match.get(mid_x, {})
        rows_q = calc_pm_combinacions(df_x, mode="quintets")
        rows_q.sort(key=lambda r: r['minuts'], reverse=True)
        lbl_partit = f"{tn_x.get(get_teams_ordered(df_x)[0],'?') if get_teams_ordered(df_x) else '?'} vs " \
                     f"{tn_x.get(get_teams_ordered(df_x)[1],'?') if len(get_teams_ordered(df_x))>1 else '?'}"
        for r_q in rows_q[:10]:  # top 10 per partit
            eq_nom_q = tn_x.get(r_q['equip'], r_q['equip'][:8])
            quintet_str = ", ".join(r_q['combinacio'])
            bg = BLAU_CLAR
            fc(ws6,row6,2,lbl_partit,bg=bg,size=8,align='left')
            fc(ws6,row6,3,quintet_str,bg=bg,size=8,align='left')
            fc(ws6,row6,4,r_q['minuts'],bg=bg)
            fc(ws6,row6,5,r_q['pf'],bg=bg)
            fc(ws6,row6,6,r_q['pc'],bg=bg)
            pm_bg = 'D5F5E3' if r_q['pm']>=0 else 'FADBD8'
            pm_fg = '0F6E56' if r_q['pm']>=0 else '993C1D'
            fc(ws6,row6,7,f"{'+'if r_q['pm']>=0 else ''}{r_q['pm']}",bold=True,bg=pm_bg,fg=pm_fg)
            fc(ws6,row6,8,r_q['pm_min'],num_fmt='+0.00;-0.00;0.00',bg=bg)
            fc(ws6,row6,9,eq_nom_q,bg=bg,size=8,align='left')
            ws6.row_dimensions[row6].height=15; row6+=1

            # Acumula
            key_q = (eq_nom_q, r_q['combinacio'])
            if key_q not in quintets_acum:
                quintets_acum[key_q] = {'minuts':0.0,'pf':0,'pc':0,'partits':0}
            quintets_acum[key_q]['minuts'] += r_q['minuts']
            quintets_acum[key_q]['pf'] += r_q['pf']
            quintets_acum[key_q]['pc'] += r_q['pc']
            quintets_acum[key_q]['partits'] += 1

    row6 += 2
    # ── Secció 2: acumulat temporada ──
    ws6.merge_cells(f'B{row6}:I{row6}')
    c=ws6[f'B{row6}']; c.value='SECCIÓ 2 — Quintets acumulats (temporada)'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=11)
    c.fill=fons(BLAU_MIG); c.alignment=Alignment(horizontal='left',vertical='center')
    ws6.row_dimensions[row6].height=20; row6+=1

    for ci,cap in zip(range(2,10),['Equip','Quintet','Partits','Min tot','Pts favor','Pts contra','+/-','+/- per min']):
        fc(ws6,row6,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,size=9)
    ws6.row_dimensions[row6].height=18; row6+=1

    quintets_sorted = sorted(quintets_acum.items(), key=lambda x: x[1]['minuts'], reverse=True)
    for i,((eq_nom_q,combo_q),d_q) in enumerate(quintets_sorted):
        bg = BLAU_CLAR if i%2==0 else BLANC
        pm_tot = d_q['pf']-d_q['pc']
        pm_min_tot = round(pm_tot/d_q['minuts'],3) if d_q['minuts']>0 else 0
        pm_bg = 'D5F5E3' if pm_tot>=0 else 'FADBD8'
        pm_fg = '0F6E56' if pm_tot>=0 else '993C1D'
        fc(ws6,row6,2,eq_nom_q,bold=True,fg=BLAU_FOSC,bg=bg,align='left',size=9)
        fc(ws6,row6,3,", ".join(combo_q),bg=bg,size=8,align='left')
        fc(ws6,row6,4,d_q['partits'],bg=bg)
        fc(ws6,row6,5,round(d_q['minuts'],1),bg=bg)
        fc(ws6,row6,6,d_q['pf'],bg=bg)
        fc(ws6,row6,7,d_q['pc'],bg=bg)
        fc(ws6,row6,8,f"{'+'if pm_tot>=0 else ''}{pm_tot}",bold=True,bg=pm_bg,fg=pm_fg)
        fc(ws6,row6,9,pm_min_tot,num_fmt='+0.00;-0.00;0.00',bg=bg)
        ws6.row_dimensions[row6].height=15; row6+=1

    # ── PESTANYA 7: PARELLES (+/-) ───────────────────────────────────────────
    ws7 = wb.create_sheet("Parelles")
    ws7.sheet_view.showGridLines=False; ws7.column_dimensions['A'].width=2
    ws7.merge_cells('B1:K1')
    c=ws7['B1']; c.value='🏀  MICKI ANALÍTICA — +/- PER PARELLES'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=14)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws7.row_dimensions[1].height=36
    ws7.merge_cells('B2:K2')
    c=ws7['B2']; c.value=f"Generat: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  +/- quan dues jugadores coincideixen a pista"
    c.font=Font(name='Arial',color=BLANC,size=10); c.fill=fons(BLAU_MIG)
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws7.row_dimensions[2].height=18; ws7.row_dimensions[3].height=6

    for ci,w in zip(range(2,12),[18,38,9,9,9,9,11,40]):
        ws7.column_dimensions[get_column_letter(ci)].width=w

    row7=4
    ws7.merge_cells(f'B{row7}:I{row7}')
    c=ws7[f'B{row7}']; c.value='SECCIÓ 1 — Parelles acumulades (temporada) — ordenat per minuts junts'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=11)
    c.fill=fons(BLAU_MIG); c.alignment=Alignment(horizontal='left',vertical='center')
    ws7.row_dimensions[row7].height=20; row7+=1

    for ci,cap in zip(range(2,10),['Equip','Parella','Partits','Min junts','Pts favor','Pts contra','+/-','+/- per min']):
        fc(ws7,row7,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,size=9)
    ws7.row_dimensions[row7].height=18; row7+=1

    parelles_acum = {}
    for _,p_x in df_p.iterrows():
        mid_x = p_x['match_id']
        df_x = load_jugades_db(mid_x)
        if df_x.empty: continue
        tn_x = eq_noms_per_match.get(mid_x, {})
        rows_par = calc_pm_combinacions(df_x, mode="parelles")
        for r_par in rows_par:
            eq_nom_par = tn_x.get(r_par['equip'], r_par['equip'][:8])
            key_par = (eq_nom_par, r_par['combinacio'])
            if key_par not in parelles_acum:
                parelles_acum[key_par] = {'minuts':0.0,'pf':0,'pc':0,'partits':set()}
            parelles_acum[key_par]['minuts'] += r_par['minuts']
            parelles_acum[key_par]['pf'] += r_par['pf']
            parelles_acum[key_par]['pc'] += r_par['pc']
            parelles_acum[key_par]['partits'].add(mid_x)

    parelles_sorted = sorted(parelles_acum.items(), key=lambda x: x[1]['minuts'], reverse=True)
    for i,((eq_nom_par,combo_par),d_par) in enumerate(parelles_sorted[:60]):
        bg = BLAU_CLAR if i%2==0 else BLANC
        pm_tot = d_par['pf']-d_par['pc']
        pm_min_tot = round(pm_tot/d_par['minuts'],3) if d_par['minuts']>0 else 0
        pm_bg = 'D5F5E3' if pm_tot>=0 else 'FADBD8'
        pm_fg = '0F6E56' if pm_tot>=0 else '993C1D'
        fc(ws7,row7,2,eq_nom_par,bold=True,fg=BLAU_FOSC,bg=bg,align='left',size=9)
        fc(ws7,row7,3," + ".join(combo_par),bg=bg,size=9,align='left')
        fc(ws7,row7,4,len(d_par['partits']),bg=bg)
        fc(ws7,row7,5,round(d_par['minuts'],1),bg=bg)
        fc(ws7,row7,6,d_par['pf'],bg=bg)
        fc(ws7,row7,7,d_par['pc'],bg=bg)
        fc(ws7,row7,8,f"{'+'if pm_tot>=0 else ''}{pm_tot}",bold=True,bg=pm_bg,fg=pm_fg)
        fc(ws7,row7,9,pm_min_tot,num_fmt='+0.00;-0.00;0.00',bg=bg)
        ws7.row_dimensions[row7].height=16; row7+=1

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

def genera_excel_temporada():
    """Genera un Excel formatat amb totes les dades de la temporada."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    BLAU_FOSC='0C3A6E'; BLAU_MIG='185FA5'; BLAU_CLAR='D6E8F7'
    BLAU_MOLT='EBF4FC'; GRIS_CAP='F2F4F6'; VERD='16A34A'
    VERMELL='C0392B'; BLANC='FFFFFF'; GROC='FFF3CD'

    def fons(c): return PatternFill('solid',fgColor=c)
    def vora():
        s=Side(style='thin',color='CCCCCC')
        return Border(top=s,bottom=s,left=s,right=s)
    def fc(ws,row,col,value,bold=False,bg=None,fg='000000',align='left',size=10):
        c=ws.cell(row=row,column=col,value=value)
        c.font=Font(name='Arial',bold=bold,color=fg,size=size)
        if bg: c.fill=fons(bg)
        c.alignment=Alignment(horizontal=align,vertical='center')
        c.border=vora()
        return c

    df_p=load_partits_db()
    df_s=load_stats_jugador_db()
    df_sh=load_shots_zones_db()

    wb=Workbook(); wb.remove(wb.active)

    # ── Pestanya 1: Temporada ─────────────────────────────────────────────
    ws1=wb.create_sheet("📋 Temporada")
    ws1.sheet_view.showGridLines=False
    ws1.column_dimensions['A'].width=2
    ws1.merge_cells('B1:H1')
    c=ws1['B1']; c.value='🏀  MICKI ANALÍTICA — RESUM DE TEMPORADA'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=15)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws1.row_dimensions[1].height=38
    ws1.merge_cells('B2:H2')
    c=ws1['B2']; c.value=f"Actualitzat: {datetime.now().strftime('%d/%m/%Y')}  ·  {len(df_p)} partits"
    c.font=Font(name='Arial',color=BLANC,size=10); c.fill=fons(BLAU_MIG)
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws1.row_dimensions[2].height=20; ws1.row_dimensions[3].height=8
    row=4
    for ci,cap,w in zip(range(2,9),['Data','Local','Pts L','Pts V','Visitant','W/L','Dif.'],[12,22,9,9,22,7,9]):
        fc(ws1,row,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,align='center',size=10)
        ws1.column_dimensions[get_column_letter(ci)].width=w
    ws1.row_dimensions[row].height=20; row+=1
    v=d=e=0
    for _,p in df_p.iterrows():
        sa,sb=int(p['score_a']),int(p['score_b'])
        if sa>sb: r='V'; v+=1; bgr='D5F5E3'; fgr=VERD
        elif sa<sb: r='D'; d+=1; bgr='FADBD8'; fgr=VERMELL
        else: r='E'; e+=1; bgr=GROC; fgr='8B6914'
        fc(ws1,row,2,str(p['data_consulta'])[:10],align='center')
        fc(ws1,row,3,p['nom_a'],bold=True,fg=BLAU_FOSC)
        fc(ws1,row,4,sa,align='center',bold=True,fg=BLAU_FOSC)
        fc(ws1,row,5,sb,align='center',bold=True,fg=VERMELL)
        fc(ws1,row,6,p['nom_b'])
        fc(ws1,row,7,r,align='center',bold=True,bg=bgr,fg=fgr)
        fc(ws1,row,8,f"{'+'if sa>sb else ''}{sa-sb}",align='center',bg=bgr,fg=fgr)
        ws1.row_dimensions[row].height=18; row+=1
    row+=1
    ws1.merge_cells(f'B{row}:H{row}')
    c=ws1[f'B{row}']
    c.value=f'  V: {v}    D: {d}    E: {e}    % Victòries: {round(v/max(len(df_p),1)*100)}%'
    c.font=Font(name='Arial',bold=True,color=BLAU_FOSC,size=11)
    c.fill=fons(BLAU_MOLT); c.alignment=Alignment(horizontal='left',vertical='center')
    from openpyxl.styles import Border, Side
    s=Side(style='thin',color=BLAU_MIG)
    c.border=Border(top=s,bottom=s,left=s,right=s)
    ws1.row_dimensions[row].height=26

    # ── Pestanya 2: Jugadores ─────────────────────────────────────────────
    ws2=wb.create_sheet("👤 Jugadores")
    ws2.sheet_view.showGridLines=False; ws2.column_dimensions['A'].width=2
    ws2.merge_cells('B1:N1')
    c=ws2['B1']; c.value='🏀  MICKI ANALÍTICA — RÀNQUING DE JUGADORES'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=15)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws2.row_dimensions[1].height=38; ws2.row_dimensions[2].height=8
    row=3
    for ci,cap,w in zip(range(2,15),
        ['#','Jugadora','Equip','Part.','Pts','Pts/P','Min','Min/P','C2','C3','TL','Faltes','Impacte'],
        [5,24,20,8,9,9,9,9,7,7,7,9,12]):
        fc(ws2,row,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,align='center',size=10)
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.row_dimensions[row].height=22; row+=1

    if not df_s.empty:
        agg=df_s.groupby(['jugador','equip_nom']).agg(
            p=('match_id','nunique'),pts=('punts','sum'),
            mn=('minuts','sum') if 'minuts' in df_s.columns else ('punts','count'),
            c2=('cistelles_2','sum'),c3=('cistelles_3','sum'),
            tl=('tirs_lliures','sum'),f=('faltes','sum'),imp=('impacte','sum')
        ).reset_index().sort_values('pts',ascending=False)
        for rank,(_,r) in enumerate(agg.iterrows(),1):
            bg=BLAU_MOLT if rank%2==0 else BLANC
            ppp=round(r['pts']/r['p'],1) if r['p']>0 else 0
            mpp=round(r['mn']/r['p'],1) if r['p']>0 else 0
            iv=f"+{int(r['imp'])}" if r['imp']>=0 else str(int(r['imp']))
            ic=VERD if r['imp']>=0 else VERMELL
            fc(ws2,row,2,rank,align='center',bg=bg,bold=True,fg=BLAU_FOSC)
            fc(ws2,row,3,r['jugador'],bold=True,bg=bg,fg=BLAU_FOSC)
            fc(ws2,row,4,r['equip_nom'],bg=bg)
            fc(ws2,row,5,int(r['p']),align='center',bg=bg)
            fc(ws2,row,6,int(r['pts']),align='center',bold=True,bg=bg,fg=BLAU_FOSC)
            fc(ws2,row,7,ppp,align='center',bg=bg)
            fc(ws2,row,8,round(r['mn'],1),align='center',bg=bg)
            fc(ws2,row,9,mpp,align='center',bg=bg)
            fc(ws2,row,10,int(r['c2']),align='center',bg=bg)
            fc(ws2,row,11,int(r['c3']),align='center',bg=bg)
            fc(ws2,row,12,int(r['tl']),align='center',bg=bg)
            fc(ws2,row,13,int(r['f']),align='center',bg=bg)
            fc(ws2,row,14,iv,align='center',bold=True,bg=bg,fg=ic)
            ws2.row_dimensions[row].height=18; row+=1

    # ── Pestanya 3: Evolució ──────────────────────────────────────────────
    ws3=wb.create_sheet("📈 Evolució per Partit")
    ws3.sheet_view.showGridLines=False; ws3.column_dimensions['A'].width=2
    ws3.merge_cells('B1:L1')
    c=ws3['B1']; c.value='🏀  MICKI ANALÍTICA — EVOLUCIÓ PER PARTIT'
    c.font=Font(name='Arial',bold=True,color=BLANC,size=15)
    c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
    ws3.row_dimensions[1].height=38; ws3.row_dimensions[2].height=8
    row=3
    if not df_s.empty:
        for jug in df_s['jugador'].unique():
            dj=df_s[df_s['jugador']==jug].sort_values('data_consulta')
            if dj.empty: continue
            ws3.merge_cells(f'B{row}:K{row}')
            c=ws3[f'B{row}']; c.value=f'  {jug}  ·  {dj.iloc[0]["equip_nom"]}'
            c.font=Font(name='Arial',bold=True,color=BLANC,size=11)
            c.fill=fons(BLAU_MIG); c.alignment=Alignment(horizontal='left',vertical='center')
            ws3.row_dimensions[row].height=22; row+=1
            for ci,cap,w in zip(range(2,12),
                ['Rival','Data','Punts','Min','C2','C3','TL','Faltes','Impacte','Pts/min'],
                [22,12,9,9,7,7,7,9,11,10]):
                fc(ws3,row,ci,cap,bold=True,bg=BLAU_CLAR,fg=BLAU_FOSC,align='center',size=9)
                ws3.column_dimensions[get_column_letter(ci)].width=max(
                    ws3.column_dimensions[get_column_letter(ci)].width or 0,w)
            ws3.row_dimensions[row].height=18; row+=1
            for _,r in dj.iterrows():
                pr=df_p[df_p['match_id']==r['match_id']]
                rival=pr.iloc[0]['nom_b'] if not pr.empty else '?'
                iv=f"+{int(r['impacte'])}" if r['impacte']>=0 else str(int(r['impacte']))
                ic=VERD if r['impacte']>=0 else VERMELL
                fc(ws3,row,2,rival,size=9); fc(ws3,row,3,str(r['data_consulta'])[:10],align='center',size=9)
                fc(ws3,row,4,int(r['punts']),align='center',bold=True,fg=BLAU_FOSC)
                fc(ws3,row,5,round(r.get('minuts',0),1),align='center')
                fc(ws3,row,6,int(r['cistelles_2']),align='center'); fc(ws3,row,7,int(r['cistelles_3']),align='center')
                fc(ws3,row,8,int(r['tirs_lliures']),align='center'); fc(ws3,row,9,int(r['faltes']),align='center')
                fc(ws3,row,10,iv,align='center',bold=True,fg=ic)
                fc(ws3,row,11,round(r['pts_per_min'],2),align='center')
                ws3.row_dimensions[row].height=16; row+=1
            it=int(dj['impacte'].sum())
            fc(ws3,row,2,'TOTAL',bold=True,bg=GRIS_CAP,fg=BLAU_FOSC,size=9)
            for ci in range(3,12): fc(ws3,row,ci,'',bg=GRIS_CAP)
            fc(ws3,row,4,int(dj['punts'].sum()),bold=True,align='center',bg=GRIS_CAP,fg=BLAU_FOSC)
            fc(ws3,row,6,int(dj['cistelles_2'].sum()),bold=True,align='center',bg=GRIS_CAP)
            fc(ws3,row,7,int(dj['cistelles_3'].sum()),bold=True,align='center',bg=GRIS_CAP)
            fc(ws3,row,8,int(dj['tirs_lliures'].sum()),bold=True,align='center',bg=GRIS_CAP)
            fc(ws3,row,9,int(dj['faltes'].sum()),bold=True,align='center',bg=GRIS_CAP)
            fc(ws3,row,10,f"+{it}" if it>=0 else str(it),bold=True,align='center',bg=GRIS_CAP,fg=VERD if it>=0 else VERMELL)
            ws3.row_dimensions[row].height=20; row+=2

    # ── Pestanya 4: Tirs ──────────────────────────────────────────────────
    if not df_sh.empty:
        ws4=wb.create_sheet("🎯 Eficiència de Tir")
        ws4.sheet_view.showGridLines=False; ws4.column_dimensions['A'].width=2
        ws4.merge_cells('B1:L1')
        c=ws4['B1']; c.value='🏀  MICKI ANALÍTICA — EFICIÈNCIA DE TIR'
        c.font=Font(name='Arial',bold=True,color=BLANC,size=15)
        c.fill=fons(BLAU_FOSC); c.alignment=Alignment(horizontal='center',vertical='center')
        ws4.row_dimensions[1].height=38; ws4.row_dimensions[2].height=8
        row=3
        for ci,cap,w in zip(range(2,13),
            ['Jugadora','Equip','TL conv/int','TL %','2pts conv/int','2pts %','3pts conv/int','3pts %','Total conv/int','Total %'],
            [24,20,14,9,16,9,16,9,16,9]):
            fc(ws4,row,ci,cap,bold=True,bg=BLAU_MIG,fg=BLANC,align='center',size=10)
            ws4.column_dimensions[get_column_letter(ci)].width=w
        ws4.row_dimensions[row].height=22; row+=1
        def ef_bg(m,t): 
            if t==0: return BLANC
            p=m/t
            return 'D5F5E3' if p>=0.55 else (GROC if p>=0.35 else 'FADBD8')
        at=df_sh[df_sh['jugador']!='__equip__'].groupby(['jugador','equip_nom']).agg(
            v1m=('val1_made','sum'),v1x=('val1_miss','sum'),
            v2m=('val2_made','sum'),v2x=('val2_miss','sum'),
            v3m=('val3_made','sum'),v3x=('val3_miss','sum'),
        ).reset_index()
        at['tm']=at['v1m']+at['v2m']+at['v3m']
        at=at.sort_values('tm',ascending=False)
        for _,r in at.iterrows():
            def ratio(m,x): return f"{int(m)}/{int(m+x)}" if (m+x)>0 else "—"
            def pct(m,x): return f"{round(m/(m+x)*100)}%" if (m+x)>0 else "—"
            fc(ws4,row,2,r['jugador'],bold=True,fg=BLAU_FOSC); fc(ws4,row,3,r['equip_nom'])
            fc(ws4,row,4,ratio(r['v1m'],r['v1x']),align='center')
            fc(ws4,row,5,pct(r['v1m'],r['v1x']),align='center',bold=True,bg=ef_bg(r['v1m'],r['v1m']+r['v1x']))
            fc(ws4,row,6,ratio(r['v2m'],r['v2x']),align='center')
            fc(ws4,row,7,pct(r['v2m'],r['v2x']),align='center',bold=True,bg=ef_bg(r['v2m'],r['v2m']+r['v2x']))
            fc(ws4,row,8,ratio(r['v3m'],r['v3x']),align='center')
            fc(ws4,row,9,pct(r['v3m'],r['v3x']),align='center',bold=True,bg=ef_bg(r['v3m'],r['v3m']+r['v3x']))
            tt=r['v1m']+r['v1x']+r['v2m']+r['v2x']+r['v3m']+r['v3x']
            fc(ws4,row,10,ratio(r['tm'],tt-r['tm']),align='center',bold=True)
            fc(ws4,row,11,pct(r['tm'],tt-r['tm']),align='center',bold=True,bg=ef_bg(r['tm'],tt),fg=BLAU_FOSC)
            ws4.row_dimensions[row].height=18; row+=1
        row+=1
        ws4.merge_cells(f'B{row}:F{row}')
        c=ws4[f'B{row}']; c.value='Llegenda:  🟢 ≥55% alta    🟡 35–54% mitja    🔴 <35% baixa'
        c.font=Font(name='Arial',italic=True,color='666666',size=9)
        c.alignment=Alignment(horizontal='left')

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

t1,t2,t3,t4,t5,t6,t7,t8,t9 = st.tabs([
    "🏀 Partit","👤 Jugadores","⏱ Ritme","⚡ Eficiència","🔄 Rotacions","📈 Hist. Jugadores","🎯 Mapa de Tir","🎬 Vídeo","📚 Històric"
])

# ══════════════════════════════════════════════════
# TAB 1: PARTIT
# ══════════════════════════════════════════════════
with t1:
    # Marcador
    # Parcials per quart: suma directa de punts del play-by-play per equip i quart
    def parcial_quart(tid, q):
        if not tid: return 0
        return int(df_orig[(df_orig["idEquip"]==tid) & (df_orig["quart"]==q)]["punts"].sum())

    # Resultat final: suma de tots els quarts
    def total_punts(tid):
        if not tid: return 0
        return int(df_orig[df_orig["idEquip"]==tid]["punts"].sum())

    # Recalculem fa/fb des del play-by-play (més fiable que el marcador de l'API)
    qs = sorted(df_orig["quart"].unique())
    # Resultat final = suma dels parcials de cada quart
    fa = sum(parcial_quart(teams[0] if teams else None, q) for q in qs)
    fb = sum(parcial_quart(teams[1] if len(teams)>1 else None, q) for q in qs)
    guanya_a = fa > fb
    guanya_b = fb > fa

    ca,cm,cb = st.columns([5,1,5])
    with ca:
        badge_a = '<div style="background:#E6F1FB;color:#0C447C;font-size:10px;font-weight:600;padding:3px 8px;border-radius:20px;display:inline-block;margin-bottom:8px">VICTÒRIA</div>' if guanya_a else ""
        border_a = f"2px solid {COLOR_A}" if guanya_a else "0.5px solid #e2e4e8"
        parc_a = " · ".join([f"Q{q}: {parcial_quart(teams[0] if teams else None, q)}" for q in qs])
        html_a = (
            f'<div style="background:#fff;border:{border_a};border-radius:16px;padding:24px 20px;text-align:center">'
            f'{badge_a}'
            f'<div style="font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:{COLOR_A};margin-bottom:8px">{nom_a}</div>'
            f'<div style="font-size:88px;font-weight:600;color:{COLOR_A};line-height:1">{fa}</div>'
            f'<div style="font-size:11px;color:#9ca3af;margin:6px 0">Local</div>'
            f'<div style="border-top:0.5px solid #f3f4f6;padding-top:8px;font-size:12px;color:{COLOR_A};opacity:.7">{parc_a}</div>'
            f'</div>'
        )
        st.markdown(html_a, unsafe_allow_html=True)
    with cm:
        st.markdown(
            f'<div style="text-align:center;padding-top:52px;font-size:20px;color:#d1d5db">vs</div>'
            f'<div style="text-align:center;margin-top:6px;font-size:10px;color:#d1d5db">{match_id[:8]}...</div>',
            unsafe_allow_html=True)
    with cb:
        badge_b = '<div style="background:#FAECE7;color:#712B13;font-size:10px;font-weight:600;padding:3px 8px;border-radius:20px;display:inline-block;margin-bottom:8px">VICTÒRIA</div>' if guanya_b else ""
        border_b = f"2px solid {COLOR_B}" if guanya_b else "0.5px solid #e2e4e8"
        parc_b = " · ".join([f"Q{q}: {parcial_quart(teams[1] if len(teams)>1 else None, q)}" for q in qs])
        html_b = (
            f'<div style="background:#fff;border:{border_b};border-radius:16px;padding:24px 20px;text-align:center">'
            f'{badge_b}'
            f'<div style="font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:{COLOR_B};margin-bottom:8px">{nom_b}</div>'
            f'<div style="font-size:88px;font-weight:600;color:{COLOR_B};line-height:1">{fb}</div>'
            f'<div style="font-size:11px;color:#9ca3af;margin:6px 0">Visitant</div>'
            f'<div style="border-top:0.5px solid #f3f4f6;padding-top:8px;font-size:12px;color:{COLOR_B};opacity:.7">{parc_b}</div>'
            f'</div>'
        )
        st.markdown(html_b, unsafe_allow_html=True)

    # Mètriques
    st.markdown(sec("Resum del partit"), unsafe_allow_html=True)
    # Resultat final directament del marcador de l'API
    try: pts_a = int(fa)
    except: pts_a = 0
    try: pts_b = int(fb)
    except: pts_b = 0
    faltes_a=int(df_orig[(df_orig["idEquip"]==teams[0])&df_orig["accio"].str.contains("falta",case=False,na=False)].shape[0]) if teams else 0
    faltes_b=int(df_orig[(df_orig["idEquip"]==teams[1])&df_orig["accio"].str.contains("falta",case=False,na=False)].shape[0]) if len(teams)>1 else 0
    c1,c2,c3,c4,c5,c6=st.columns(6)
    for col,lab,val,sub,col_ in zip([c1,c2,c3,c4,c5,c6],
        ["Jugades","Punts","Punts","Faltes","Faltes","Quarts"],
        [len(df_orig),pts_a,pts_b,faltes_a,faltes_b,df_orig["quart"].nunique()],
        ["total",nom_a,nom_b,nom_a,nom_b,"períodes"],
        ["#374151",COLOR_A,COLOR_B,COLOR_A,COLOR_B,"#374151"]):
        with col: st.markdown(card(lab,val,sub,col_),unsafe_allow_html=True)

    # Evolució marcador
    st.markdown(sec("Evolució del marcador"), unsafe_allow_html=True)
    if not score_df.empty:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=score_df["num"],y=score_df["scoreA"],name=nom_a,
            line=dict(color=COLOR_A,width=2.5),fill="tozeroy",fillcolor="rgba(24,95,165,0.07)"))
        fig.add_trace(go.Scatter(x=score_df["num"],y=score_df["scoreB"],name=nom_b,
            line=dict(color=COLOR_B,width=2.5),fill="tozeroy",fillcolor="rgba(153,60,29,0.07)"))
        for q in score_df["quart"].unique():
            fq=score_df[score_df["quart"]==q]["num"].min()
            if fq>1: fig.add_vline(x=fq,line_dash="dot",line_color="#e2e4e8",
                annotation_text=f"Q{q}",annotation_font_color="#9ca3af",annotation_font_size=10)
        fig.update_xaxes(title="Jugada"); fig.update_yaxes(title="Punts")
        fig.update_layout(
            legend=dict(orientation="h", yanchor="bottom", y=1.02,
                        xanchor="right", x=1, bgcolor="rgba(0,0,0,0)",
                        font=dict(size=12)))
        st.plotly_chart(chart_style(fig,300),use_container_width=True)

    # Punts per quart
    st.markdown(sec("Punts per quart"), unsafe_allow_html=True)
    pq=df_orig.groupby(["quart","equip_nom"])["punts"].sum().reset_index()
    if not pq.empty and pq["punts"].sum()>0:
        # Reconstruïm el mapa de colors a partir dels idEquip reals
        cmap = {}
        if teams: cmap[team_names.get(teams[0], nom_a)] = COLOR_A
        if len(teams)>1: cmap[team_names.get(teams[1], nom_b)] = COLOR_B
        fig2=px.bar(pq,x="quart",y="punts",color="equip_nom",barmode="group",
            color_discrete_map=cmap,labels={"quart":"Quart","punts":"Punts","equip_nom":"Equip"})
        st.plotly_chart(chart_style(fig2,240),use_container_width=True)

    # Mapa de calor minuts
    st.markdown(sec("En quins minuts marca cada equip?"), unsafe_allow_html=True)
    df_cist=df_orig[df_orig["punts"]>0].copy()
    df_cist["min_quart"]=df_cist["min_num"].apply(lambda x: int(x%10) if x>0 else 0)
    tab_ha,tab_hb=st.tabs([nom_a,nom_b])
    for htab,tid,ch in [(tab_ha,teams[0] if teams else None,COLOR_A),
                        (tab_hb,teams[1] if len(teams)>1 else None,COLOR_B)]:
        with htab:
            if tid is None: continue
            de=df_cist[df_cist["idEquip"]==tid]
            if de.empty: st.info("Sense cistelles."); continue
            hd=de.groupby(["quart","min_quart"])["punts"].sum().reset_index()
            piv=hd.pivot(index="quart",columns="min_quart",values="punts").fillna(0)
            fh=go.Figure(go.Heatmap(z=piv.values,x=[f"min {c}" for c in piv.columns],
                y=[f"Q{q}" for q in piv.index],colorscale=[[0,"#f9fafb"],[1,ch]],
                text=piv.values.astype(int),texttemplate="%{text}",
                hovertemplate="Q%{y} min%{x}: %{z}pts<extra></extra>"))
            fh.update_layout(paper_bgcolor="#fff",plot_bgcolor="#fff",
                font=dict(color="#374151",family="Inter"),margin=dict(l=0,r=0,t=10,b=0),height=200)
            st.plotly_chart(fh,use_container_width=True)

    # Play-by-Play
    st.markdown(sec("Play-by-Play"), unsafe_allow_html=True)
    df_f=df_orig.copy()
    if quart_sel: df_f=df_f[df_f["quart"].isin(quart_sel)]
    if accio_cerca: df_f=df_f[df_f["accio"].str.contains(accio_cerca,case=False,na=False)]
    if jugador_cerca: df_f=df_f[df_f["jugador"].str.contains(jugador_cerca,case=False,na=False)]
    st.caption(f"{len(df_f)} jugades")
    # Afegir badge d'equip a la taula
    def fmt_equip(eq):
        if eq==nom_a: return f"● {nom_a}"
        if eq==nom_b: return f"● {nom_b}"
        return eq
    # Play-by-play com taula HTML (única manera de renderitzar badges de color)
    rows_html = []
    for _,r in df_f.iterrows():
        eq = r["equip_nom"]
        if eq == nom_a:
            bdg = f'<span style="background:#E6F1FB;color:#0C447C;font-size:10px;font-weight:600;padding:2px 7px;border-radius:20px;white-space:nowrap">{nom_a[:4].upper()}</span>'
        elif eq == nom_b:
            bdg = f'<span style="background:#FAECE7;color:#712B13;font-size:10px;font-weight:600;padding:2px 7px;border-radius:20px;white-space:nowrap">{nom_b[:4].upper()}</span>'
        else:
            bdg = ""
        rows_html.append(
            f'<tr style="border-bottom:0.5px solid #f3f4f6">'
            f'<td style="padding:5px 8px;color:#9ca3af;font-size:11px;text-align:right">{int(r["num"])}</td>'
            f'<td style="padding:5px 4px;color:#9ca3af;font-size:11px;text-align:center">{r["quart"]}</td>'
            f'<td style="padding:5px 8px;color:#6b7280;font-size:11px;font-variant-numeric:tabular-nums">{r["temps"]}</td>'
            f'<td style="padding:5px 8px">{bdg}</td>'
            f'<td style="padding:5px 4px;color:#9ca3af;font-size:11px;text-align:center">{r["dorsal"]}</td>'
            f'<td style="padding:5px 8px;color:#374151;font-size:12px;font-weight:500">{r["jugador"]}</td>'
            f'<td style="padding:5px 8px;color:#374151;font-size:12px">{r["accio"]}</td>'
            f'<td style="padding:5px 8px;color:#6b7280;font-size:11px;font-variant-numeric:tabular-nums;text-align:right">{r["marcador"]}</td>'
            f'</tr>'
        )
    table_html = (
        '<div style="background:#fff;border:0.5px solid #e2e4e8;border-radius:10px;overflow:auto;max-height:400px">'
        '<table style="width:100%;border-collapse:collapse">'
        '<thead><tr style="border-bottom:1px solid #e2e4e8;background:#f9fafb;position:sticky;top:0">'
        '<th style="padding:6px 8px;font-size:10px;color:#9ca3af;font-weight:600;text-align:right;white-space:nowrap">#</th>'
        '<th style="padding:6px 4px;font-size:10px;color:#9ca3af;font-weight:600;text-align:center">Q</th>'
        '<th style="padding:6px 8px;font-size:10px;color:#9ca3af;font-weight:600">Temps</th>'
        '<th style="padding:6px 8px;font-size:10px;color:#9ca3af;font-weight:600">Equip</th>'
        '<th style="padding:6px 4px;font-size:10px;color:#9ca3af;font-weight:600;text-align:center">D</th>'
        '<th style="padding:6px 8px;font-size:10px;color:#9ca3af;font-weight:600">Jugadora</th>'
        '<th style="padding:6px 8px;font-size:10px;color:#9ca3af;font-weight:600">Acció</th>'
        '<th style="padding:6px 8px;font-size:10px;color:#9ca3af;font-weight:600;text-align:right">Marc</th>'
        '</tr></thead>'
        '<tbody>' + "".join(rows_html) + '</tbody>'
        '</table></div>'
    )
    st.markdown(table_html, unsafe_allow_html=True)
    csv_data=df_f[["num","quart","temps","idEquip","equip_nom","dorsal","jugador","accio","marcador","punts"]].to_csv(index=False).encode("utf-8")
    st.download_button("⬇ Descarregar CSV",csv_data,f"pbp_{match_id}.csv","text/csv")

# ══════════════════════════════════════════════════
# TAB 2: JUGADORES
# ══════════════════════════════════════════════════
with t2:
    st.markdown(sec("Punts per jugadora i quart"), unsafe_allow_html=True)
    stats_per=df_orig.groupby(["equip_nom","jugador","quart"]).agg(
        Punts=("punts","sum"),Accions=("accio","count"),
        Cistelles=("accio",lambda x: x.str.contains("Cistella",case=False,na=False).sum()),
        Faltes=("accio",lambda x: x.str.contains("falta",case=False,na=False).sum()),
    ).reset_index()
    tab_pa,tab_pb=st.tabs([nom_a,nom_b])
    for ptab,pnom,ph in [(tab_pa,nom_a,COLOR_A),(tab_pb,nom_b,COLOR_B)]:
        with ptab:
            dfe=stats_per[stats_per["equip_nom"]==pnom]
            if dfe.empty: st.info("Sense dades."); continue
            piv_p=dfe.pivot_table(index="jugador",columns="quart",values="Punts",aggfunc="sum",fill_value=0)
            piv_p["TOTAL"]=piv_p.sum(axis=1)
            piv_p=piv_p.sort_values("TOTAL",ascending=False)
            piv_p.columns=[f"Q{c}" if c!="TOTAL" else "TOTAL" for c in piv_p.columns]
            cols_q=[c for c in piv_p.columns if c!="TOTAL"]
            fig_hp=go.Figure(go.Heatmap(z=piv_p[cols_q].values,x=cols_q,y=piv_p.index.tolist(),
                colorscale=[[0,"#f9fafb"],[1,ph]],
                text=piv_p[cols_q].values.astype(int),texttemplate="%{text}",
                hovertemplate="%{y} — %{x}: %{z}pts<extra></extra>"))
            fig_hp.update_layout(paper_bgcolor="#fff",plot_bgcolor="#fff",
                font=dict(color="#374151",family="Inter"),
                margin=dict(l=0,r=0,t=10,b=0),height=max(180,36*len(piv_p)))
            st.plotly_chart(fig_hp,use_container_width=True)
            st.dataframe(piv_p,use_container_width=True)

    st.markdown(sec("Impacte en pista"), unsafe_allow_html=True)
    st.caption("Parcial de l'equip durant els minuts reals que la jugadora és a pista (intervals Entra/Surt).")
    imp_rows=[]
    MINS_Q_IMP = 10
    # Precalcula t_abs
    df_orig_imp = df_orig.copy()
    df_orig_imp["t_abs"] = df_orig_imp.apply(
        lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
        if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)

    col_j_imp = "jugador" if "jugador" in df_orig.columns else "jugadora"
    # Noms d'equip per filtrar
    noms_equip = [nom_a.upper(), nom_b.upper(), nom_a, nom_b]
    for jug in df_orig[col_j_imp].unique():
        if not jug or str(jug) in ("","nan"): continue
        # Filtra si el "jugador" és realment el nom de l'equip
        if str(jug).upper() in [n.upper() for n in noms_equip]: continue
        if len(str(jug).split()) > 4: continue  # noms d'equip solen ser llargs
        dj=df_orig[df_orig[col_j_imp]==jug]
        eq_id=dj["idEquip"].iloc[0]; eq_nom=dj["equip_nom"].iloc[0]
        rival=[t for t in teams if t!=eq_id]
        rival_id_imp = rival[0] if rival else None

        # Calcula intervals reals
        intervals_disp = []; en_pista_disp = {}
        primer_acc_disp = str(dj.sort_values("num").iloc[0].get("accio",""))
        primer_q_disp = int(dj.sort_values("num").iloc[0].get("quart",1))
        if "Surt" in primer_acc_disp and "camp" in primer_acc_disp:
            en_pista_disp[jug] = (primer_q_disp-1)*MINS_Q_IMP

        for _,row_d in df_orig.sort_values("num").iterrows():
            if str(row_d.get(col_j_imp,"")) != str(jug): continue
            acc_d = str(row_d.get("accio",""))
            q_d = int(row_d.get("quart",1))
            m_d = float(row_d.get("min_num",0))
            t_d = (q_d-1)*MINS_Q_IMP + (MINS_Q_IMP - m_d if m_d<=MINS_Q_IMP else m_d)
            t_d = max(0, min(t_d, q_d*MINS_Q_IMP))
            if "Entra" in acc_d and "camp" in acc_d:
                en_pista_disp[jug] = t_d
            elif "Surt" in acc_d and "camp" in acc_d:
                ti_d = en_pista_disp.pop(jug, (q_d-1)*MINS_Q_IMP)
                if t_d > ti_d: intervals_disp.append((ti_d, t_d))
            elif "Final de període" in acc_d:
                if jug in en_pista_disp:
                    ti_d = en_pista_disp.pop(jug)
                    fi_d = q_d*MINS_Q_IMP
                    if fi_d > ti_d: intervals_disp.append((ti_d, fi_d))
        for ti_o in en_pista_disp.values():
            fi_o = float(df_orig["quart"].max() * MINS_Q_IMP)
            if fi_o > ti_o: intervals_disp.append((ti_o, fi_o))
        intervals_disp = [(float(ti),float(tf)) for ti,tf in intervals_disp]

        if not intervals_disp:
            n_min,n_max=dj["num"].min(),dj["num"].max()
            dr=df_orig[(df_orig["num"]>=n_min)&(df_orig["num"]<=n_max)]
            pf=int(dr[dr["idEquip"]==eq_id]["punts"].sum())
            pc=int(dr[dr["idEquip"]==rival_id_imp]["punts"].sum()) if rival_id_imp else 0
        else:
            pf=pc=0
            for ti_d,tf_d in intervals_disp:
                df_i=df_orig_imp[(df_orig_imp["t_abs"]>=ti_d)&(df_orig_imp["t_abs"]<=tf_d)]
                pf+=int(df_i[df_i["idEquip"]==eq_id]["punts"].sum())
                if rival_id_imp:
                    pc+=int(df_i[df_i["idEquip"]==rival_id_imp]["punts"].sum())

        # Usage Rate per display
        if intervals_disp:
            df_orig_imp2 = df_orig.copy()
            df_orig_imp2["t_abs"] = df_orig_imp2.apply(
                lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
                if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)
            mask_us = df_orig_imp2["t_abs"].apply(
                lambda t: any(ti<=t<=tf for ti,tf in intervals_disp))
            df_eq_on_us = df_orig_imp2[mask_us & (df_orig_imp2["idEquip"]==eq_id)]
            usage_disp = calc_usage_rate(dj, df_eq_on_us)
        else:
            usage_disp = calc_usage_rate(dj, df_orig[df_orig["idEquip"]==eq_id])
        imp_rows.append({"Equip":eq_nom,"Jugadora":jug,"Pts favor":pf,"Pts contra":pc,
            "Parcial":f"+{pf-pc}" if pf>=pc else str(pf-pc),
            "Usage%": f"{usage_disp}%",
            "_diff":pf-pc})
    df_imp=pd.DataFrame(imp_rows).sort_values("_diff",ascending=False).drop(columns="_diff")
    t_ia,t_ib=st.tabs([nom_a,nom_b])
    for it,in_nom in [(t_ia,nom_a),(t_ib,nom_b)]:
        with it:
            di=df_imp[df_imp["Equip"]==in_nom].drop(columns="Equip")
            st.dataframe(di,use_container_width=True,hide_index=True)

    # ── Usage% vs Eficiència (scatter) ─────────────────────────────────────
    st.markdown(sec("Usage% vs Eficiència"), unsafe_allow_html=True)
    st.caption("Eix X = % de possessions usades · Eix Y = Pts/min · Quadrant ideal: dalt a la dreta")

    if imp_rows:
        df_scatter = pd.DataFrame(imp_rows).copy()
        df_scatter["_usage_num"] = df_scatter["Usage%"].str.replace("%","").astype(float)
        # Pts totals de la jugadora (de la taula d'estadístiques)
        df_scatter["_pts"] = df_scatter["Jugadora"].apply(
            lambda j: int(df_orig[df_orig[col_j_imp]==j]["punts"].sum()))
        # Minuts reals (dels intervals calculats)
        MINS_Q_SC = 10
        def get_minuts(jug):
            ivs = []; ep = {}
            dj_sc = df_orig[df_orig[col_j_imp]==jug]
            if dj_sc.empty: return 1
            pr = dj_sc.sort_values("num").iloc[0]
            if "Surt" in str(pr.get("accio","")) and "camp" in str(pr.get("accio","")):
                ep[jug] = (int(pr.get("quart",1))-1)*MINS_Q_SC
            for _,r in df_orig.sort_values("num").iterrows():
                if str(r.get(col_j_imp,"")) != str(jug): continue
                a = str(r.get("accio","")); q = int(r.get("quart",1))
                m = float(r.get("min_num",0))
                t = (q-1)*MINS_Q_SC + (MINS_Q_SC-m if m<=MINS_Q_SC else m)
                t = max(0, min(t, q*MINS_Q_SC))
                if "Entra" in a and "camp" in a: ep[jug] = t
                elif "Surt" in a and "camp" in a:
                    ti = ep.pop(jug, (q-1)*MINS_Q_SC)
                    if t > ti: ivs.append((ti, t))
                elif "Final de període" in a:
                    if jug in ep:
                        ti = ep.pop(jug); fi = float(q*MINS_Q_SC)
                        if fi > ti: ivs.append((ti, fi))
            for ti_o in ep.values():
                fi_o = float(df_orig["quart"].max()*MINS_Q_SC)
                if fi_o > ti_o: ivs.append((ti_o, fi_o))
            total = sum(tf-ti for ti,tf in ivs)
            return max(total, 1)
        df_scatter["_min"] = df_scatter["Jugadora"].apply(get_minuts)
        df_scatter["Pts/min"] = (df_scatter["_pts"] / df_scatter["_min"]).round(2)
        df_scatter["_color"] = df_scatter["Equip"].map({nom_a: COLOR_A, nom_b: COLOR_B})

        fig_sc = go.Figure()
        for eq, color in [(nom_a, COLOR_A), (nom_b, COLOR_B)]:
            df_eq = df_scatter[df_scatter["Equip"]==eq]
            if df_eq.empty: continue
            fig_sc.add_trace(go.Scatter(
                x=df_eq["_usage_num"],
                y=df_eq["Pts/min"],
                mode="markers+text",
                name=eq,
                marker=dict(size=14, color=color,
                            line=dict(width=1.5, color="white")),
                text=df_eq["Jugadora"].apply(
                    lambda n: n.split()[1] if len(n.split())>1 else n),
                textposition="top center",
                textfont=dict(size=9),
                hovertemplate=(
                    "<b>%{text}</b><br>"
                    "Usage: %{x:.1f}%<br>"
                    "Pts/min: %{y:.2f}<br>"
                    "<extra></extra>"
                )
            ))

        # Línies de mitjana
        mitjana_usage = df_scatter["_usage_num"].mean()
        mitjana_pts   = df_scatter["Pts/min"].mean()
        fig_sc.add_vline(x=mitjana_usage, line_dash="dot",
            line_color="#e2e4e8",
            annotation_text=f"Mitjana {mitjana_usage:.0f}%",
            annotation_font_size=9, annotation_font_color="#9ca3af")
        fig_sc.add_hline(y=mitjana_pts, line_dash="dot",
            line_color="#e2e4e8",
            annotation_text=f"Mitjana {mitjana_pts:.2f}",
            annotation_font_size=9, annotation_font_color="#9ca3af")

        # Etiquetes dels quadrants
        x_max = df_scatter["_usage_num"].max() * 1.1
        y_max = df_scatter["Pts/min"].max() * 1.1
        for txt, x, y, color in [
            ("⭐ Estrella", x_max*0.95, y_max*0.95, "#16a34a"),
            ("⚠️ Massa ús", x_max*0.95, y_max*0.05, "#dc2626"),
            ("💡 Infravalorada", x_max*0.05, y_max*0.95, "#185FA5"),
            ("🔄 Rol secundari", x_max*0.05, y_max*0.05, "#9ca3af"),
        ]:
            fig_sc.add_annotation(x=x, y=y, text=txt,
                showarrow=False, font=dict(size=9, color=color),
                xanchor="center", yanchor="middle", opacity=0.5)

        fig_sc.update_layout(
            xaxis=dict(title="Usage% (% possessions usades)",
                       showgrid=True, gridcolor="#f3f4f6", color="#9ca3af"),
            yaxis=dict(title="Pts/min",
                       showgrid=True, gridcolor="#f3f4f6", color="#9ca3af"),
            paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
            font=dict(color="#374151", family="Inter"),
            legend=dict(bgcolor="#ffffff", bordercolor="#e2e4e8",
                        borderwidth=1, orientation="h",
                        yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(l=0,r=0,t=40,b=0), height=380)
        st.plotly_chart(fig_sc, use_container_width=True)

    st.markdown(sec("Combinació de jugadores"), unsafe_allow_html=True)
    eq_combo=st.selectbox("Equip",[nom_a,nom_b],key="combo_eq")
    eq_id_combo=teams[0] if eq_combo==nom_a else (teams[1] if len(teams)>1 else None)
    if eq_id_combo:
        jugs_eq=sorted(df_orig[df_orig["idEquip"]==eq_id_combo]["jugador"].unique().tolist())
        jugs_sel=st.multiselect("Selecciona 2–5 jugadores",jugs_eq,max_selections=5,key="combo_jugs")
        if len(jugs_sel)>=2:
            mask_c=df_orig["jugador"].isin(jugs_sel)&(df_orig["idEquip"]==eq_id_combo)
            df_c=df_orig[mask_c]; n1,n2=df_c["num"].min(),df_c["num"].max()
            dr=df_orig[(df_orig["num"]>=n1)&(df_orig["num"]<=n2)]
            rival_c=[t for t in teams if t!=eq_id_combo]
            pf_c=int(dr[dr["idEquip"]==eq_id_combo]["punts"].sum())
            pc_c=int(dr[dr["idEquip"]==rival_c[0]]["punts"].sum()) if rival_c else 0
            diff_c=pf_c-pc_c; cr="#16a34a" if diff_c>=0 else "#dc2626"
            rt="GUANYEN" if diff_c>0 else ("EMPATEN" if diff_c==0 else "PERDEN")
            c1,c2,c3,c4=st.columns(4)
            with c1: st.markdown(card("Parcial favor",pf_c,"","#16a34a"),unsafe_allow_html=True)
            with c2: st.markdown(card("Parcial contra",pc_c,"","#dc2626"),unsafe_allow_html=True)
            with c3: st.markdown(card("Diferència",f"{'+'if diff_c>=0 else ''}{diff_c}","",cr),unsafe_allow_html=True)
            with c4: st.markdown(card("Resultat",rt,"",cr),unsafe_allow_html=True)
        elif len(jugs_sel)==1: st.info("Selecciona almenys 2 jugadores.")

    st.markdown(sec("Jugadora: puntua quan guanya o perd?"), unsafe_allow_html=True)
    tots_jugs=sorted(df_orig["jugador"].unique().tolist())
    jug_sel=st.selectbox("Jugadora",tots_jugs,key="jug_analisi")
    if jug_sel:
        dj2=df_orig[df_orig["jugador"]==jug_sel].copy()
        dj2["estat"]=dj2.apply(lambda r: estat_marc(r,teams),axis=1)
        res2=dj2.groupby("estat").agg(Punts=("punts","sum"),Accions=("accio","count"),
            Cistelles=("accio",lambda x: x.str.contains("Cistella",case=False,na=False).sum())
        ).reindex(["Guanyant","Empatat","Perdent"]).fillna(0).astype(int)
        ec={"Guanyant":"#16a34a","Empatat":"#d97706","Perdent":"#dc2626"}
        ei={"Guanyant":"📈","Empatat":"➡️","Perdent":"📉"}
        cg,ce,cp=st.columns(3)
        for col,estat in zip([cg,ce,cp],["Guanyant","Empatat","Perdent"]):
            with col:
                pts=int(res2.loc[estat,"Punts"]) if estat in res2.index else 0
                acc=int(res2.loc[estat,"Accions"]) if estat in res2.index else 0
                cis=int(res2.loc[estat,"Cistelles"]) if estat in res2.index else 0
                st.markdown(card(f"{ei[estat]} {estat}",pts,f"{acc} acc · {cis} cist",ec[estat]),unsafe_allow_html=True)
        if res2["Punts"].sum()>0:
            fig_j=px.bar(res2.reset_index().rename(columns={"estat":"Estat"}),
                x="Estat",y="Punts",color="Estat",color_discrete_map=ec,text="Punts")
            fig_j.update_traces(textposition="outside")
            st.plotly_chart(chart_style(fig_j,240),use_container_width=True)
        with st.expander("Detall cistelles"):
            dj2_pts=dj2[dj2["punts"]>0][["quart","temps","accio","marcador","punts","estat"]]
            st.dataframe(dj2_pts.rename(columns={"quart":"Q","temps":"Temps","accio":"Acció",
                "marcador":"Marc","punts":"Pts","estat":"Estat"}),use_container_width=True,hide_index=True)

# ══════════════════════════════════════════════════
# TAB 3: RITME
# ══════════════════════════════════════════════════
with t3:
    st.markdown(sec("Temps entre cistelles"), unsafe_allow_html=True)
    df_cist2=df_orig[df_orig["punts"]>0].sort_values("num").copy()
    for tid,tnom,tc in [(teams[0] if teams else None,nom_a,COLOR_A),
                        (teams[1] if len(teams)>1 else None,nom_b,COLOR_B)]:
        if tid is None: continue
        dc=df_cist2[df_cist2["idEquip"]==tid].copy()
        if len(dc)<2: continue
        dc["seg_entre"]=((dc["min_num"].shift(-1)-dc["min_num"])*60).abs()
        dc=dc.dropna(subset=["seg_entre"]); dc=dc[dc["seg_entre"]<600]
        mit=dc["seg_entre"].mean(); med=dc["seg_entre"].median()
        c1,c2,c3=st.columns(3)
        with c1: st.markdown(card(f"{tnom} — Mitjana",f"{mit:.0f}s","",tc),unsafe_allow_html=True)
        with c2: st.markdown(card("Mediana",f"{med:.0f}s","",tc),unsafe_allow_html=True)
        with c3: st.markdown(card("Cistelles",len(dc),"",tc),unsafe_allow_html=True)
        fig_t=px.histogram(dc,x="seg_entre",nbins=20,color_discrete_sequence=[tc],
            labels={"seg_entre":"Segons"})
        st.plotly_chart(chart_style(fig_t,200,f"{tnom} — temps entre cistelles"),use_container_width=True)

    st.markdown(sec("Ritme de puntuació (pts/min)"), unsafe_allow_html=True)
    ritme_rows=[]
    for q in sorted(df_orig["quart"].unique()):
        for tid,tnom in [(teams[0] if teams else None,nom_a),(teams[1] if len(teams)>1 else None,nom_b)]:
            if tid is None: continue
            pts_q=df_orig[(df_orig["quart"]==q)&(df_orig["idEquip"]==tid)]["punts"].sum()
            ritme_rows.append({"Quart":f"Q{q}","Equip":tnom,"Pts/min":round(pts_q/10,2)})
    df_ritme=pd.DataFrame(ritme_rows)
    if not df_ritme.empty:
        fig_r=px.line(df_ritme,x="Quart",y="Pts/min",color="Equip",
            color_discrete_map={nom_a:COLOR_A,nom_b:COLOR_B},markers=True)
        st.plotly_chart(chart_style(fig_r,260,"Ritme per quart"),use_container_width=True)

    st.markdown(sec("Eficiència ofensiva per quart"), unsafe_allow_html=True)
    ef_rows=[]
    for q in sorted(df_orig["quart"].unique()):
        for tid,tnom in [(teams[0] if teams else None,nom_a),(teams[1] if len(teams)>1 else None,nom_b)]:
            if tid is None: continue
            dq=df_orig[(df_orig["quart"]==q)&(df_orig["idEquip"]==tid)]
            cist=dq["accio"].str.contains("Cistella",case=False,na=False).sum()
            fall=dq["accio"].str.contains("Tir|fallit|fallat",case=False,na=False).sum()
            tot=cist+fall
            ef=round(cist/tot*100,1) if tot>0 else 0
            ef_rows.append({"Quart":f"Q{q}","Equip":tnom,"Eficiència %":ef})
    df_ef=pd.DataFrame(ef_rows)
    if not df_ef.empty:
        fig_ef=px.bar(df_ef,x="Quart",y="Eficiència %",color="Equip",barmode="group",
            color_discrete_map={nom_a:COLOR_A,nom_b:COLOR_B},text="Eficiència %")
        fig_ef.update_traces(texttemplate="%{text}%",textposition="outside")
        st.plotly_chart(chart_style(fig_ef,260,"% Eficiència per quart"),use_container_width=True)

    # ── Anàlisi de temps morts ────────────────────────────────────────────
    st.markdown(sec("⏸ Temps morts — qui anota després?"), unsafe_allow_html=True)
    st.caption("Primera cistella de l'equip que demana el temps mort, i quant triga a anotar-la.")

    to_data = analyze_timeouts(df_orig, team_names)
    if not to_data:
        st.info("No s'han detectat temps morts en aquest partit.")
    else:
        df_to = pd.DataFrame(to_data)

        # Mètriques generals
        c1,c2,c3,c4 = st.columns(4)
        total_to = len(df_to)
        anotats = df_to["va_anotar"].sum()
        efectivitat = round(anotats/total_to*100) if total_to>0 else 0
        seg_mitjana = df_to[df_to["va_anotar"]==1]["segons_resposta"].mean()

        with c1: st.markdown(card("Temps morts",total_to,"total","#374151"),unsafe_allow_html=True)
        with c2: st.markdown(card("Anoten després",int(anotats),"cistella","#16a34a"),unsafe_allow_html=True)
        with c3: st.markdown(card("Efectivitat",f"{efectivitat}%","","#185FA5"),unsafe_allow_html=True)
        with c4: st.markdown(card("Seg. fins cistella",f"{seg_mitjana:.0f}s" if not pd.isna(seg_mitjana) else "—","mitjana","#d97706"),unsafe_allow_html=True)

        # Taula detallada
        df_to_show = df_to.copy()
        df_to_show["Q"] = df_to_show["quart"]
        df_to_show["Equip"] = df_to_show["equip_nom"]
        df_to_show["Anota?"] = df_to_show["va_anotar"].map({1:"✅ Sí", 0:"❌ No"})
        df_to_show["Jugadora"] = df_to_show["jugadora"]
        df_to_show["Acció"] = df_to_show["accio"]
        df_to_show["Seg."] = df_to_show["segons_resposta"].apply(lambda x: f"{x:.0f}s" if x and not pd.isna(x) else "—")
        df_to_show["≤24s?"] = df_to_show.get("dins_24s", pd.Series([0]*len(df_to_show))).map({1:"✅ Sí", 0:"❌ No"})
        st.dataframe(df_to_show[["Q","Equip","Anota?","≤24s?","Jugadora","Acció","Seg."]],
            use_container_width=True, hide_index=True)

        # Gràfic per equip
        if len(df_to) > 1:
            resum_eq = df_to.groupby("equip_nom").agg(
                Total=("va_anotar","count"),
                Anotats=("va_anotar","sum"),
                Seg_mitjana=("segons_resposta","mean")
            ).reset_index()
            resum_eq["Efectivitat %"] = (resum_eq["Anotats"]/resum_eq["Total"]*100).round(0)
            fig_to = px.bar(resum_eq, x="equip_nom", y="Efectivitat %",
                color="equip_nom", color_discrete_map=color_map_eq,
                text="Efectivitat %", labels={"equip_nom":"Equip"})
            fig_to.update_traces(texttemplate="%{text}%", textposition="outside")
            st.plotly_chart(chart_style(fig_to, 220, "Efectivitat dels temps morts per equip"), use_container_width=True)

    st.markdown(sec("Momentum shifts"), unsafe_allow_html=True)
    st.caption("Runs de 5+ punts consecutius sense resposta del rival.")
    THRESHOLD=5; shift_rows=[]
    if not score_df.empty:
        prev_diff,run_team,run_pts,run_start=0,None,0,None
        for _,row in score_df.iterrows():
            diff=row["diff"]; delta=diff-prev_diff
            cur=(teams[0] if teams else None) if delta>0 else ((teams[1] if len(teams)>1 else None) if delta<0 else None)
            if cur and cur==run_team: run_pts+=abs(delta)
            else:
                if run_pts>=THRESHOLD and run_team:
                    shift_rows.append({"Equip":team_names.get(run_team,"?"),
                        "Jugada inici":run_start,"Jugada fi":row["num"],
                        "Parcial":f"+{run_pts}","Temps":row["temps"],"Quart":row["quart"]})
                run_team,run_pts,run_start=cur,abs(delta),row["num"]
            prev_diff=diff
    if shift_rows:
        df_sh=pd.DataFrame(shift_rows)
        st.dataframe(df_sh,use_container_width=True,hide_index=True)
    else:
        st.info(f"No s'han detectat runs de {THRESHOLD}+ punts.")

# ══════════════════════════════════════════════════
# TAB 4: HISTÒRIC PARTITS
# ══════════════════════════════════════════════════
with t9:
    st.markdown(sec("Partits consultats"), unsafe_allow_html=True)
    df_hist=load_partits_db()
    if df_hist.empty:
        st.info("Encara no hi ha partits. Carrega un partit per començar.")
    else:
        df_hs=df_hist.copy()
        df_hs["Resultat"]=df_hs.apply(lambda r: f"{r['score_a']}–{r['score_b']}",axis=1)
        st.dataframe(df_hs[["data_consulta","nom_a","Resultat","nom_b","total_jugades","match_id"]].rename(
            columns={"data_consulta":"Data","nom_a":"Local","nom_b":"Visitant",
                     "total_jugades":"Jugades","match_id":"ID"}),
            use_container_width=True,hide_index=True)

        st.markdown(sec("Carregar un partit de l'històric"), unsafe_allow_html=True)
        ids_hist=df_hist["match_id"].tolist()
        sel_hist=st.selectbox("Selecciona",ids_hist,
            format_func=lambda x: f"{df_hist[df_hist['match_id']==x]['nom_a'].values[0]} vs {df_hist[df_hist['match_id']==x]['nom_b'].values[0]}",
            key="sel_hist")
        if st.button("📂 Carregar",key="load_hist"):
            df_loaded=load_jugades_db(sel_hist)
            st.session_state.df=df_loaded
            st.session_state.match_id=sel_hist
            row=df_hist[df_hist["match_id"]==sel_hist].iloc[0]
            t_list=get_teams(df_loaded)
            noms={}
            if len(t_list)>=1: noms[t_list[0]]=row["nom_a"]
            if len(t_list)>=2: noms[t_list[1]]=row["nom_b"]
            st.session_state.team_names=noms
            st.session_state.score_a = int(row["score_a"])
            st.session_state.score_b = int(row["score_b"])
            st.success("Carregat! Ves a 🏀 Partit."); st.rerun()

        col_del1, col_del2 = st.columns([2,1])
        with col_del1:
            del_id=st.selectbox("Eliminar un partit",ids_hist,
                format_func=lambda x: f"{df_hist[df_hist['match_id']==x]['nom_a'].values[0]} vs {df_hist[df_hist['match_id']==x]['nom_b'].values[0]}",
                key="del_hist")
        with col_del2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button("🗑 Eliminar",key="btn_del"):
                delete_partit_db(del_id); st.success("Eliminat."); st.rerun()

        # Botó per descarregar Excel de temporada
        st.markdown(sec("Exporta la temporada a Excel"), unsafe_allow_html=True)
        st.caption("Excel formatat amb 4 pestanyes: Temporada · Jugadores · Evolució per Partit · Eficiència de Tir")
        if st.button("⬇ Descarregar Excel de temporada", key="btn_excel"):
            excel_data = genera_excel_temporada()
            st.download_button(
                label="📥 Clic per descarregar",
                data=excel_data,
                file_name=f"miki_analitica_temporada_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel"
            )

        if len(df_hist)>1:
            st.markdown(sec("Evolució de resultats"), unsafe_allow_html=True)
            rows_comp=[]
            for _,rh in df_hist.iterrows():
                lbl=f"{rh['nom_a']} vs {rh['nom_b']} ({rh['data_consulta'][:10]})"
                rows_comp.append({"Partit":lbl,"Equip":rh["nom_a"],"Punts":rh["score_a"]})
                rows_comp.append({"Partit":lbl,"Equip":rh["nom_b"],"Punts":rh["score_b"]})
            fig_comp=px.line(pd.DataFrame(rows_comp),x="Partit",y="Punts",color="Equip",markers=True)
            fig_comp.update_xaxes(tickangle=-30)
            st.plotly_chart(chart_style(fig_comp,280),use_container_width=True)

# ══════════════════════════════════════════════════
# TAB 5: HISTÒRIC JUGADORES
# ══════════════════════════════════════════════════
with t4:
    # ── ROT — Índex de gestió de rotacions ─────────────────────────────────
    st.markdown(sec("ROT — Índex de gestió de rotacions"), unsafe_allow_html=True)
    st.caption("ROT = 5 · (ρ + 1)  on ρ = correlació de Pearson entre minuts jugats i +/- per minut de cada jugadora. Escala 0-10. ROT alt = les jugadores que juguen més aporten més.")

    col_j_rot = "jugador" if "jugador" in df_orig.columns else "jugadora"
    rot_data = []

    for eq_rot in teams:
        rival_rot = [t for t in teams if t != eq_rot]
        rival_id_rot = rival_rot[0] if rival_rot else None
        if rival_id_rot is None: continue

        df_t_rot = df_orig.copy()
        df_t_rot["t_abs"] = df_t_rot.apply(
            lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
            if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)

        jugs_rot = [j for j in df_orig[df_orig["idEquip"]==eq_rot][col_j_rot].unique()
                    if j and str(j) not in ("","nan")]

        for jug_rot in jugs_rot:
            # Calcula intervals reals (el mateix mètode que usem a tot arreu)
            MINS_Q_R = 10
            ivs_rot = []; ep_rot = {}
            dj_rot = df_orig[df_orig[col_j_rot]==jug_rot]
            if dj_rot.empty: continue
            pr_rot = dj_rot.sort_values("num").iloc[0]
            if "Surt" in str(pr_rot.get("accio","")) and "camp" in str(pr_rot.get("accio","")):
                ep_rot[jug_rot] = (int(pr_rot.get("quart",1))-1)*MINS_Q_R
            for _,row_r in df_orig.sort_values("num").iterrows():
                if str(row_r.get(col_j_rot,"")) != jug_rot: continue
                a_r=str(row_r.get("accio","")); q_r=int(row_r.get("quart",1))
                m_r=float(row_r.get("min_num",0))
                t_r=(q_r-1)*MINS_Q_R+(MINS_Q_R-m_r if m_r<=MINS_Q_R else m_r)
                t_r=max(0,min(t_r,q_r*MINS_Q_R))
                if "Entra" in a_r and "camp" in a_r: ep_rot[jug_rot]=t_r
                elif "Surt" in a_r and "camp" in a_r:
                    ti_r=ep_rot.pop(jug_rot,(q_r-1)*MINS_Q_R)
                    if t_r>ti_r: ivs_rot.append((float(ti_r),float(t_r)))
                elif "Final de període" in a_r:
                    if jug_rot in ep_rot:
                        ti_r=ep_rot.pop(jug_rot); fi_r=float(q_r*MINS_Q_R)
                        if fi_r>ti_r: ivs_rot.append((ti_r,fi_r))
            for ti_o in ep_rot.values():
                fi_o=float(df_orig["quart"].max()*MINS_Q_R)
                if fi_o>ti_o: ivs_rot.append((float(ti_o),fi_o))

            if not ivs_rot: continue

            # Minuts des dels intervals reals (igual que a Rotacions)
            min_rot = round(sum(tf-ti for ti,tf in ivs_rot), 1)
            if min_rot < 0.5: continue

            # +/- des dels mateixos intervals
            pf_rot=pc_rot=0
            for ti_r,tf_r in ivs_rot:
                df_i_rot=df_t_rot[(df_t_rot["t_abs"]>=ti_r)&(df_t_rot["t_abs"]<=tf_r)]
                pf_rot+=int(df_i_rot[df_i_rot["idEquip"]==eq_rot]["punts"].sum())
                pc_rot+=int(df_i_rot[df_i_rot["idEquip"]==rival_id_rot]["punts"].sum())
            pm_rot = pf_rot - pc_rot
            pm_min_rot = pm_rot / min_rot if min_rot > 0 else 0
            rot_data.append({"equip": eq_rot, "jugadora": jug_rot, "minuts": min_rot, "pm_min": pm_min_rot})

    # Calcula ROT per equip
    for eq_rot in teams:
        df_rot_eq = pd.DataFrame([r for r in rot_data if r["equip"]==eq_rot])
        if df_rot_eq.empty or len(df_rot_eq) < 3: continue

        from scipy import stats as sp_stats_rot
        rho_rot, pval_rot = sp_stats_rot.pearsonr(df_rot_eq["minuts"], df_rot_eq["pm_min"])
        rot_val = round(5 * (rho_rot + 1), 2)

        eq_nom_rot = nom_a if eq_rot==teams[0] else nom_b
        color_rot = COLOR_A if eq_rot==teams[0] else COLOR_B

        if rot_val >= 7.5: rot_interp = "🟢 Rotacions excel·lents"
        elif rot_val >= 5.5: rot_interp = "🟡 Rotacions bones"
        elif rot_val >= 3.5: rot_interp = "🟠 Rotacions millorables"
        else: rot_interp = "🔴 Rotacions per revisar"

        with st.container():
            st.markdown(f'<div style="font-size:13px;font-weight:700;color:{color_rot};margin-bottom:6px">🏀 {eq_nom_rot}</div>', unsafe_allow_html=True)
            c1,c2,c3,c4 = st.columns(4)
            with c1: st.markdown(card("ROT", f"{rot_val}/10", rot_interp, color_rot), unsafe_allow_html=True)
            with c2: st.markdown(card("Pearson (ρ)", f"{rho_rot:+.3f}", "minuts vs +/-/min", color_rot), unsafe_allow_html=True)
            with c3: st.markdown(card("p-value", f"{pval_rot:.3f}", "sig. si <0.05", "#374151"), unsafe_allow_html=True)
            with c4: st.markdown(card("Jugadores", len(df_rot_eq), "analitzades", "#374151"), unsafe_allow_html=True)

            # Scatter Minuts vs +/- per minut
            import numpy as np
            fig_rot = go.Figure()
            colors_rot = [COLOR_A if eq_rot==teams[0] else COLOR_B] * len(df_rot_eq)
            fig_rot.add_trace(go.Scatter(
                x=df_rot_eq["minuts"], y=df_rot_eq["pm_min"],
                mode="markers+text",
                marker=dict(size=10, color=color_rot, line=dict(width=1, color="white")),
                text=df_rot_eq["jugadora"].apply(lambda n: n.split()[1] if len(n.split())>1 else n),
                textposition="top center", textfont=dict(size=8),
                hovertemplate="<b>%{text}</b><br>Minuts: %{x:.1f}<br>+/- per min: %{y:+.3f}<extra></extra>"
            ))
            m_rot, b_rot = np.polyfit(df_rot_eq["minuts"], df_rot_eq["pm_min"], 1)
            x_lr = [df_rot_eq["minuts"].min(), df_rot_eq["minuts"].max()]
            y_lr = [m_rot*x+b_rot for x in x_lr]
            fig_rot.add_trace(go.Scatter(
                x=x_lr, y=y_lr, mode="lines",
                line=dict(color="#d97706", width=2, dash="dot"),
                name=f"Tendència (ρ={rho_rot:+.3f})", showlegend=True))
            fig_rot.add_hline(y=0, line_dash="solid", line_color="#e2e4e8")
            fig_rot.update_xaxes(title="Minuts jugats")
            fig_rot.update_yaxes(title="+/- per minut")
            st.plotly_chart(chart_style(fig_rot, 260, f"{eq_nom_rot} — ROT {rot_val}/10 (ρ={rho_rot:+.3f})"), use_container_width=True)

    st.markdown(sec("⚡ Eficiència i On/Off Rating per jugadora"), unsafe_allow_html=True)
    st.caption(
        "On/Off Net Rating = diferència de Net Rating (pts/100 poss) quan la jugadora és a pista vs quan no hi és. "
        "Exemple: +5.2 = l'equip guanya 5.2 pts més per 100 poss amb la jugadora. "
        "⚠️ = poques possessions Off, valor poc fiable."
    )
    df_pr_onoff = load_partits_db()
    if df_pr_onoff.empty:
        st.info("Carrega partits per veure l'On/Off Rating.")
    else:
        ids_onoff = df_pr_onoff["match_id"].tolist()
        sel_onoff = st.selectbox("Selecciona un partit", ids_onoff,
            format_func=lambda x: f"{df_pr_onoff[df_pr_onoff['match_id']==x]['nom_a'].values[0]} vs "
                                   f"{df_pr_onoff[df_pr_onoff['match_id']==x]['nom_b'].values[0]} "
                                   f"({df_pr_onoff[df_pr_onoff['match_id']==x]['data_consulta'].values[0][:10]})",
            key="sel_onoff2")
        df_onoff2 = load_jugades_db(sel_onoff)
        if not df_onoff2.empty:
            teams_oo2 = get_teams_ordered(df_onoff2)
            p_row2 = df_pr_onoff[df_pr_onoff["match_id"]==sel_onoff].iloc[0]
            tn_oo2 = {}
            if len(teams_oo2)>=1: tn_oo2[teams_oo2[0]] = p_row2["nom_a"]
            if len(teams_oo2)>=2: tn_oo2[teams_oo2[1]] = p_row2["nom_b"]
            df_onoff2["equip_nom"] = df_onoff2["idEquip"].map(tn_oo2).fillna("?")
            df_onoff2["jugador"] = df_onoff2.get("jugadora", df_onoff2.get("jugador",""))
            st.markdown("**Eficiències d'equip**")
            ef2 = calc_eficiencies(df_onoff2, teams_oo2, tn_oo2)
            col_ea2, col_eb2 = st.columns(2)
            for col_e2, tid2 in [(col_ea2, teams_oo2[0] if teams_oo2 else None),
                                  (col_eb2, teams_oo2[1] if len(teams_oo2)>1 else None)]:
                with col_e2:
                    if tid2 and tid2 in ef2:
                        e2 = ef2[tid2]
                        tcol2 = COLOR_A if tid2 == teams_oo2[0] else COLOR_B
                        c1,c2,c3,c4 = st.columns(4)
                        with c1: st.markdown(card("Possessions", e2["poss_of"], "", tcol2), unsafe_allow_html=True)
                        with c2: st.markdown(card("Off Rating", e2["off_rtg"], "pts/100 poss", tcol2), unsafe_allow_html=True)
                        with c3: st.markdown(card("Def Rating", e2["def_rtg"], "pts/100 poss", "#dc2626"), unsafe_allow_html=True)
                        nc2 = "#16a34a" if e2["net_rtg"] >= 0 else "#dc2626"
                        with c4: st.markdown(card("Net Rating",
                            f"{'+'if e2['net_rtg']>=0 else ''}{e2['net_rtg']}",
                            e2["nom"], nc2), unsafe_allow_html=True)
            st.markdown("**On/Off Rating per jugadora**")
            eq_oo2 = st.selectbox("Equip", [tn_oo2.get(t,"?") for t in teams_oo2], key="eq_oo2")
            tid_oo2 = teams_oo2[0] if eq_oo2 == tn_oo2.get(teams_oo2[0],"?") else (teams_oo2[1] if len(teams_oo2)>1 else None)
            if tid_oo2:
                col_jug2 = "jugadora" if "jugadora" in df_onoff2.columns else "jugador"
                jugs_oo2 = sorted(df_onoff2[
                    (df_onoff2["idEquip"]==tid_oo2) & (df_onoff2[col_jug2] != "")
                ][col_jug2].unique().tolist())
                oo_rows2 = []
                for jug2 in jugs_oo2:
                    df_jug_oo2 = df_onoff2.copy()
                    df_jug_oo2["jugador"] = df_jug_oo2[col_jug2]
                    oo2 = calc_onoff(df_jug_oo2, jug2, tid_oo2, teams_oo2)
                    if oo2:
                        if oo2 and oo2.get("diff") is not None:
                            off_poss = oo2.get("off_poss", 0)
                            fiable = off_poss >= 4
                            label_diff = f"{'+'if oo2['diff']>=0 else ''}{oo2['diff']}"
                            if not fiable:
                                label_diff += " ⚠️"
                            oo_rows2.append({
                                "Jugadora": jug2,
                                "On Poss": oo2.get("on_poss","—"),
                                "On Off Rtg": oo2["on_off_rtg"] if oo2["on_off_rtg"] is not None else "N/D",
                                "On Def Rtg": oo2["on_def_rtg"] if oo2["on_def_rtg"] is not None else "N/D",
                                "On Net Rtg": oo2["on_net_rtg"] if oo2["on_net_rtg"] is not None else "N/D",
                                "Off Poss": off_poss,
                                "Off Off Rtg": oo2["off_off_rtg"] if oo2["off_off_rtg"] is not None else "N/D",
                                "Off Def Rtg": oo2["off_def_rtg"] if oo2["off_def_rtg"] is not None else "N/D",
                                "Off Net Rtg": oo2["off_net_rtg"] if oo2["off_net_rtg"] is not None else "N/D",
                                "Diferència": label_diff,
                                "_diff": oo2["diff"],
                                "_fiable": fiable
                            })
                        elif oo2:
                            oo_rows2.append({
                                "Jugadora": jug2,
                                "On Poss": oo2.get("on_poss","—"),
                                "On Off Rtg": "N/D", "On Def Rtg": "N/D", "On Net Rtg": "N/D",
                                "Off Poss": oo2.get("off_poss","—"),
                                "Off Off Rtg": "N/D", "Off Def Rtg": "N/D", "Off Net Rtg": "N/D",
                                "Diferència": "N/D (poques poss.)", "_diff": 0, "_fiable": False
                            })
                if oo_rows2:
                    df_oo2 = pd.DataFrame(oo_rows2).sort_values("_diff", ascending=False)
                    colors_bar2 = []
                    for _, row_oo in df_oo2.iterrows():
                        if not row_oo.get("_fiable", True):
                            colors_bar2.append("#d97706")  # taronja = poc fiable
                        elif row_oo["_diff"] >= 0:
                            colors_bar2.append("#16a34a")
                        else:
                            colors_bar2.append("#dc2626")
                    fig_oo2 = go.Figure()
                    fig_oo2.add_trace(go.Bar(
                        x=df_oo2["Jugadora"], y=df_oo2["_diff"],
                        marker_color=colors_bar2,
                        text=[f"{'+'if d>=0 else ''}{d}" for d in df_oo2["_diff"]],
                        textposition="outside"))
                    fig_oo2.add_hline(y=0, line_dash="solid", line_color="#e2e4e8")
                    fig_oo2.update_layout(
                        yaxis_title="On - Off Net Rating",
                        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                        font=dict(color="#374151", family="Inter"),
                        margin=dict(l=0,r=0,t=30,b=0), height=280)
                    st.plotly_chart(fig_oo2, use_container_width=True)
                    with st.expander("Veure detall complet"):
                        st.caption("On = quan la jugadora és a pista · Off = quan no hi és · ⚠️ = poques possessions Off, valor poc fiable")
                        # Usa taula HTML en lloc de st.dataframe per evitar problemes de CSS
                        df_show_oo = df_oo2.drop(columns=["_diff","_fiable"], errors="ignore")
                        cols_oo = list(df_show_oo.columns)
                        html_oo = '<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px;color:#1a2744">'
                        html_oo += '<tr>' + ''.join(f'<th style="background:#D6E8F7;color:#0C447C;padding:6px 10px;text-align:center;border:1px solid #B5D4F4;font-weight:600">{c}</th>' for c in cols_oo) + '</tr>'
                        for _, row_oo in df_show_oo.iterrows():
                            html_oo += '<tr>'
                            for ci_oo, col_oo in enumerate(cols_oo):
                                val_oo = row_oo[col_oo]
                                bg = '#ffffff' if ci_oo > 0 else '#EBF4FC'
                                align = 'left' if ci_oo == 0 else 'center'
                                html_oo += f'<td style="padding:5px 10px;border:1px solid #B5D4F4;background:{bg};color:#1a2744;text-align:{align}">{val_oo}</td>'
                            html_oo += '</tr>'
                        html_oo += '</table></div>'
                        st.markdown(html_oo, unsafe_allow_html=True)
                else:
                    st.info("No hi ha prou dades per calcular l'On/Off Rating.")

    # ── Exporta Excel ───────────────────────────────────────────────────────
    st.markdown(sec("Exporta a Excel"), unsafe_allow_html=True)
    st.caption("Excel amb totes les mètriques avançades de tots els partits de la base de dades.")
    if st.button("⬇ Descarregar Excel d'anàlisi complet", key="btn_excel_analisi"):
        excel_data = genera_excel_analisi()
        if excel_data:
            st.download_button(
                label="📥 Clic per descarregar",
                data=excel_data,
                file_name=f"miki_analisi_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_analisi"
            )
        else:
            st.info("No hi ha partits a la base de dades.")

with t5:
    # ══════════════════════════════════════════════════
    # PESTANYA ROTACIONS
    # ══════════════════════════════════════════════════

    # ── Càlcul de minuts reals per jugadora des de l'API ──────────────────
    def get_intervals_jugadores(df):
        """Retorna dict jugadora -> [(t_ini, t_fi, equip_id)] en minuts absoluts de partit."""
        MINS_Q = 10
        intervals = {}
        en_pista  = {}
        col_j = "jugador" if "jugador" in df.columns else "jugadora"

        # Detecta jugadores que comencen a pista (primer event és Surt sense Entra previ)
        for jug_x in df[col_j].unique():
            if not jug_x or str(jug_x) in ("","nan"): continue
            df_jx = df[df[col_j]==jug_x].sort_values("num")
            if df_jx.empty: continue
            primer_acc = str(df_jx.iloc[0].get("accio",""))
            primer_q   = int(df_jx.iloc[0].get("quart",1))
            primer_ei  = str(df_jx.iloc[0].get("idEquip",""))
            if "Surt" in primer_acc and "camp" in primer_acc:
                en_pista[jug_x] = ((primer_q-1)*MINS_Q, primer_ei)

        for _, row in df.sort_values("num").iterrows():
            jug = row.get(col_j, "")
            if not jug or str(jug) in ("","nan"): continue
            accio = str(row.get("accio",""))
            quart = int(row.get("quart",1)) if row.get("quart","") != "" else 1
            # min_num és el minut DINS del quart (cronòmetre enrere: 10→0)
            # El convertim a minut absolut: inici_quart + (MINS_Q - min_dins_quart)
            min_dins = float(row.get("min_num", 0))
            # Si min_num ja és absolut (>10), ho detectem
            if min_dins > MINS_Q:
                t_min = min_dins  # ja és absolut
            else:
                t_min = (quart-1)*MINS_Q + (MINS_Q - min_dins)
            # Clipa al rang vàlid
            t_min = max(0, min(t_min, quart * MINS_Q))
            eq_id = str(row.get("idEquip",""))
            if "Entra al camp" in accio:
                en_pista[jug] = (t_min, eq_id)
            elif "Surt del camp" in accio:
                ini_t = en_pista.pop(jug, ((quart-1)*MINS_Q, eq_id))
                t_ini_real = ini_t[0]; eq_real = ini_t[1]
                if t_min > t_ini_real:
                    intervals.setdefault(jug, []).append((t_ini_real, t_min, eq_real))
            elif "Final de període" in accio:
                fi = quart * MINS_Q
                for j, (ti, ei) in list(en_pista.items()):
                    if fi > ti:
                        intervals.setdefault(j, []).append((ti, fi, ei))
                en_pista = {}
        # Tanca intervals que han quedat oberts
        for j, (ti, ei) in en_pista.items():
            fi = df["quart"].max() * MINS_Q if not df.empty else 40
            if fi > ti:
                intervals.setdefault(j, []).append((ti, fi, ei))
        return intervals

    intervals_jug = get_intervals_jugadores(df_orig)
    score_df_rot = score_df.copy() if not score_df.empty else pd.DataFrame()

    # ── 1. Gràfic de quintets (Gantt de rotacions) ────────────────────────
    st.markdown(sec("Gràfic de rotacions — qui juga cada minut"), unsafe_allow_html=True)
    st.caption("Cada barra indica un tram de joc d'una jugadora. La línia vermella/blava és el ±parcial de l'equip.")

    eq_rot = st.selectbox("Equip", [nom_a, nom_b], key="rot_eq")
    tid_rot = teams[0] if eq_rot == nom_a else (teams[1] if len(teams)>1 else None)
    color_rot = COLOR_A if eq_rot == nom_a else COLOR_B
    rival_rot = teams[1] if eq_rot == nom_a else (teams[0] if teams else None)

    if tid_rot and intervals_jug:
        # Filtra jugadores de l'equip seleccionat
        jugs_rot = {j: ivs for j,ivs in intervals_jug.items()
                    if any(ei == tid_rot for _,_,ei in ivs)}

        if not jugs_rot:
            st.info("No s'han detectat events d'entrada/sortida per a aquest equip.")
        else:
            # Ordena per primer minut de joc
            jugs_sorted = sorted(jugs_rot.items(), key=lambda x: min(i[0] for i in x[1]))
            n_jugs = len(jugs_sorted)
            MINS_TOTAL = max(df_orig["quart"].max() * 10 if not df_orig.empty else 40, 40)

            fig_rot = go.Figure()

            # Barres de rotació per jugadora
            for yi, (jug, ivs) in enumerate(jugs_sorted):
                for (t_ini, t_fi, ei) in ivs:
                    if ei != tid_rot: continue
                    fig_rot.add_trace(go.Bar(
                        x=[t_fi - t_ini],
                        y=[jug],
                        base=[t_ini],
                        orientation='h',
                        marker_color=color_rot,
                        marker_opacity=0.75,
                        marker_line=dict(width=0.5, color='white'),
                        name=jug,
                        showlegend=False,
                        hovertemplate=f"{jug}<br>Minut {t_ini:.1f}–{t_fi:.1f}<br>Durada: {t_fi-t_ini:.1f} min<extra></extra>"
                    ))

            # Línies de parcial de l'equip (±)
            if not score_df_rot.empty:
                score_df_rot["t_min"] = score_df_rot.apply(
                    lambda r: (int(r["quart"])-1)*10 + r.get("min_num", 0)
                    if "min_num" in r else (int(r["quart"])-1)*10, axis=1)

                if "min_num" not in score_df_rot.columns:
                    # Estima minut des del número de jugada
                    score_df_rot["t_min"] = score_df_rot["num"] / score_df_rot["num"].max() * MINS_TOTAL

                parcial_eq  = score_df_rot["scoreA"] if tid_rot == teams[0] else score_df_rot["scoreB"]
                parcial_riv = score_df_rot["scoreB"] if tid_rot == teams[0] else score_df_rot["scoreA"]
                diff_parcial = parcial_eq - parcial_riv

                # Normalitza per mostrar com a línia sobre el gràfic
                d_max = max(abs(diff_parcial.max()), abs(diff_parcial.min()), 1)

                fig_rot.add_trace(go.Scatter(
                    x=score_df_rot["t_min"] if "t_min" in score_df_rot else score_df_rot.index,
                    y=diff_parcial,
                    mode="lines",
                    name="Parcial equip",
                    line=dict(color="#374151", width=1.5, dash="dot"),
                    yaxis="y2",
                    hovertemplate="Minut %{x:.1f}<br>Parcial: %{y:+d}<extra></extra>"
                ))

            # Línies de quart
            for q in range(1, 5):
                fig_rot.add_vline(x=q*10, line_dash="dot", line_color="#e2e4e8",
                    annotation_text=f"Fi Q{q}", annotation_font_size=9,
                    annotation_font_color="#9ca3af")

            fig_rot.update_layout(
                barmode="overlay",
                height=max(280, n_jugs * 36 + 80),
                paper_bgcolor="#ffffff", plot_bgcolor="#f9fafb",
                font=dict(color="#374151", family="Inter"),
                xaxis=dict(title="Minut de joc", range=[0, MINS_TOTAL],
                           showgrid=True, gridcolor="#f3f4f6", color="#9ca3af"),
                yaxis=dict(showgrid=False, color="#374151"),
                yaxis2=dict(overlaying="y", side="right", title="Parcial ±",
                            showgrid=False, zeroline=True, zerolinecolor="#e2e4e8",
                            color="#9ca3af"),
                margin=dict(l=0,r=60,t=20,b=40),
                legend=dict(bgcolor="#ffffff",bordercolor="#e2e4e8",borderwidth=1)
            )
            st.plotly_chart(fig_rot, use_container_width=True)

    # ── 2. +/- per minuts jugats ──────────────────────────────────────────
    st.markdown(sec("+/- per minut jugat"), unsafe_allow_html=True)
    st.caption("Parcial de l'equip per minut jugat per cada jugadora. Valors positius = l'equip guanya mentre juga.")

    if tid_rot and not score_df_rot.empty and intervals_jug:
        pm_rows = []
        for jug, ivs in intervals_jug.items():
            ivs_eq = [(ti,tf) for ti,tf,ei in ivs if ei==tid_rot]
            if not ivs_eq: continue
            total_min = sum(tf-ti for ti,tf in ivs_eq)
            if total_min < 0.5: continue

            # Usa el mateix mètode que la pestanya Jugadores:
            # compta punts del play-by-play durant els intervals
            rival_rot_id = teams[1] if tid_rot == teams[0] else (teams[0] if teams else None)
            parcial_favor = 0; parcial_contra = 0
            for t_ini, t_fi in ivs_eq:
                # Converteix minuts a nums de jugada
                df_interval = df_orig[
                    (df_orig["min_num"].apply(lambda m, q=0: m) >= 0)
                ]
                # Filtra per temps absolut
                df_orig_t = df_orig.copy()
                df_orig_t["t_abs"] = df_orig_t.apply(
                    lambda r: (int(r["quart"])-1)*10 + (10 - float(r["min_num"]))
                    if float(r.get("min_num",0)) <= 10
                    else float(r.get("min_num",0)), axis=1)
                df_interval = df_orig_t[
                    (df_orig_t["t_abs"] >= t_ini) &
                    (df_orig_t["t_abs"] <= t_fi)]
                parcial_favor  += int(df_interval[df_interval["idEquip"]==tid_rot]["punts"].sum())
                if rival_rot_id:
                    parcial_contra += int(df_interval[df_interval["idEquip"]==rival_rot_id]["punts"].sum())

            pm = parcial_favor - parcial_contra
            pm_per_min = round(pm / total_min, 2) if total_min > 0 else 0
            pm_rows.append({
                "Jugadora": jug,
                "Minuts": round(total_min, 1),
                "+/-": pm,
                "+/- per min": pm_per_min
            })

        if pm_rows:
            df_pm = pd.DataFrame(pm_rows).sort_values("+/- per min", ascending=False)

            # Gràfic scatter: X = minuts, Y = +/-
            fig_pm = go.Figure()
            colors_pm = ["#16a34a" if v >= 0 else "#dc2626" for v in df_pm["+/- per min"]]
            fig_pm.add_trace(go.Scatter(
                x=df_pm["Minuts"],
                y=df_pm["+/- per min"],
                mode="markers+text",
                marker=dict(size=12, color=colors_pm,
                            line=dict(width=1.5, color="white")),
                text=df_pm["Jugadora"].apply(lambda n: n.split()[1] if len(n.split())>1 else n),
                textposition="top center",
                textfont=dict(size=9),
                hovertemplate="%{text}<br>Minuts: %{x:.1f}<br>+/- per min: %{y:+.2f}<extra></extra>"
            ))
            fig_pm.add_hline(y=0, line_dash="solid", line_color="#e2e4e8")
            fig_pm.update_layout(
                xaxis=dict(title="Minuts jugats", showgrid=True, gridcolor="#f3f4f6", color="#9ca3af"),
                yaxis=dict(title="+/- per minut", showgrid=True, gridcolor="#f3f4f6", color="#9ca3af",
                           zeroline=True, zerolinecolor="#e2e4e8"),
                paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                font=dict(color="#374151", family="Inter"),
                margin=dict(l=0,r=0,t=20,b=0), height=320)
            st.plotly_chart(fig_pm, use_container_width=True)
            st.dataframe(df_pm, use_container_width=True, hide_index=True)

    # ── 3. +/- per tram de minuts jugats ─────────────────────────────────
    st.markdown(sec("+/- per tram de joc"), unsafe_allow_html=True)
    st.caption("Per cada tram que una jugadora juga seguit, veus el parcial de l'equip minut a minut.")

    if tid_rot and not score_df_rot.empty and intervals_jug:
        jugs_tram = [j for j,ivs in intervals_jug.items()
                     if any(ei==tid_rot for _,_,ei in ivs)]
        jug_tram = st.selectbox("Jugadora", sorted(jugs_tram), key="jug_tram")

        if jug_tram and jug_tram in intervals_jug:
            ivs_jug = [(ti,tf) for ti,tf,ei in intervals_jug[jug_tram] if ei==tid_rot]

            fig_tram = go.Figure()
            colors_tram = [color_rot, "#16a34a", "#d97706", "#6366f1", "#ec4899"]

            for bi, (t_ini, t_fi) in enumerate(ivs_jug):
                if "t_min" in score_df_rot.columns:
                    df_tr = score_df_rot[
                        (score_df_rot["t_min"] >= t_ini) &
                        (score_df_rot["t_min"] <= t_fi)].copy()
                    x_vals = df_tr["t_min"] - t_ini
                else:
                    n_ini = int(t_ini / MINS_TOTAL * len(score_df_rot))
                    n_fi  = int(t_fi  / MINS_TOTAL * len(score_df_rot))
                    df_tr = score_df_rot.iloc[n_ini:n_fi].copy()
                    x_vals = pd.Series(range(len(df_tr))) * (t_fi-t_ini) / max(len(df_tr),1)

                if df_tr.empty: continue

                if tid_rot == teams[0]:
                    parcial_tr = df_tr["scoreA"] - df_tr["scoreA"].iloc[0] -                                  (df_tr["scoreB"] - df_tr["scoreB"].iloc[0])
                else:
                    parcial_tr = df_tr["scoreB"] - df_tr["scoreB"].iloc[0] -                                  (df_tr["scoreA"] - df_tr["scoreA"].iloc[0])

                color_bi = colors_tram[bi % len(colors_tram)]
                fig_tram.add_trace(go.Scatter(
                    x=x_vals,
                    y=parcial_tr,
                    mode="lines+markers",
                    name=f"Tram {bi+1} (min {t_ini:.0f}–{t_fi:.0f})",
                    line=dict(color=color_bi, width=2),
                    marker=dict(size=6, color=color_bi),
                    hovertemplate="Min %{x:.1f} del tram<br>Parcial: %{y:+d}<extra></extra>"
                ))

            fig_tram.add_hline(y=0, line_dash="solid", line_color="#e2e4e8")
            fig_tram.update_layout(
                xaxis=dict(title="Minuts dins del tram", showgrid=True,
                           gridcolor="#f3f4f6", color="#9ca3af"),
                yaxis=dict(title="Parcial ±", showgrid=True,
                           gridcolor="#f3f4f6", color="#9ca3af",
                           zeroline=True, zerolinecolor="#e2e4e8"),
                paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                font=dict(color="#374151", family="Inter"),
                legend=dict(bgcolor="#ffffff", bordercolor="#e2e4e8", borderwidth=1,
                            orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                margin=dict(l=0,r=0,t=40,b=0), height=300)
            st.plotly_chart(fig_tram, use_container_width=True)

    # ── 4. Mapa de calor de parelles ─────────────────────────────────────
    st.markdown(sec("Mapa de calor +/- per parelles"), unsafe_allow_html=True)
    st.caption("Color de cada casella = +/- conjunt de la parella. Verd = l'equip guanya quan juguen juntes.")

    if tid_rot and intervals_jug:
        jugs_eq_all = sorted([j for j,ivs in intervals_jug.items()
                              if any(ei==tid_rot for _,_,ei in ivs)])
        rival_rot_id2 = teams[1] if tid_rot == teams[0] else (teams[0] if teams else None)

        if len(jugs_eq_all) >= 2:
            # Precalcula t_abs per eficiència
            df_orig_tab = df_orig.copy()
            df_orig_tab["t_abs"] = df_orig_tab.apply(
                lambda r: (int(r["quart"])-1)*10 + (10 - float(r["min_num"]))
                if float(r.get("min_num",0)) <= 10
                else float(r.get("min_num",0)), axis=1)

            def pm_parella(j1, j2):
                ivs1 = [(ti,tf) for ti,tf,ei in intervals_jug.get(j1,[]) if ei==tid_rot]
                ivs2 = [(ti,tf) for ti,tf,ei in intervals_jug.get(j2,[]) if ei==tid_rot]
                juntes = []
                for a1,a2 in ivs1:
                    for b1,b2 in ivs2:
                        ini=max(a1,b1); fi=min(a2,b2)
                        if fi > ini: juntes.append((ini,fi))
                if not juntes: return None
                pf=pc=0
                for ti,tf in juntes:
                    df_j = df_orig_tab[(df_orig_tab["t_abs"]>=ti)&(df_orig_tab["t_abs"]<=tf)]
                    pf += int(df_j[df_j["idEquip"]==tid_rot]["punts"].sum())
                    if rival_rot_id2:
                        pc += int(df_j[df_j["idEquip"]==rival_rot_id2]["punts"].sum())
                return pf - pc

            # Construeix matriu
            n = len(jugs_eq_all)
            matrix = [[None]*n for _ in range(n)]
            for i,j1 in enumerate(jugs_eq_all):
                for j,j2 in enumerate(jugs_eq_all):
                    if i == j:
                        matrix[i][j] = 0
                    elif i < j:
                        val = pm_parella(j1, j2)
                        matrix[i][j] = val
                        matrix[j][i] = val

            # Noms curts per a l'eix
            noms_curts = [n.split()[1] if len(n.split())>1 else n for n in jugs_eq_all]

            # Substitueix None per 0 per al heatmap
            z_vals = [[v if v is not None else 0 for v in row] for row in matrix]
            text_vals = [[f"+{v}" if v is not None and v > 0
                          else (str(v) if v is not None else "—")
                          for v in row] for row in matrix]

            fig_hm = go.Figure(go.Heatmap(
                z=z_vals,
                x=noms_curts,
                y=noms_curts,
                text=text_vals,
                texttemplate="%{text}",
                textfont=dict(size=11, color="white"),
                colorscale=[
                    [0.0, "#dc2626"],
                    [0.5, "#f9fafb"],
                    [1.0, "#16a34a"]
                ],
                zmid=0,
                showscale=True,
                colorbar=dict(title="+/-", thickness=12)
            ))
            fig_hm.update_layout(
                height=max(320, n*48+80),
                paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                font=dict(color="#374151", family="Inter", size=11),
                xaxis=dict(side="top", tickangle=-30),
                margin=dict(l=0,r=60,t=60,b=0)
            )
            st.plotly_chart(fig_hm, use_container_width=True)
            st.caption("Diagonal = la pròpia jugadora (0). Caselles buides = no han jugat juntes.")
        else:
            st.info("Cal tenir almenys 2 jugadores amb events d'entrada/sortida.")

    # ── 5. +/- per parella (detall) ───────────────────────────────────────
    st.markdown(sec("+/- per parella de jugadores (detall)"), unsafe_allow_html=True)
    st.caption("Parcial de l'equip durant els minuts que les dues jugadores seleccionades han jugat juntes.")

    if tid_rot and intervals_jug:
        jugs_par = [j for j,ivs in intervals_jug.items()
                    if any(ei==tid_rot for _,_,ei in ivs)]
        col_p1, col_p2 = st.columns(2)
        with col_p1: jug_p1 = st.selectbox("Jugadora 1", sorted(jugs_par), key="par_j1")
        with col_p2: jug_p2 = st.selectbox("Jugadora 2",
            [j for j in sorted(jugs_par) if j != jug_p1], key="par_j2")

        if jug_p1 and jug_p2:
            ivs1 = [(ti,tf) for ti,tf,ei in intervals_jug.get(jug_p1,[]) if ei==tid_rot]
            ivs2 = [(ti,tf) for ti,tf,ei in intervals_jug.get(jug_p2,[]) if ei==tid_rot]

            # Troba intervals on les dues juguen juntes
            juntes = []
            for a1,a2 in ivs1:
                for b1,b2 in ivs2:
                    ini = max(a1,b1); fi = min(a2,b2)
                    if fi > ini: juntes.append((ini,fi))

            if not juntes:
                st.info(f"{jug_p1} i {jug_p2} no han jugat juntes en aquest partit.")
            else:
                total_min_j = sum(f-i for i,f in juntes)
                parcial_j = 0
                for t_ini,t_fi in juntes:
                    if "t_min" in score_df_rot.columns:
                        df_j = score_df_rot[(score_df_rot["t_min"]>=t_ini)&(score_df_rot["t_min"]<=t_fi)]
                    else:
                        n1=int(t_ini/MINS_TOTAL*len(score_df_rot))
                        n2=int(t_fi/MINS_TOTAL*len(score_df_rot))
                        df_j = score_df_rot.iloc[n1:n2]
                    if not df_j.empty:
                        if tid_rot == teams[0]:
                            parcial_j += (df_j["scoreA"].iloc[-1]-df_j["scoreA"].iloc[0]) -                                          (df_j["scoreB"].iloc[-1]-df_j["scoreB"].iloc[0])
                        else:
                            parcial_j += (df_j["scoreB"].iloc[-1]-df_j["scoreB"].iloc[0]) -                                          (df_j["scoreA"].iloc[-1]-df_j["scoreA"].iloc[0])

                c1,c2,c3 = st.columns(3)
                col_p = "#16a34a" if parcial_j >= 0 else "#dc2626"
                with c1: st.markdown(card("Minuts juntes", round(total_min_j,1), "min", color_rot), unsafe_allow_html=True)
                with c2: st.markdown(card("Parcial", f"{'+'if parcial_j>=0 else ''}{parcial_j}", "", col_p), unsafe_allow_html=True)
                with c3: st.markdown(card("+/- per min",
                    f"{'+'if parcial_j>=0 else ''}{round(parcial_j/total_min_j,2) if total_min_j>0 else 0}",
                    "", col_p), unsafe_allow_html=True)
                st.caption(f"Trams juntes: {', '.join([f'{i:.0f}–{f:.0f} min' for i,f in juntes])}")

with t6:
    st.markdown(sec("Rànquing acumulat"), unsafe_allow_html=True)
    df_sj=load_stats_jugador_db()
    if df_sj.empty:
        st.info("Carrega partits per generar l'històric de jugadores.")
    else:
        df_pr=load_partits_db()
        def label_p(mid):
            r=df_pr[df_pr["match_id"]==mid]
            if r.empty: return mid[:10]+"..."
            return f"{r.iloc[0]['nom_a']} vs {r.iloc[0]['nom_b']} ({r.iloc[0]['data_consulta'][:10]})"
        df_sj["Partit"]=df_sj["match_id"].apply(label_p)
        df_sj["Data"]=df_sj["data_consulta"]

        ranking=df_sj.groupby(["jugador","equip_nom"]).agg(
            Partits=("match_id","nunique"),Punts=("punts","sum"),
            C2=("cistelles_2","sum"),C3=("cistelles_3","sum"),
            TL=("tirs_lliures","sum"),Faltes=("faltes","sum"),Impacte=("impacte","sum"),
        ).reset_index()
        ranking["Pts/p"]=(ranking["Punts"]/ranking["Partits"]).round(1)
        ranking["Imp/p"]=(ranking["Impacte"]/ranking["Partits"]).round(1)
        ranking=ranking.sort_values("Punts",ascending=False).rename(columns={"jugador":"Jugadora","equip_nom":"Equip"})

        equips_hist=["Tots"]+sorted(df_sj["equip_nom"].unique().tolist())
        eq_rank=st.selectbox("Filtra per equip",equips_hist,key="eq_rank")
        df_rk=ranking if eq_rank=="Tots" else ranking[ranking["Equip"]==eq_rank]
        # Afegir minuts totals si existeix la columna
        cols_rank = ["Jugadora","Equip","Partits","Punts","Pts/p","C2","C3","TL","Faltes","Impacte","Imp/p"]
        if "minuts" in ranking.columns:
            ranking["Min/p"] = (ranking.get("minuts",0) / ranking["Partits"]).round(1)
            cols_rank = ["Jugadora","Equip","Partits","Punts","Pts/p","C2","C3","TL","Faltes","Min/p","Impacte","Imp/p"]
        # Afegir eficiència de tir al rànquing
        if "minuts" in ranking.columns:
            ranking["Ef2%"] = (ranking.get("C2",0) /
                (ranking.get("C2",0) + df_sj.groupby(["jugador","equip_nom"])["cistelles_2"].sum().reset_index()["cistelles_2"] * 0 + 1)
            ).round(0)

        st.dataframe(df_rk[cols_rank],
            use_container_width=True,hide_index=True)

        # Tirs ficats/tirats per tipus
        st.markdown(sec("Eficiència de tir — ficats/tirats"), unsafe_allow_html=True)
        st.caption("Exemple: 12/17 vol dir 12 cistelles de 17 intents.")

        df_sz_rank = load_shots_zones_db()
        if not df_sz_rank.empty:
            # Agrega per jugadora tots els partits
            df_sz_agg = df_sz_rank[df_sz_rank["jugador"]!="__equip__"].groupby(["jugador","equip_nom"]).agg(
                v1m=("val1_made","sum"), v1x=("val1_miss","sum"),
                v2m=("val2_made","sum"), v2x=("val2_miss","sum"),
                v3m=("val3_made","sum"), v3x=("val3_miss","sum"),
            ).reset_index()

            def fmt_ratio(made, miss):
                total = made + miss
                ef = round(made/total*100) if total > 0 else 0
                color = "#16a34a" if ef >= 55 else ("#d97706" if ef >= 35 else "#dc2626")
                return f"{made}/{total} ({ef}%)", color

            # Filtra per equip
            eq_tir = st.selectbox("Equip", ["Tots"] + sorted(df_sz_agg["equip_nom"].unique().tolist()), key="eq_tir_rank")
            df_sz_show = df_sz_agg if eq_tir == "Tots" else df_sz_agg[df_sz_agg["equip_nom"]==eq_tir]
            df_sz_show = df_sz_show.copy()
            df_sz_show["TL (1pt)"]  = df_sz_show.apply(lambda r: f"{r.v1m}/{r.v1m+r.v1x} ({round(r.v1m/(r.v1m+r.v1x)*100) if (r.v1m+r.v1x)>0 else 0}%)", axis=1)
            df_sz_show["2pts"]      = df_sz_show.apply(lambda r: f"{r.v2m}/{r.v2m+r.v2x} ({round(r.v2m/(r.v2m+r.v2x)*100) if (r.v2m+r.v2x)>0 else 0}%)", axis=1)
            df_sz_show["3pts"]      = df_sz_show.apply(lambda r: f"{r.v3m}/{r.v3m+r.v3x} ({round(r.v3m/(r.v3m+r.v3x)*100) if (r.v3m+r.v3x)>0 else 0}%)", axis=1)
            df_sz_show["Total"]     = df_sz_show.apply(lambda r: f"{r.v1m+r.v2m+r.v3m}/{r.v1m+r.v1x+r.v2m+r.v2x+r.v3m+r.v3x} ({round((r.v1m+r.v2m+r.v3m)/(r.v1m+r.v1x+r.v2m+r.v2x+r.v3m+r.v3x)*100) if (r.v1m+r.v1x+r.v2m+r.v2x+r.v3m+r.v3x)>0 else 0}%)", axis=1)
            df_sz_show = df_sz_show.sort_values("v2m", ascending=False)
            st.dataframe(df_sz_show[["jugador","equip_nom","TL (1pt)","2pts","3pts","Total"]].rename(
                columns={"jugador":"Jugadora","equip_nom":"Equip"}),
                use_container_width=True, hide_index=True)
        else:
            st.info("Consulta més partits per veure les estadístiques de tir.")

        top5=df_rk.head(5)
        if not top5.empty:
            fig_rank=px.bar(top5,x="Jugadora",y="Punts",color="Equip",
                color_discrete_map=dict(zip(df_sj["equip_nom"].unique(),[COLOR_A,COLOR_B])),
                text="Punts")
            fig_rank.update_traces(textposition="outside")
            st.plotly_chart(chart_style(fig_rank,240,"Top 5 anotadores acumulades"),use_container_width=True)

        st.markdown(sec("Evolució d'una jugadora per partit"), unsafe_allow_html=True)
        tots_jugs_hist=sorted(df_sj["jugador"].unique().tolist())
        jug_hist=st.selectbox("Selecciona jugadora",tots_jugs_hist,key="jug_hist")
        if jug_hist:
            df_jh=df_sj[df_sj["jugador"]==jug_hist].sort_values("Data")
            if df_jh.empty:
                st.info("Sense dades.")
            else:
                c1,c2,c3,c4=st.columns(4)
                with c1: st.markdown(card("Partits",len(df_jh),"","#374151"),unsafe_allow_html=True)
                with c2: st.markdown(card("Punts totals",int(df_jh["punts"].sum()),"",COLOR_A),unsafe_allow_html=True)
                with c3: st.markdown(card("Pts/partit",f"{df_jh['punts'].mean():.1f}","mitjana",COLOR_A),unsafe_allow_html=True)
                with c4:
                    imp=int(df_jh["impacte"].sum())
                    ci="#16a34a" if imp>=0 else "#dc2626"
                    st.markdown(card("Impacte total",f"{'+'if imp>=0 else ''}{imp}","acumulat",ci),unsafe_allow_html=True)

                tcol_jug=COLOR_A
                fig_evo=go.Figure()
                fig_evo.add_trace(go.Scatter(x=df_jh["Partit"],y=df_jh["punts"],
                    mode="lines+markers",name="Punts",
                    line=dict(color=tcol_jug,width=2.5),marker=dict(size=8),
                    fill="tozeroy",fillcolor="rgba(24,95,165,0.07)"))
                fig_evo.add_trace(go.Scatter(x=df_jh["Partit"],y=df_jh["punts"].expanding().mean(),
                    mode="lines",name="Mitjana",line=dict(color="#d97706",width=2,dash="dot")))
                fig_evo.update_xaxes(tickangle=-30)
                st.plotly_chart(chart_style(fig_evo,280,f"{jug_hist} — punts per partit"),use_container_width=True)

                ci_list=["#16a34a" if v>=0 else "#dc2626" for v in df_jh["impacte"]]
                fig_imp=go.Figure()
                fig_imp.add_trace(go.Scatter(x=df_jh["Partit"],y=df_jh["impacte"],
                    mode="lines+markers",name="Impacte",
                    line=dict(color="#6366f1",width=2.5),
                    marker=dict(size=9,color=ci_list,line=dict(width=1,color="#fff"))))
                fig_imp.add_hline(y=0,line_dash="solid",line_color="#e2e4e8")
                fig_imp.update_xaxes(tickangle=-30)
                st.plotly_chart(chart_style(fig_imp,240,f"{jug_hist} — impacte per partit"),use_container_width=True)

                fig_cist=go.Figure()
                fig_cist.add_trace(go.Scatter(x=df_jh["Partit"],y=df_jh["cistelles_2"],
                    mode="lines+markers",name="C2",line=dict(color=COLOR_A,width=2),marker=dict(size=7)))
                fig_cist.add_trace(go.Scatter(x=df_jh["Partit"],y=df_jh["cistelles_3"],
                    mode="lines+markers",name="C3",line=dict(color="#16a34a",width=2),marker=dict(size=7)))
                fig_cist.add_trace(go.Scatter(x=df_jh["Partit"],y=df_jh["tirs_lliures"],
                    mode="lines+markers",name="TL",line=dict(color="#d97706",width=2,dash="dot"),marker=dict(size=7)))
                fig_cist.update_xaxes(tickangle=-30)
                st.plotly_chart(chart_style(fig_cist,240,f"{jug_hist} — tipus cistelles per partit"),use_container_width=True)

                with st.expander("Taula completa"):
                    _df_tc = df_jh[["Partit","punts","cistelles_2","cistelles_3","tirs_lliures","faltes","impacte","pts_per_min"]].rename(
                        columns={"punts":"Pts","cistelles_2":"C2","cistelles_3":"C3","tirs_lliures":"TL",
                                 "faltes":"Faltes","impacte":"Impacte","pts_per_min":"Pts/min"})
                    _cols_tc = list(_df_tc.columns)
                    _html_tc = '<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px;color:#1a2744">'
                    _html_tc += '<tr>' + ''.join(f'<th style="background:#D6E8F7;color:#0C447C;padding:6px 10px;text-align:center;border:1px solid #B5D4F4;font-weight:600">{c}</th>' for c in _cols_tc) + '</tr>'
                    for _, _row_tc in _df_tc.iterrows():
                        _html_tc += '<tr>'
                        for _ci_tc, _col_tc in enumerate(_cols_tc):
                            _bg_tc = '#EBF4FC' if _ci_tc == 0 else '#ffffff'
                            _align_tc = 'left' if _ci_tc == 0 else 'center'
                            _html_tc += f'<td style="padding:5px 10px;border:1px solid #B5D4F4;background:{_bg_tc};color:#1a2744;text-align:{_align_tc}">{_row_tc[_col_tc]}</td>'
                        _html_tc += '</tr>'
                    _html_tc += '</table></div>'
                    st.markdown(_html_tc, unsafe_allow_html=True)

        # ── Rendiment per bloc de minuts ────────────────────────────────────────
        st.markdown(sec("Rendiment per bloc de minuts — primers vs últims"), unsafe_allow_html=True)
        st.caption("Compara si la jugadora anota més al principi o al final de cada bloc de minuts que juga.")

        BLOC_MINS = 3  # minuts a considerar com 'principi' i 'final' del bloc

        jug_bloc = st.selectbox("Jugadora", tots_jugs_hist, key="jug_bloc")
        if jug_bloc:
            # Agafem totes les jugades d'aquesta jugadora de tots els partits
            con_bloc = sqlite3.connect(DB_PATH)
            df_bloc = pd.read_sql(
                "SELECT * FROM jugades WHERE jugador=? ORDER BY match_id, num",
                con_bloc, params=(jug_bloc,))
            con_bloc.close()

            if df_bloc.empty:
                st.info("Sense dades de play-by-play per a aquesta jugadora.")
            else:
                # Per cada partit, identifica blocs de joc continus
                bloc_rows = []
                for mid, df_mid in df_bloc.groupby("match_id"):
                    df_mid = df_mid.sort_values("min_num")
                    # Detecta ruptures de bloc (>2 min sense acció = fora de pista)
                    df_mid["gap"] = df_mid["min_num"].diff().fillna(0)
                    df_mid["bloc_id"] = (df_mid["gap"] > 2).cumsum()

                    for bloc_id, df_b in df_mid.groupby("bloc_id"):
                        if len(df_b) < 2: continue
                        min_inici = df_b["min_num"].min()
                        min_fi    = df_b["min_num"].max()
                        durada    = min_fi - min_inici
                        if durada < BLOC_MINS * 2: continue  # bloc massa curt

                        # Primers N minuts del bloc
                        df_primers = df_b[df_b["min_num"] <= min_inici + BLOC_MINS]
                        pts_primers = int(df_primers["punts"].sum())

                        # Últims N minuts del bloc
                        df_ultims = df_b[df_b["min_num"] >= min_fi - BLOC_MINS]
                        pts_ultims = int(df_ultims["punts"].sum())

                        # Etiqueta del partit
                        df_pr_b = load_partits_db()
                        row_p = df_pr_b[df_pr_b["match_id"]==mid]
                        lbl = f"{row_p.iloc[0]['nom_a']} vs {row_p.iloc[0]['nom_b']}" if not row_p.empty else mid[:8]

                        bloc_rows.append({
                            "Partit": lbl,
                            "Bloc": f"Bloc {int(bloc_id)+1}",
                            "Durada (min)": round(durada, 1),
                            f"Pts primers {BLOC_MINS} min": pts_primers,
                            f"Pts últims {BLOC_MINS} min": pts_ultims,
                            "Tendència": "📈 Millora" if pts_ultims > pts_primers
                                         else ("📉 Baixa" if pts_ultims < pts_primers
                                               else "➡️ Estable")
                        })

                if bloc_rows:
                    df_blocs = pd.DataFrame(bloc_rows)
                    st.dataframe(df_blocs, use_container_width=True, hide_index=True)

                    # Resum global
                    col_b1, col_b2, col_b3 = st.columns(3)
                    mit_p = df_blocs[f"Pts primers {BLOC_MINS} min"].mean()
                    mit_u = df_blocs[f"Pts últims {BLOC_MINS} min"].mean()
                    tendencia = "📈 Millora al final" if mit_u > mit_p else ("📉 Baixa al final" if mit_u < mit_p else "➡️ Estable")
                    with col_b1: st.markdown(card(f"Pts/bloc inici",f"{mit_p:.1f}","mitjana","#185FA5"),unsafe_allow_html=True)
                    with col_b2: st.markdown(card(f"Pts/bloc final",f"{mit_u:.1f}","mitjana","#185FA5"),unsafe_allow_html=True)
                    with col_b3: st.markdown(card("Tendència global",tendencia,"","#374151"),unsafe_allow_html=True)

                    fig_bloc = go.Figure()
                    fig_bloc.add_trace(go.Bar(name=f"Primers {BLOC_MINS} min",
                        x=df_blocs["Partit"]+" "+df_blocs["Bloc"],
                        y=df_blocs[f"Pts primers {BLOC_MINS} min"],
                        marker_color=COLOR_A, opacity=0.8))
                    fig_bloc.add_trace(go.Bar(name=f"Últims {BLOC_MINS} min",
                        x=df_blocs["Partit"]+" "+df_blocs["Bloc"],
                        y=df_blocs[f"Pts últims {BLOC_MINS} min"],
                        marker_color="#16a34a", opacity=0.8))
                    fig_bloc.update_layout(barmode="group")
                    fig_bloc.update_xaxes(tickangle=-30)
                    st.plotly_chart(chart_style(fig_bloc,260,f"{jug_bloc} — primers vs últims minuts del bloc"),use_container_width=True)
                else:
                    st.info(f"No hi ha blocs de més de {BLOC_MINS*2} minuts per a aquesta jugadora.")

    # ── 1. +/- per minut acumulat de temporada ──────────────────────────────
    st.markdown(sec("+/- per minut acumulat — temporada"), unsafe_allow_html=True)
    st.caption("Parcial de l'equip per minut jugat, acumulat de tots els partits.")

    jug_pm = st.selectbox("Jugadora", tots_jugs_hist, key="jug_pm_acum")

    if jug_pm:
        pm_acum_rows = []
        for mid_ac, df_ac in [(p, load_jugades_db(p)) for p in load_partits_db()["match_id"].tolist()]:
            if df_ac.empty: continue
            teams_ac = get_teams_ordered(df_ac)
            tn_ac = {}
            df_pr_ac = load_partits_db()
            p_ac = df_pr_ac[df_pr_ac["match_id"]==mid_ac]
            if not p_ac.empty:
                if len(teams_ac)>=1: tn_ac[teams_ac[0]] = p_ac.iloc[0]["nom_a"]
                if len(teams_ac)>=2: tn_ac[teams_ac[1]] = p_ac.iloc[0]["nom_b"]

            col_jac = "jugador" if "jugador" in df_ac.columns else "jugadora"
            df_ac["jugador"] = df_ac[col_jac].fillna("")
            if jug_pm not in df_ac["jugador"].values: continue

            # Calcula intervals reals
            MINS_Q_AC = 10
            intervals_ac = []; ep_ac = {}
            dj_ac = df_ac[df_ac["jugador"]==jug_pm]
            primer_ac = dj_ac.sort_values("num").iloc[0]
            if "Surt" in str(primer_ac.get("accio","")) and "camp" in str(primer_ac.get("accio","")):
                ep_ac[jug_pm] = (int(primer_ac.get("quart",1))-1)*MINS_Q_AC

            for _,row_ac in df_ac.sort_values("num").iterrows():
                if str(row_ac.get("jugador","")) != jug_pm: continue
                acc_ac = str(row_ac.get("accio",""))
                q_ac = int(row_ac.get("quart",1))
                m_ac = float(row_ac.get("min_num",0))
                t_ac = (q_ac-1)*MINS_Q_AC + (MINS_Q_AC-m_ac if m_ac<=MINS_Q_AC else m_ac)
                t_ac = max(0, min(t_ac, q_ac*MINS_Q_AC))
                if "Entra" in acc_ac and "camp" in acc_ac: ep_ac[jug_pm]=t_ac
                elif "Surt" in acc_ac and "camp" in acc_ac:
                    ti_ac=ep_ac.pop(jug_pm,(q_ac-1)*MINS_Q_AC)
                    if t_ac>ti_ac: intervals_ac.append((float(ti_ac),float(t_ac)))
                elif "Final de període" in acc_ac:
                    if jug_pm in ep_ac:
                        ti_ac=ep_ac.pop(jug_pm); fi_ac=float(q_ac*MINS_Q_AC)
                        if fi_ac>ti_ac: intervals_ac.append((ti_ac,fi_ac))
            for ti_o in ep_ac.values():
                fi_o=float(df_ac["quart"].max()*MINS_Q_AC)
                if fi_o>ti_o: intervals_ac.append((float(ti_o),fi_o))
            intervals_ac=[(float(ti),float(tf)) for ti,tf in intervals_ac]
            if not intervals_ac: continue

            total_min_ac = sum(tf-ti for ti,tf in intervals_ac)
            if total_min_ac < 1: continue

            # Calcula +/-
            eq_id_ac = dj_ac["idEquip"].iloc[0]
            rival_ac = [t for t in teams_ac if t!=eq_id_ac]
            rival_id_ac = rival_ac[0] if rival_ac else None
            df_t_ac = df_ac.copy()
            df_t_ac["t_abs"] = df_t_ac.apply(
                lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
                if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)
            pf_ac=pc_ac=0
            for ti_r,tf_r in intervals_ac:
                df_i_ac=df_t_ac[(df_t_ac["t_abs"]>=ti_r)&(df_t_ac["t_abs"]<=tf_r)]
                pf_ac+=int(df_i_ac[df_i_ac["idEquip"]==eq_id_ac]["punts"].sum())
                if rival_id_ac:
                    pc_ac+=int(df_i_ac[df_i_ac["idEquip"]==rival_id_ac]["punts"].sum())
            pm_val = pf_ac-pc_ac
            pm_per_min = round(pm_val/total_min_ac,3)

            lbl_ac = f"{tn_ac.get(eq_id_ac,'?')} vs {tn_ac.get(rival_id_ac,'?')}" if rival_id_ac else mid_ac[:8]
            pm_acum_rows.append({
                "Partit": lbl_ac,
                "match_id": mid_ac,
                "Minuts": round(total_min_ac,1),
                "+/-": pm_val,
                "+/- per min": pm_per_min,
            })

        if pm_acum_rows:
            df_pm_acum = pd.DataFrame(pm_acum_rows)
            mitj_pm = df_pm_acum["+/- per min"].mean()

            # Pearson entre Minuts i +/- per minut
            if len(df_pm_acum) >= 3:
                from scipy import stats as sp_stats
                rho, pval = sp_stats.pearsonr(df_pm_acum["Minuts"], df_pm_acum["+/- per min"])
                rho_txt = f"ρ = {rho:+.3f}"
                if abs(rho) >= 0.7: rho_interp = "correlació forta"
                elif abs(rho) >= 0.4: rho_interp = "correlació moderada"
                else: rho_interp = "correlació feble"
                rho_dir = "positiva" if rho > 0 else "negativa"
                pval_txt = f"p = {pval:.3f}" + (" (sig.)" if pval < 0.05 else " (no sig.)")
            else:
                rho, pval = None, None
                rho_txt = "Cal ≥3 partits"
                rho_interp = ""; rho_dir = ""; pval_txt = ""

            col_pa1,col_pa2,col_pa3,col_pa4 = st.columns(4)
            color_pm = "#16a34a" if mitj_pm>=0 else "#dc2626"
            color_rho = "#16a34a" if (rho or 0)>0.4 else ("#dc2626" if (rho or 0)<-0.4 else "#374151")
            with col_pa1: st.markdown(card("Partits analitzats",len(df_pm_acum),"",COLOR_A),unsafe_allow_html=True)
            with col_pa2: st.markdown(card("Minuts totals",round(df_pm_acum["Minuts"].sum(),1),"",COLOR_A),unsafe_allow_html=True)
            with col_pa3: st.markdown(card("+/- per min (mitj.)",f"{'+'if mitj_pm>=0 else ''}{mitj_pm:.3f}","",color_pm),unsafe_allow_html=True)
            with col_pa4: st.markdown(card("Pearson (Min vs +/-/min)",rho_txt,f"{rho_interp} {rho_dir}" if rho is not None else "",color_rho),unsafe_allow_html=True)

            if rho is not None:
                st.caption(f"ρ de Pearson (Minuts jugats vs +/- per minut): **{rho_txt}** — {rho_interp} {rho_dir} · {pval_txt}")

            # Gràfic barres +/- per minut per partit
            colors_pm_acum = ["#16a34a" if v>=0 else "#dc2626" for v in df_pm_acum["+/- per min"]]
            fig_pm_acum = go.Figure()
            fig_pm_acum.add_trace(go.Bar(
                x=df_pm_acum["Partit"],
                y=df_pm_acum["+/- per min"],
                marker_color=colors_pm_acum,
                text=[f"{'+'if v>=0 else ''}{v:.3f}" for v in df_pm_acum["+/- per min"]],
                textposition="outside"))
            fig_pm_acum.add_hline(y=0, line_dash="solid", line_color="#e2e4e8")
            fig_pm_acum.add_hline(y=mitj_pm, line_dash="dot", line_color="#d97706",
                annotation_text=f"Mitjana: {mitj_pm:+.3f}", annotation_font_size=10)
            fig_pm_acum.update_layout(xaxis_tickangle=-30)
            st.plotly_chart(chart_style(fig_pm_acum,280,f"{jug_pm} — +/- per minut (temporada)"),use_container_width=True)

            # Scatter Minuts vs +/- per minut amb línia de tendència
            if len(df_pm_acum) >= 3:
                import numpy as np
                fig_scatter_pm = go.Figure()
                fig_scatter_pm.add_trace(go.Scatter(
                    x=df_pm_acum["Minuts"], y=df_pm_acum["+/- per min"],
                    mode="markers+text",
                    marker=dict(size=10, color=colors_pm_acum, line=dict(width=1, color="white")),
                    text=df_pm_acum["Partit"].apply(lambda s: s[:12]),
                    textposition="top center", textfont=dict(size=8),
                    hovertemplate="<b>%{text}</b><br>Minuts: %{x}<br>+/- per min: %{y:+.3f}<extra></extra>"
                ))
                # Línia de tendència
                m, b = np.polyfit(df_pm_acum["Minuts"], df_pm_acum["+/- per min"], 1)
                x_line = [df_pm_acum["Minuts"].min(), df_pm_acum["Minuts"].max()]
                y_line = [m*x+b for x in x_line]
                fig_scatter_pm.add_trace(go.Scatter(
                    x=x_line, y=y_line, mode="lines",
                    line=dict(color="#d97706", width=2, dash="dot"),
                    name=f"Tendència (ρ={rho:+.3f})", showlegend=True
                ))
                fig_scatter_pm.add_hline(y=0, line_dash="solid", line_color="#e2e4e8")
                fig_scatter_pm.update_xaxes(title="Minuts jugats")
                fig_scatter_pm.update_yaxes(title="+/- per minut")
                st.plotly_chart(chart_style(fig_scatter_pm, 260,
                    f"{jug_pm} — Minuts vs +/- per minut (ρ = {rho:+.3f})"),
                    use_container_width=True)

            st.dataframe(df_pm_acum[["Partit","Minuts","+/-","+/- per min"]],
                use_container_width=True, hide_index=True)
        else:
            st.info("No hi ha prou dades per a aquesta jugadora.")

    # ── 2. Rendiment per posició dins del tram ──────────────────────────────
    st.markdown(sec("En quins minuts del tram aporta?"), unsafe_allow_html=True)
    st.caption("Parcial de l'equip al minut 1, 2, 3... de cada tram que juga la jugadora. Acumulat de tots els partits.")

    jug_tram_acum = st.selectbox("Jugadora", tots_jugs_hist, key="jug_tram_acum")

    if jug_tram_acum:
        tram_minuts = {}  # {minut_dins_tram: [parcials]}

        for mid_tr, df_tr in [(p, load_jugades_db(p)) for p in load_partits_db()["match_id"].tolist()]:
            if df_tr.empty: continue
            teams_tr = get_teams_ordered(df_tr)
            col_jtr = "jugador" if "jugador" in df_tr.columns else "jugadora"
            df_tr["jugador"] = df_tr[col_jtr].fillna("")
            if jug_tram_acum not in df_tr["jugador"].values: continue

            eq_id_tr = df_tr[df_tr["jugador"]==jug_tram_acum]["idEquip"].iloc[0]
            rival_tr = [t for t in teams_tr if t!=eq_id_tr]
            rival_id_tr = rival_tr[0] if rival_tr else None

            MINS_Q_TR = 10
            intervals_tr = []; ep_tr = {}
            dj_tr = df_tr[df_tr["jugador"]==jug_tram_acum]
            primer_tr = dj_tr.sort_values("num").iloc[0]
            if "Surt" in str(primer_tr.get("accio","")) and "camp" in str(primer_tr.get("accio","")):
                ep_tr[jug_tram_acum] = (int(primer_tr.get("quart",1))-1)*MINS_Q_TR

            for _,row_tr in df_tr.sort_values("num").iterrows():
                if str(row_tr.get("jugador","")) != jug_tram_acum: continue
                acc_tr = str(row_tr.get("accio",""))
                q_tr = int(row_tr.get("quart",1))
                m_tr = float(row_tr.get("min_num",0))
                t_tr = (q_tr-1)*MINS_Q_TR + (MINS_Q_TR-m_tr if m_tr<=MINS_Q_TR else m_tr)
                t_tr = max(0, min(t_tr, q_tr*MINS_Q_TR))
                if "Entra" in acc_tr and "camp" in acc_tr: ep_tr[jug_tram_acum]=t_tr
                elif "Surt" in acc_tr and "camp" in acc_tr:
                    ti_tr=ep_tr.pop(jug_tram_acum,(q_tr-1)*MINS_Q_TR)
                    if t_tr>ti_tr: intervals_tr.append((float(ti_tr),float(t_tr)))
                elif "Final de període" in acc_tr:
                    if jug_tram_acum in ep_tr:
                        ti_tr=ep_tr.pop(jug_tram_acum); fi_tr=float(q_tr*MINS_Q_TR)
                        if fi_tr>ti_tr: intervals_tr.append((ti_tr,fi_tr))
            for ti_o in ep_tr.values():
                fi_o=float(df_tr["quart"].max()*MINS_Q_TR)
                if fi_o>ti_o: intervals_tr.append((float(ti_o),fi_o))
            intervals_tr=[(float(ti),float(tf)) for ti,tf in intervals_tr]

            if not intervals_tr or rival_id_tr is None: continue

            df_t_tr = df_tr.copy()
            df_t_tr["t_abs"] = df_t_tr.apply(
                lambda r: (int(r["quart"])-1)*10+(10-float(r["min_num"]))
                if float(r.get("min_num",0))<=10 else float(r.get("min_num",0)), axis=1)

            # Per cada tram, calcula el parcial a cada minut
            for ti_tr, tf_tr in intervals_tr:
                durada_tr = tf_tr - ti_tr
                if durada_tr < 1: continue
                for min_dins in range(1, min(int(durada_tr)+1, 10)):
                    t_inici_min = ti_tr + min_dins - 1
                    t_fi_min = ti_tr + min_dins
                    df_min_tr = df_t_tr[(df_t_tr["t_abs"]>=t_inici_min)&(df_t_tr["t_abs"]<t_fi_min)]
                    pf_min = int(df_min_tr[df_min_tr["idEquip"]==eq_id_tr]["punts"].sum())
                    pc_min = int(df_min_tr[df_min_tr["idEquip"]==rival_id_tr]["punts"].sum())
                    tram_minuts.setdefault(min_dins, []).append(pf_min - pc_min)

        if tram_minuts:
            mins_sorted = sorted(tram_minuts.keys())
            mitjes_tram = [round(sum(tram_minuts[m])/len(tram_minuts[m]),2) for m in mins_sorted]
            mostres_tram = [len(tram_minuts[m]) for m in mins_sorted]

            fig_tram_acum = go.Figure()
            colors_tram_acum = ["#16a34a" if v>=0 else "#dc2626" for v in mitjes_tram]
            fig_tram_acum.add_trace(go.Bar(
                x=[f"Min {m}" for m in mins_sorted],
                y=mitjes_tram,
                marker_color=colors_tram_acum,
                text=[f"{'+'if v>=0 else ''}{v:.2f}" for v in mitjes_tram],
                textposition="outside",
                customdata=mostres_tram,
                hovertemplate="Min %{x}<br>Parcial mig: %{y:+.2f}<br>Mostres: %{customdata}<extra></extra>"
            ))
            fig_tram_acum.add_hline(y=0, line_dash="solid", line_color="#e2e4e8")
            fig_tram_acum.update_layout(
                xaxis_title="Minut dins del tram",
                yaxis_title="Parcial mig de l'equip")
            st.plotly_chart(chart_style(fig_tram_acum,280,
                f"{jug_tram_acum} — parcial per minut dins del tram"),use_container_width=True)
            st.caption("Cada barra = parcial mig de l'equip en aquell minut de cada tram que juga. "
                       "Verd = l'equip guanya en aquell minut. Vermell = l'equip perd.")

            # Interpretació
            best_min = mins_sorted[mitjes_tram.index(max(mitjes_tram))]
            worst_min = mins_sorted[mitjes_tram.index(min(mitjes_tram))]
            st.info(f"📊 Millor minut del tram: **Minut {best_min}** ({max(mitjes_tram):+.2f} parcial mig) · "
                    f"Pitjor: **Minut {worst_min}** ({min(mitjes_tram):+.2f})")
        else:
            st.info("No hi ha prou dades per a aquesta jugadora.")

    # ── 3. Evolució +/- per minut al llarg de la temporada ─────────────────
    st.markdown(sec("Evolució del +/- per minut — temporada"), unsafe_allow_html=True)
    st.caption("Com ha evolucionat l'aportació per minut jugat al llarg de la temporada.")

    jug_evo = st.selectbox("Jugadora", tots_jugs_hist, key="jug_evo_pm")
    if jug_evo and pm_acum_rows and jug_evo == jug_pm:
        # Reutilitza les dades calculades
        fig_evo = go.Figure()
        fig_evo.add_trace(go.Scatter(
            x=list(range(1, len(df_pm_acum)+1)),
            y=df_pm_acum["+/- per min"].tolist(),
            mode="lines+markers",
            name="+/- per min",
            line=dict(color=COLOR_A, width=2),
            marker=dict(size=8, color=["#16a34a" if v>=0 else "#dc2626"
                                        for v in df_pm_acum["+/- per min"]]),
            text=df_pm_acum["Partit"].tolist(),
            hovertemplate="%{text}<br>+/- per min: %{y:+.3f}<extra></extra>"
        ))
        # Línia de mitjana mòbil
        if len(df_pm_acum) >= 3:
            rolling = df_pm_acum["+/- per min"].rolling(2, min_periods=1).mean()
            fig_evo.add_trace(go.Scatter(
                x=list(range(1, len(df_pm_acum)+1)),
                y=rolling.tolist(),
                mode="lines",
                name="Tendència (2P)",
                line=dict(color="#d97706", width=2, dash="dot")
            ))
        fig_evo.add_hline(y=0, line_dash="solid", line_color="#e2e4e8")
        fig_evo.add_hline(y=mitj_pm, line_dash="dot", line_color="#185FA5",
            annotation_text=f"Mitjana: {mitj_pm:+.3f}")
        fig_evo.update_xaxes(title="Número de partit")
        fig_evo.update_yaxes(title="+/- per minut")
        st.plotly_chart(chart_style(fig_evo,280,f"{jug_evo} — evolució +/- per minut"),use_container_width=True)
    else:
        st.info("Selecciona la mateixa jugadora que als apartats anteriors per veure l'evolució.")

        st.markdown(sec("Comparativa entre jugadores"), unsafe_allow_html=True)
        col_j,col_m=st.columns([2,1])
        with col_j:
            jugs_comp=st.multiselect("Jugadores (màx 4)",tots_jugs_hist,max_selections=4,key="jugs_comp")
        with col_m:
            metrica_comp=st.selectbox("Mètrica",
                ["punts","impacte","cistelles_2","cistelles_3","faltes","pts_per_min"],
                format_func=lambda x: {"punts":"Punts","impacte":"Impacte","cistelles_2":"C2",
                    "cistelles_3":"C3","faltes":"Faltes","pts_per_min":"Pts/min"}[x],
                key="metrica_comp")
        if len(jugs_comp)>=2:
            df_comp2=df_sj[df_sj["jugador"].isin(jugs_comp)].sort_values("Data")
            fig_comp2=px.line(df_comp2,x="Partit",y=metrica_comp,color="jugador",markers=True,
                labels={"jugador":"Jugadora"})
            fig_comp2.update_traces(line_width=2,marker_size=7)
            fig_comp2.update_xaxes(tickangle=-30)
            st.plotly_chart(chart_style(fig_comp2,280,f"Evolució per partit: {metrica_comp}"),use_container_width=True)

            mitt=df_comp2.groupby("jugador")[metrica_comp].mean().reset_index()
            mitt.columns=["Jugadora","Mitjana"]
            mitt=mitt.sort_values("Mitjana",ascending=False)
            mitt["Mitjana"]=mitt["Mitjana"].round(2)
            paleta=[COLOR_A,COLOR_B,"#16a34a","#d97706"]
            fig_mit=go.Figure()
            for i,row_m in mitt.iterrows():
                fig_mit.add_trace(go.Bar(x=[row_m["Jugadora"]],y=[row_m["Mitjana"]],
                    name=row_m["Jugadora"],
                    marker_color=paleta[list(mitt["Jugadora"]).index(row_m["Jugadora"])%len(paleta)],
                    text=[f"{row_m['Mitjana']:.2f}"],textposition="outside"))
            fig_mit.update_layout(showlegend=False,barmode="group")
            st.plotly_chart(chart_style(fig_mit,220,f"Mitjana per partit: {metrica_comp}"),use_container_width=True)
            st.dataframe(mitt,use_container_width=True,hide_index=True)

            metr=["punts","cistelles_2","cistelles_3","tirs_lliures","faltes","impacte"]
            labs=["Punts","C2","C3","TL","Faltes","Impacte"]
            fig_rad=go.Figure()
            for i,jug_r in enumerate(jugs_comp):
                dj_r=df_sj[df_sj["jugador"]==jug_r]
                vals=[dj_r[m].mean() for m in metr]
                maxv=[df_sj[m].max() for m in metr]
                norm=[round(v/mx*10,1) if mx>0 else 0 for v,mx in zip(vals,maxv)]
                fig_rad.add_trace(go.Scatterpolar(
                    r=norm+[norm[0]],theta=labs+[labs[0]],
                    fill="toself",name=jug_r,opacity=0.7,
                    line=dict(color=paleta[i%len(paleta)],width=2)))
            fig_rad.update_layout(
                polar=dict(bgcolor="#f9fafb",
                    radialaxis=dict(visible=True,range=[0,10],color="#9ca3af",gridcolor="#e2e4e8"),
                    angularaxis=dict(color="#374151",gridcolor="#e2e4e8")),
                paper_bgcolor="#ffffff",font=dict(color="#374151",family="Inter"),
                legend=dict(bgcolor="#fff",bordercolor="#e2e4e8",borderwidth=1),
                margin=dict(l=40,r=40,t=50,b=40),height=360,
                title=dict(text="Radar de rendiment (0–10 normalitzat)",font=dict(color="#374151",size=13)))
            st.plotly_chart(fig_rad,use_container_width=True)
        elif len(jugs_comp)==1:
            st.info("Selecciona almenys 2 jugadores.")

    # ── Històric temps morts ────────────────────────────────────────────────
    st.markdown(sec("⏸ Efectivitat dels temps morts — temporada"), unsafe_allow_html=True)
    df_to_hist = load_timeouts_db()
    if df_to_hist.empty:
        st.info("Consulta més partits per veure l'evolució dels temps morts.")
    else:
        df_pr_to = load_partits_db()
        def lp_to(mid):
            r = df_pr_to[df_pr_to["match_id"]==mid]
            if r.empty: return mid[:8]+"..."
            return f"{r.iloc[0]['nom_a']} vs {r.iloc[0]['nom_b']} ({r.iloc[0]['data_consulta'][:10]})"
        df_to_hist["Partit"] = df_to_hist["match_id"].apply(lp_to)

        equips_to = sorted(df_to_hist["equip_nom"].unique().tolist())
        eq_to = st.selectbox("Equip", equips_to, key="eq_to_hist")

        df_eq_to = df_to_hist[df_to_hist["equip_nom"]==eq_to]

        # Evolució efectivitat per partit
        evo = df_eq_to.groupby("Partit").agg(
            Total=("va_anotar","count"),
            Anotats=("va_anotar","sum"),
            Seg_mit=("segons_resposta","mean")
        ).reset_index()
        evo["Efectivitat %"] = (evo["Anotats"]/evo["Total"]*100).round(0)

        fig_to_evo = go.Figure()
        fig_to_evo.add_trace(go.Scatter(
            x=evo["Partit"], y=evo["Efectivitat %"],
            mode="lines+markers", name="Efectivitat %",
            line=dict(color=COLOR_A, width=2.5), marker=dict(size=8)))
        fig_to_evo.add_hline(y=50, line_dash="dot", line_color="#e2e4e8",
            annotation_text="50%", annotation_font_color="#9ca3af", annotation_font_size=10)
        fig_to_evo.update_layout(yaxis=dict(range=[0,100], ticksuffix="%"))
        fig_to_evo.update_xaxes(tickangle=-30)
        st.plotly_chart(chart_style(fig_to_evo, 260, f"{eq_to} — efectivitat temps morts per partit"), use_container_width=True)

        # Qui anota més després dels temps morts
        df_anotades = df_eq_to[df_eq_to["va_anotar"]==1]
        if not df_anotades.empty:
            top_jug = df_anotades.groupby("jugadora").size().reset_index(name="Cistelles post-TM")
            top_jug = top_jug.sort_values("Cistelles post-TM", ascending=False).head(8)
            fig_jug_to = px.bar(top_jug, x="jugadora", y="Cistelles post-TM",
                color_discrete_sequence=[COLOR_A], text="Cistelles post-TM",
                labels={"jugadora":"Jugadora"})
            fig_jug_to.update_traces(textposition="outside")
            st.plotly_chart(chart_style(fig_jug_to, 240,
                f"{eq_to} — qui anota després dels temps morts"), use_container_width=True)

        # Temps mitjà de resposta per partit
        fig_seg = go.Figure()
        fig_seg.add_trace(go.Scatter(
            x=evo["Partit"], y=evo["Seg_mit"],
            mode="lines+markers", name="Seg. fins cistella",
            line=dict(color="#d97706", width=2.5), marker=dict(size=8)))
        fig_seg.update_xaxes(tickangle=-30)
        fig_seg.update_yaxes(title="Segons")
        st.plotly_chart(chart_style(fig_seg, 220,
            f"{eq_to} — temps mitjà fins anotar després del temps mort"), use_container_width=True)

# ══════════════════════════════════════════════════
# TAB 6: MAPA DE TIR
# ══════════════════════════════════════════════════
with t7:
    st.markdown(sec("Mapa de tir — partit actual"), unsafe_allow_html=True)
    st.caption("Mida del cercle = volum · Color = eficiència: verd >55%, taronja 35–55%, vermell <35%")

    # ── Migració BD tirs_fcbq ───────────────────────────────────────────────
    try:
        con_t = sqlite3.connect(DB_PATH)
        con_t.execute("""CREATE TABLE IF NOT EXISTS tirs_fcbq (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id TEXT, data_consulta TEXT,
            equip_nom TEXT, x REAL, y REAL, fet INTEGER
        )""")
        con_t.commit(); con_t.close()
    except: pass

    def save_tirs_fcbq(match_id, data_consulta, equip_local, equip_visitant, tirs_local, tirs_visit):
        con_t = sqlite3.connect(DB_PATH)
        con_t.execute("DELETE FROM tirs_fcbq WHERE match_id=?", (match_id,))
        rows = []
        for t in tirs_local:
            rows.append((match_id, data_consulta, equip_local, float(t['x']), float(t['y']), 1 if t.get('fet') else 0))
        for t in tirs_visit:
            rows.append((match_id, data_consulta, equip_visitant, float(t['x']), float(t['y']), 1 if t.get('fet') else 0))
        con_t.executemany("INSERT INTO tirs_fcbq (match_id,data_consulta,equip_nom,x,y,fet) VALUES (?,?,?,?,?,?)", rows)
        con_t.commit(); con_t.close()

    def load_tirs_fcbq(match_id=None, equip_nom=None):
        con_t = sqlite3.connect(DB_PATH)
        q = "SELECT * FROM tirs_fcbq WHERE 1=1"
        params = []
        if match_id: q += " AND match_id=?"; params.append(match_id)
        if equip_nom: q += " AND equip_nom=?"; params.append(equip_nom)
        df_t = pd.read_sql(q, con_t, params=params)
        con_t.close()
        return df_t

    def dibuixa_mapa_tir_fcbq(tirs_list, titol="Mapa de tir", W=340, H=400):
        """Mig camp vertical. Cistella dalt al centre.
        Cada tir (x,y) és 0-100% relatiu al SEU PROPI mig camp individual.
        Cistella a x≈48-50%, y≈12-16% (dalt-centre).
        Mapeig directe: px=x%*W, py=y%*H (escalat per deixar marge dalt).
        """
        import math
        svg = []
        svg.append(f'<svg viewBox="0 0 {W} {H+38}" xmlns="http://www.w3.org/2000/svg" '
                   f'style="width:100%;max-width:420px;border-radius:10px">')
        svg.append(f'<rect width="{W}" height="{H}" fill="#e8e0d0" rx="6"/>')
        s  = 'stroke="#2d5a2d" stroke-width="1.5" fill="none"'
        sf = 'stroke="#2d5a2d" stroke-width="1.5" fill="rgba(255,255,255,0.18)"'

        cx_c = W // 2
        cy_c = 22

        svg.append(f'<rect x="4" y="4" width="{W-8}" height="{H-8}" {s} rx="3"/>')
        svg.append(f'<circle cx="{cx_c}" cy="{cy_c}" r="9" {sf}/>')
        svg.append(f'<circle cx="{cx_c}" cy="{cy_c}" r="3" fill="#2d5a2d"/>')

        z_w = int(W * 0.40); z_h = int(H * 0.34)
        z_x = (W - z_w) // 2
        svg.append(f'<rect x="{z_x}" y="4" width="{z_w}" height="{z_h}" {sf}/>')
        r_zona = z_w // 2
        svg.append(f'<path d="M {z_x} {4+z_h} A {r_zona} {r_zona} 0 0 0 {z_x+z_w} {4+z_h}" {s}/>')

        # Arc de triple: radi = 1.25 × distància lateral → sempre visible als dos costats
        dx_paret = cx_c - 4
        r_triple = int(dx_paret * 1.25)
        dy_paret = int(math.sqrt(max(r_triple**2 - dx_paret**2, 0)))
        y_paret = cy_c + dy_paret
        svg.append(f'<line x1="4" y1="4" x2="4" y2="{y_paret}" {s}/>')
        svg.append(f'<line x1="{W-4}" y1="4" x2="{W-4}" y2="{y_paret}" {s}/>')
        svg.append(f'<path d="M 4 {y_paret} A {r_triple} {r_triple} 0 0 0 {W-4} {y_paret}" {s}/>')

        svg.append(f'<line x1="4" y1="{H-4}" x2="{W-4}" y2="{H-4}" {s}/>')

        # ── Tirs ──────────────────────────────────────────────────────────
        # Mapeig directe: x → eix horitzontal, y → eix vertical
        # y mínim real ≈ 11 (línia de fons, a prop de la cistella, dalt)
        # y màxim ≈ 70-100 (línia de mig camp, baix)
        Y_MIN, Y_MAX = 8, 100
        fets_n = 0; fallats_n = 0
        for t in tirs_list:
            x_raw = float(t['x'])
            y_raw = float(t['y'])

            px = (x_raw / 100) * W
            py = ((y_raw - Y_MIN) / (Y_MAX - Y_MIN)) * H

            px = max(8, min(W-8, px))
            py = max(8, min(H-8, py))
            fet = bool(int(t.get('fet', 0)))

            if fet:
                fets_n += 1
                svg.append(f'<circle cx="{px:.1f}" cy="{py:.1f}" r="7" fill="#16a34a" '
                           f'opacity="0.85" stroke="white" stroke-width="1.5">'
                           f'<title>✅ Cistella</title></circle>')
            else:
                fallats_n += 1
                svg.append(f'<line x1="{px-5:.1f}" y1="{py-5:.1f}" x2="{px+5:.1f}" y2="{py+5:.1f}" '
                           f'stroke="#dc2626" stroke-width="2.5" opacity="0.8"/>')
                svg.append(f'<line x1="{px+5:.1f}" y1="{py-5:.1f}" x2="{px-5:.1f}" y2="{py+5:.1f}" '
                           f'stroke="#dc2626" stroke-width="2.5" opacity="0.8"/>')

        ef = round(fets_n / max(fets_n + fallats_n, 1) * 100)
        svg.append(f'<text x="{W//2}" y="{H+16}" font-family="Arial" font-size="11" '
                   f'fill="#374151" text-anchor="middle" font-weight="bold">{titol}</text>')
        svg.append(f'<text x="{W//2}" y="{H+32}" font-family="Arial" font-size="10" '
                   f'fill="#374151" text-anchor="middle">'
                   f'✅ {fets_n} cistelles  ❌ {fallats_n} fallats  ·  {ef}% ef.</text>')
        svg.append('</svg>')
        return ''.join(svg)

    def dibuixa_heatmap_fcbq(tirs_list, titol="Mapa de calor", mode="tots", W=340, H=400):
        """Heatmap de densitat de tirs. mode: 'tots','cistelles','fallats'"""
        import math
        # Filtra segons mode
        if mode == "cistelles":
            tirs_f = [t for t in tirs_list if t.get('fet')]
        elif mode == "fallats":
            tirs_f = [t for t in tirs_list if not t.get('fet')]
        else:
            tirs_f = tirs_list

        if not tirs_f:
            return f'<svg viewBox="0 0 {W} {H+38}" xmlns="http://www.w3.org/2000/svg"><text x="{W//2}" y="{H//2}" text-anchor="middle" font-family="Arial" font-size="14" fill="#888">Sense dades</text></svg>'

        Y_MIN, Y_MAX = 8, 100
        svg = []
        svg.append(f'<svg viewBox="0 0 {W} {H+38}" xmlns="http://www.w3.org/2000/svg" '
                   f'style="width:100%;max-width:420px;border-radius:10px">')
        svg.append(f'<rect width="{W}" height="{H}" fill="#e8e0d0" rx="6"/>')

        # Grid de cel·les per calcular densitat
        COLS, ROWS = 12, 15
        cell_w = W / COLS
        cell_h = H / ROWS
        grid = [[0]*COLS for _ in range(ROWS)]
        for t in tirs_f:
            x_raw = float(t['x'])
            y_raw = float(t['y'])
            px = (x_raw / 100) * W
            py = ((y_raw - Y_MIN) / (Y_MAX - Y_MIN)) * H
            col_g = min(int(px / cell_w), COLS-1)
            row_g = min(int(py / cell_h), ROWS-1)
            if 0 <= col_g < COLS and 0 <= row_g < ROWS:
                grid[row_g][col_g] += 1

        max_v = max(max(r) for r in grid) or 1

        # Colors del heatmap (transparent → groc → taronja → vermell)
        def heat_color(v, max_v):
            if v == 0: return None
            ratio = v / max_v
            if mode == "cistelles":
                # Verd clar → verd fosc
                r = int(200 - ratio * 150)
                g = int(240 - ratio * 100)
                b = int(150 - ratio * 130)
            elif mode == "fallats":
                # Groc → vermell
                r = int(220 + ratio * 35)
                g = int(200 - ratio * 180)
                b = int(50 - ratio * 50)
            else:
                # Blau clar → blau fosc → vermell
                if ratio < 0.5:
                    r = int(100 + ratio * 100)
                    g = int(150 + ratio * 50)
                    b = int(240 - ratio * 80)
                else:
                    r = int(200 + (ratio-0.5)*110)
                    g = int(175 - (ratio-0.5)*350)
                    b = int(200 - (ratio-0.5)*400)
            r = max(0, min(255, r))
            g = max(0, min(255, g))
            b = max(0, min(255, b))
            opacity = 0.3 + ratio * 0.65
            return f'rgba({r},{g},{b},{opacity:.2f})'

        for row_g in range(ROWS):
            for col_g in range(COLS):
                v = grid[row_g][col_g]
                color = heat_color(v, max_v)
                if color:
                    x0 = col_g * cell_w
                    y0 = row_g * cell_h
                    svg.append(f'<rect x="{x0:.1f}" y="{y0:.1f}" '
                               f'width="{cell_w:.1f}" height="{cell_h:.1f}" '
                               f'fill="{color}" rx="3"/>')

        # Línies de la pista sobre el heatmap
        s = 'stroke="#2d5a2d" stroke-width="1.5" fill="none"'
        sf = 'stroke="#2d5a2d" stroke-width="1" fill="none" stroke-dasharray="3,3"'
        cx_c = W // 2; cy_c = 22
        svg.append(f'<rect x="4" y="4" width="{W-8}" height="{H-8}" {s} rx="3"/>')
        svg.append(f'<circle cx="{cx_c}" cy="{cy_c}" r="9" stroke="#2d5a2d" stroke-width="2" fill="rgba(255,255,255,0.7)"/>')
        svg.append(f'<circle cx="{cx_c}" cy="{cy_c}" r="3" fill="#2d5a2d"/>')
        z_w = int(W*0.40); z_h = int(H*0.34); z_x = (W-z_w)//2
        svg.append(f'<rect x="{z_x}" y="4" width="{z_w}" height="{z_h}" {sf}/>')
        r_zona = z_w//2
        svg.append(f'<path d="M {z_x} {4+z_h} A {r_zona} {r_zona} 0 0 0 {z_x+z_w} {4+z_h}" {s}/>')
        dx_p = cx_c-4
        r_t = int(dx_p * 1.25)
        dy_p = int(math.sqrt(max(r_t**2-dx_p**2,0)))
        y_p = cy_c+dy_p
        svg.append(f'<line x1="4" y1="4" x2="4" y2="{y_p}" {s}/>')
        svg.append(f'<line x1="{W-4}" y1="4" x2="{W-4}" y2="{y_p}" {s}/>')
        svg.append(f'<path d="M 4 {y_p} A {r_t} {r_t} 0 0 0 {W-4} {y_p}" {s}/>')
        svg.append(f'<line x1="4" y1="{H-4}" x2="{W-4}" y2="{H-4}" {s}/>')

        # Títol i stats
        n_show = len(tirs_f)
        fets_n = sum(1 for t in tirs_f if t.get('fet'))
        svg.append(f'<text x="{W//2}" y="{H+16}" font-family="Arial" font-size="11" '
                   f'fill="#374151" text-anchor="middle" font-weight="bold">{titol}</text>')
        svg.append(f'<text x="{W//2}" y="{H+32}" font-family="Arial" font-size="10" '
                   f'fill="#374151" text-anchor="middle">{n_show} tirs · '
                   f'{fets_n} cistelles · {round(fets_n/max(n_show,1)*100)}% ef.</text>')
        svg.append('</svg>')
        return ''.join(svg)

    # ── Secció mapa FCBQ ────────────────────────────────────────────────────
    st.markdown(sec("Mapa de tir — dades de la FCBQ"), unsafe_allow_html=True)
    st.caption("Extreu les coordenades del web de la FCBQ i guarda-les aquí per veure el mapa i l'acumulat.")

    with st.expander("📋 Script per extreure les dades (copia i enganxa a la Console F12 del navegador)"):
        st.code("""// 1. Ves al web FCBQ → mapa de tir del partit
// 2. F12 → Console → escriu "allow pasting" → Enter
// 3. Enganxa aquest script → Enter
// 4. Es copiarà automàticament al portapapers

const punts = [];
document.querySelectorAll('*').forEach(el => {
  const style = el.getAttribute('style') || '';
  const left = style.match(/left:\\s*([\\d.]+)%/);
  const top = style.match(/top:\\s*([\\d.]+)%/);
  if (!left || !top) return;
  let color = '';
  const check = (e) => {
    const s = window.getComputedStyle(e);
    const bg = s.backgroundColor;
    const bc = s.borderColor;
    if (bg && bg !== 'rgba(0, 0, 0, 0)') return bg;
    if (bc && bc !== 'rgba(0, 0, 0, 0)' && bc !== 'rgb(0, 0, 0)') return bc;
    return '';
  };
  color = check(el);
  if (!color) el.querySelectorAll('*').forEach(c => { if (!color) color = check(c); });
  if (!color || (!color.includes('144') && !color.includes('220'))) return;
  const key = `${parseFloat(left[1]).toFixed(1)}_${parseFloat(top[1]).toFixed(1)}`;
  if (!punts.find(p => p.key === key)) {
    punts.push({ key, x: parseFloat(left[1]), y: parseFloat(top[1]), fet: color.includes('144') });
  }
});
copy(JSON.stringify(punts.map(p => ({x:p.x, y:p.y, fet:p.fet}))));
console.log(`✅ Copiat! Total: ${punts.length} | Cistelles: ${punts.filter(p=>p.fet).length} | Fallats: ${punts.filter(p=>!p.fet).length}`);""", language="javascript")

    df_pr_mapa = load_partits_db()
    if not df_pr_mapa.empty:
        mid_mapa = st.selectbox("Associa al partit",
            df_pr_mapa["match_id"].tolist(),
            format_func=lambda x: f"{df_pr_mapa[df_pr_mapa['match_id']==x]['nom_a'].values[0]} vs "
                                   f"{df_pr_mapa[df_pr_mapa['match_id']==x]['nom_b'].values[0]}",
            key="mid_mapa_fcbq")
        p_mapa = df_pr_mapa[df_pr_mapa["match_id"]==mid_mapa].iloc[0]
        nom_local_mapa = p_mapa["nom_a"]; nom_visit_mapa = p_mapa["nom_b"]
    else:
        mid_mapa = "manual"; nom_local_mapa = nom_a; nom_visit_mapa = nom_b

    json_tirs = st.text_area("Enganxa aquí el JSON dels tirs",
        height=80, placeholder='[{"x":46.05,"y":13.76,"fet":true},...]',
        key="json_tirs_fcbq")

    if json_tirs.strip():
        try:
            import json as json_mod
            tirs = json_mod.loads(json_tirs)
            n_total = len(tirs)

            split_idx = st.number_input(
                f"A partir de quina posició comencen els tirs de {nom_visit_mapa}? "
                f"(els primers tirs són de {nom_local_mapa})",
                min_value=0, max_value=n_total, value=n_total//2, step=1,
                key="split_idx_mapa",
                help=f"Total de tirs: {n_total}. Per exemple, si els primers 60 són del local i la resta del visitant, escriu 60."
            )

            tirs_local = tirs[:split_idx]
            tirs_visit = tirs[split_idx:]

            if st.button("💾 Guardar tirs a la BD", key="btn_save_tirs"):
                save_tirs_fcbq(mid_mapa, datetime.now().strftime("%Y-%m-%d"),
                               nom_local_mapa, nom_visit_mapa, tirs_local, tirs_visit)
                st.success(f"✅ {len(tirs)} tirs guardats! ({len(tirs_local)} {nom_local_mapa} + {len(tirs_visit)} {nom_visit_mapa})")

            eq_sel = st.radio("Mostra", ["Tots dos", nom_local_mapa, nom_visit_mapa],
                              horizontal=True, key="eq_sel_mapa")

            viz_sel = st.radio("Visualització", ["🎯 Punts individuals", "🔥 Mapa de calor", "🔥 Calor cistelles", "🔥 Calor fallats"],
                               horizontal=True, key="viz_sel_mapa")

            if eq_sel == nom_local_mapa:
                tirs_show = tirs_local
            elif eq_sel == nom_visit_mapa:
                tirs_show = tirs_visit
            else:
                tirs_show = tirs

            fets_n = sum(1 for t in tirs_show if t.get('fet'))
            tot_n = len(tirs_show)
            c1,c2,c3 = st.columns(3)
            with c1: st.markdown(card("Total tirs",tot_n,eq_sel,"#374151"),unsafe_allow_html=True)
            with c2: st.markdown(card("Cistelles",fets_n,"convertides","#16a34a"),unsafe_allow_html=True)
            with c3: st.markdown(card("Eficiència",f"{round(fets_n/max(tot_n,1)*100)}%","","#185FA5"),unsafe_allow_html=True)

            def mostra_mapa(tirs, titol, viz):
                if viz == "🎯 Punts individuals":
                    st.markdown(dibuixa_mapa_tir_fcbq(tirs, titol), unsafe_allow_html=True)
                elif viz == "🔥 Mapa de calor":
                    st.markdown(dibuixa_heatmap_fcbq(tirs, titol, "tots"), unsafe_allow_html=True)
                elif viz == "🔥 Calor cistelles":
                    st.markdown(dibuixa_heatmap_fcbq(tirs, f"{titol} — cistelles", "cistelles"), unsafe_allow_html=True)
                else:
                    st.markdown(dibuixa_heatmap_fcbq(tirs, f"{titol} — fallats", "fallats"), unsafe_allow_html=True)

            if eq_sel == "Tots dos":
                col_ma, col_mb = st.columns(2)
                with col_ma: mostra_mapa(tirs_local, nom_local_mapa, viz_sel)
                with col_mb: mostra_mapa(tirs_visit, nom_visit_mapa, viz_sel)
            else:
                mostra_mapa(tirs_show, eq_sel, viz_sel)

        except Exception as e:
            st.error(f"Error: {e}")

    # ── Mapa acumulat ───────────────────────────────────────────────────────
    st.markdown(sec("Mapa de tir acumulat — temporada"), unsafe_allow_html=True)
    df_tirs_bd = load_tirs_fcbq()
    if df_tirs_bd.empty:
        st.info("Guarda tirs de partits per veure l'acumulat de temporada.")
    else:
        partits_tirs = sorted(df_tirs_bd["match_id"].unique().tolist())
        df_pr_acum = load_partits_db()
        sel_partits = st.multiselect("Partits", partits_tirs, default=partits_tirs,
            format_func=lambda x: f"{df_pr_acum[df_pr_acum['match_id']==x]['nom_a'].values[0]} vs "
                                   f"{df_pr_acum[df_pr_acum['match_id']==x]['nom_b'].values[0]}"
                                   if not df_pr_acum.empty and x in df_pr_acum["match_id"].values else x[:8],
            key="sel_partits_mapa")
        if sel_partits:
            df_acum = df_tirs_bd[df_tirs_bd["match_id"].isin(sel_partits)]
            tirs_acum = df_acum[["x","y","fet"]].to_dict("records")
            fets_a = int(df_acum["fet"].sum()); tot_a = len(df_acum)
            c1,c2,c3 = st.columns(3)
            with c1: st.markdown(card("Total tirs",tot_a,f"{len(sel_partits)} partits","#374151"),unsafe_allow_html=True)
            with c2: st.markdown(card("Cistelles",fets_a,"convertides","#16a34a"),unsafe_allow_html=True)
            with c3: st.markdown(card("Eficiència",f"{round(fets_a/max(tot_a,1)*100)}%","","#185FA5"),unsafe_allow_html=True)

            viz_acum = st.radio("Visualització",
                ["🎯 Punts individuals","🔥 Mapa de calor","🔥 Calor cistelles","🔥 Calor fallats"],
                horizontal=True, key="viz_acum_mapa")

            titol_acum = f"{len(sel_partits)} partits seleccionats"
            if viz_acum == "🎯 Punts individuals":
                st.markdown(dibuixa_mapa_tir_fcbq(tirs_acum, titol_acum), unsafe_allow_html=True)
            elif viz_acum == "🔥 Mapa de calor":
                st.markdown(dibuixa_heatmap_fcbq(tirs_acum, titol_acum, "tots"), unsafe_allow_html=True)
            elif viz_acum == "🔥 Calor cistelles":
                st.markdown(dibuixa_heatmap_fcbq(tirs_acum, f"{titol_acum} — cistelles", "cistelles"), unsafe_allow_html=True)
            else:
                st.markdown(dibuixa_heatmap_fcbq(tirs_acum, f"{titol_acum} — fallats", "fallats"), unsafe_allow_html=True)

    col_ma,col_mb=st.columns(2)
    for col_m,tid,tnom,tcol in [
        (col_ma,teams[0] if teams else None,nom_a,COLOR_A),
        (col_mb,teams[1] if len(teams)>1 else None,nom_b,COLOR_B)
    ]:
        with col_m:
            if tid is None: st.info("Sense dades."); continue
            de=df_orig[df_orig["idEquip"]==tid]
            v1m,v1x,v2m,v2x,v3m,v3x=get_shot_counts(de)
            tot=v1m+v1x+v2m+v2x+v3m+v3x; made=v1m+v2m+v3m
            ef=round(made/tot*100) if tot>0 else 0
            c1,c2,c3=st.columns(3)
            with c1: st.markdown(card("Tirs",tot,"","#374151"),unsafe_allow_html=True)
            with c2: st.markdown(card("Convertits",made,"","#16a34a"),unsafe_allow_html=True)
            with c3: st.markdown(card("Eficiència",f"{ef}%","",tcol),unsafe_allow_html=True)
            svg_html=shot_map_svg([(v1m,v1x),(v2m,v2x),(v3m,v3x)])
            st.markdown(f"""<div style="background:#fff;border:0.5px solid #e2e4e8;border-radius:12px;padding:14px">
                <div style="font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;color:{tcol};margin-bottom:8px">{tnom}</div>
                {svg_html}</div>""",unsafe_allow_html=True)

    st.markdown("""<div style="display:flex;gap:16px;margin-top:8px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:6px;font-size:11px;color:#6b7280"><div style="width:10px;height:10px;border-radius:50%;background:#16a34a"></div>Alta (&gt;55%)</div>
        <div style="display:flex;align-items:center;gap:6px;font-size:11px;color:#6b7280"><div style="width:10px;height:10px;border-radius:50%;background:#d97706"></div>Mitja (35–55%)</div>
        <div style="display:flex;align-items:center;gap:6px;font-size:11px;color:#6b7280"><div style="width:10px;height:10px;border-radius:50%;background:#dc2626"></div>Baixa (&lt;35%)</div>
    </div>""",unsafe_allow_html=True)

    st.markdown(sec("Mapa de tir per jugadora"), unsafe_allow_html=True)
    eq_shot=st.selectbox("Equip",[nom_a,nom_b],key="shot_eq")
    eq_id_shot=teams[0] if eq_shot==nom_a else (teams[1] if len(teams)>1 else None)
    if eq_id_shot:
        jugs_shot=sorted(df_orig[(df_orig["idEquip"]==eq_id_shot)&(df_orig["jugador"]!="")]["jugador"].unique().tolist())
        jug_shot=st.selectbox("Jugadora",jugs_shot,key="shot_jug")
        if jug_shot:
            dj=df_orig[df_orig["jugador"]==jug_shot]
            tcol_s=COLOR_A if eq_shot==nom_a else COLOR_B
            v1m,v1x,v2m,v2x,v3m,v3x=get_shot_counts(dj)
            tot=v1m+v1x+v2m+v2x+v3m+v3x; made=v1m+v2m+v3m
            ef=round(made/tot*100) if tot>0 else 0
            pts_tot=v1m+v2m*2+v3m*3
            c1,c2,c3,c4=st.columns(4)
            with c1: st.markdown(card("Tirs",tot,"",tcol_s),unsafe_allow_html=True)
            with c2: st.markdown(card("Convertits",made,"","#16a34a"),unsafe_allow_html=True)
            with c3: st.markdown(card("Eficiència",f"{ef}%","",tcol_s),unsafe_allow_html=True)
            with c4: st.markdown(card("Punts",pts_tot,"anotats",tcol_s),unsafe_allow_html=True)
            col_js,col_jd=st.columns([1,1])
            with col_js:
                svg_j=shot_map_svg([(v1m,v1x),(v2m,v2x),(v3m,v3x)])
                st.markdown(f"""<div style="background:#fff;border:0.5px solid #e2e4e8;border-radius:12px;padding:14px">
                    <div style="font-size:11px;font-weight:600;color:{tcol_s};margin-bottom:8px">{jug_shot}</div>
                    {svg_j}</div>""",unsafe_allow_html=True)
            with col_jd:
                rows_d=[
                    {"Zona":"Tirs lliures (1pt)","Conv":v1m,"Fall":v1x,"Total":v1m+v1x,
                     "Ef":f"{round(v1m/(v1m+v1x)*100) if (v1m+v1x)>0 else 0}%"},
                    {"Zona":"Tirs de 2pts","Conv":v2m,"Fall":v2x,"Total":v2m+v2x,
                     "Ef":f"{round(v2m/(v2m+v2x)*100) if (v2m+v2x)>0 else 0}%"},
                    {"Zona":"Tirs de 3pts","Conv":v3m,"Fall":v3x,"Total":v3m+v3x,
                     "Ef":f"{round(v3m/(v3m+v3x)*100) if (v3m+v3x)>0 else 0}%"},
                ]
                st.dataframe(pd.DataFrame(rows_d),use_container_width=True,hide_index=True)

    # ── Evolució temporal tirs ──────────────────────────────────────────────
    st.markdown(sec("Evolució de l'eficiència per zona — temporada"), unsafe_allow_html=True)
    st.caption("Línia temporal de l'eficiència de cada zona al llarg dels partits consultats.")
    df_sz=load_shots_zones_db()
    if df_sz.empty:
        st.info("Consulta més partits per veure l'evolució temporal.")
    else:
        df_pr2=load_partits_db()
        def lp(mid):
            r=df_pr2[df_pr2["match_id"]==mid]
            if r.empty: return mid[:8]+"..."
            return f"{r.iloc[0]['nom_a']} vs {r.iloc[0]['nom_b']} ({r.iloc[0]['data_consulta'][:10]})"
        df_sz["Partit"]=df_sz["match_id"].apply(lp)

        tab_eq_sz, tab_jug_sz = st.tabs(["Per equip","Per jugadora"])

        with tab_eq_sz:
            equips_sz=sorted(df_sz[df_sz["jugador"]=="__equip__"]["equip_nom"].unique().tolist())
            eq_sz=st.selectbox("Equip",equips_sz,key="sz_eq") if equips_sz else None
            if eq_sz:
                df_eq_sz=df_sz[(df_sz["equip_nom"]==eq_sz)&(df_sz["jugador"]=="__equip__")].sort_values("data_consulta").copy()
                if not df_eq_sz.empty:
                    df_eq_sz["ef1"]=[round(r.val1_made/(r.val1_made+r.val1_miss)*100) if (r.val1_made+r.val1_miss)>0 else None for _,r in df_eq_sz.iterrows()]
                    df_eq_sz["ef2"]=[round(r.val2_made/(r.val2_made+r.val2_miss)*100) if (r.val2_made+r.val2_miss)>0 else None for _,r in df_eq_sz.iterrows()]
                    df_eq_sz["ef3"]=[round(r.val3_made/(r.val3_made+r.val3_miss)*100) if (r.val3_made+r.val3_miss)>0 else None for _,r in df_eq_sz.iterrows()]
                    fig_sz=go.Figure()
                    for col_ef,label,color_ef in [("ef1","Tirs lliures (1pt)","#6366f1"),("ef2","Tirs de 2pts",COLOR_A),("ef3","Tirs de 3pts","#16a34a")]:
                        fig_sz.add_trace(go.Scatter(x=df_eq_sz["Partit"],y=df_eq_sz[col_ef].tolist(),
                            mode="lines+markers",name=label,line=dict(color=color_ef,width=2.5),
                            marker=dict(size=8,color=color_ef),connectgaps=True))
                    fig_sz.add_hline(y=50,line_dash="dot",line_color="#e2e4e8",
                        annotation_text="50%",annotation_font_color="#9ca3af",annotation_font_size=10)
                    fig_sz.update_layout(yaxis=dict(range=[0,100],ticksuffix="%"))
                    fig_sz.update_xaxes(tickangle=-30)
                    st.plotly_chart(chart_style(fig_sz,300,f"{eq_sz} — eficiència per zona"),use_container_width=True)

        with tab_jug_sz:
            jugs_sz=sorted(df_sz[df_sz["jugador"]!="__equip__"]["jugador"].unique().tolist())
            if jugs_sz:
                jug_sz=st.selectbox("Jugadora",jugs_sz,key="sz_jug")
                if jug_sz:
                    df_jug_sz=df_sz[df_sz["jugador"]==jug_sz].sort_values("data_consulta").copy()
                    if not df_jug_sz.empty:
                        df_jug_sz["ef1"]=[round(r.val1_made/(r.val1_made+r.val1_miss)*100) if (r.val1_made+r.val1_miss)>0 else None for _,r in df_jug_sz.iterrows()]
                        df_jug_sz["ef2"]=[round(r.val2_made/(r.val2_made+r.val2_miss)*100) if (r.val2_made+r.val2_miss)>0 else None for _,r in df_jug_sz.iterrows()]
                        df_jug_sz["ef3"]=[round(r.val3_made/(r.val3_made+r.val3_miss)*100) if (r.val3_made+r.val3_miss)>0 else None for _,r in df_jug_sz.iterrows()]
                        fig_jz=go.Figure()
                        for col_ef,label,color_ef in [("ef1","Tirs lliures (1pt)","#6366f1"),("ef2","Tirs de 2pts",COLOR_A),("ef3","Tirs de 3pts","#16a34a")]:
                            fig_jz.add_trace(go.Scatter(x=df_jug_sz["Partit"],y=df_jug_sz[col_ef].tolist(),
                                mode="lines+markers",name=label,line=dict(color=color_ef,width=2.5),
                                marker=dict(size=8,color=color_ef),connectgaps=True))
                        fig_jz.add_hline(y=50,line_dash="dot",line_color="#e2e4e8",
                            annotation_text="50%",annotation_font_color="#9ca3af",annotation_font_size=10)
                        fig_jz.update_layout(yaxis=dict(range=[0,100],ticksuffix="%"))
                        fig_jz.update_xaxes(tickangle=-30)
                        st.plotly_chart(chart_style(fig_jz,300,f"{jug_sz} — eficiència per zona"),use_container_width=True)
                        with st.expander("Detall per partit"):
                            df_det=df_jug_sz[["Partit","val1_made","val1_miss","ef1","val2_made","val2_miss","ef2","val3_made","val3_miss","ef3"]].copy()
                            df_det.columns=["Partit","1pt Conv","1pt Fall","Ef 1pt%","2pts Conv","2pts Fall","Ef 2pts%","3pts Conv","3pts Fall","Ef 3pts%"]
                            st.dataframe(df_det,use_container_width=True,hide_index=True)
            else:
                st.info("Consulta més partits per veure l'evolució per jugadora.")

# ══════════════════════════════════════════════════
# TAB 7: ANÀLISI DE VÍDEO
# ══════════════════════════════════════════════════
with t8:
    st.markdown(sec("🎬 Anàlisi de vídeo — carrega els CSV del notebook"), unsafe_allow_html=True)
    st.caption("Carrega els fitxers generats pel notebook de Google Colab per veure les dades del vídeo.")

    col_u1, col_u2, col_u3 = st.columns(3)
    with col_u1:
        f_accions  = st.file_uploader("CSV d'accions", type="csv", key="up_accions",
                                       help="dades_partit_accions.csv")
    with col_u2:
        f_tracking = st.file_uploader("CSV de tracking", type="csv", key="up_tracking",
                                       help="dades_partit_tracking.csv")
    with col_u3:
        f_resum    = st.file_uploader("CSV de resum", type="csv", key="up_resum",
                                       help="dades_partit_resum.csv")

    if f_accions is None:
        st.info("Carrega almenys el CSV d'accions per començar.")
    else:
        df_vid_acc = pd.read_csv(f_accions)
        df_vid_tra = pd.read_csv(f_tracking) if f_tracking else pd.DataFrame()
        df_vid_res = pd.read_csv(f_resum)    if f_resum    else pd.DataFrame()

        st.success(f"✅ {len(df_vid_acc)} accions carregades")

        # ── Mètriques generals ──────────────────────────────────────────
        st.markdown(sec("Resum del partit"), unsafe_allow_html=True)
        equips_vid = [e for e in df_vid_acc["equip"].unique() if e and e != "?"]

        cols_m = st.columns(len(equips_vid) * 3 + 1)
        idx = 0
        cols_m[idx].markdown(card("Accions totals", len(df_vid_acc), "", "#374151"), unsafe_allow_html=True)
        idx += 1
        for eq in equips_vid:
            df_e = df_vid_acc[df_vid_acc["equip"] == eq]
            color_e = COLOR_A if idx <= 3 else COLOR_B
            cist  = int((df_e["tipus"] == "cistella").sum())
            falts = int((df_e["tipus"] == "falta").sum())
            tfall = int((df_e["tipus"] == "tir_fallat").sum())
            cols_m[idx].markdown(card(f"Cistelles {eq}", cist, "", color_e), unsafe_allow_html=True); idx+=1
            cols_m[idx].markdown(card(f"Faltes {eq}", falts, "", color_e), unsafe_allow_html=True); idx+=1
            cols_m[idx].markdown(card(f"Tirs fallats {eq}", tfall, "", color_e), unsafe_allow_html=True); idx+=1

        # ── Gràfic accions per equip ────────────────────────────────────
        st.markdown(sec("Accions per equip i tipus"), unsafe_allow_html=True)
        df_grp = df_vid_acc[df_vid_acc["equip"].isin(equips_vid)].groupby(["equip","tipus"]).size().reset_index(name="n")
        if not df_grp.empty:
            pal = {equips_vid[0]: COLOR_A, equips_vid[1]: COLOR_B} if len(equips_vid) >= 2 else {equips_vid[0]: COLOR_A}
            fig_a = px.bar(df_grp, x="tipus", y="n", color="equip", barmode="group",
                color_discrete_map=pal,
                labels={"tipus":"Tipus","n":"Accions","equip":"Equip"})
            fig_a.update_layout(xaxis_title="", paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                font=dict(color="#374151",family="Inter"),
                legend=dict(bgcolor="#ffffff",bordercolor="#e2e4e8",borderwidth=1,title=""),
                margin=dict(l=0,r=0,t=30,b=0),height=280)
            st.plotly_chart(fig_a, use_container_width=True)

        # ── Accions per quart ───────────────────────────────────────────
        st.markdown(sec("Accions per quart"), unsafe_allow_html=True)
        if "quart" in df_vid_acc.columns:
            df_q = df_vid_acc[df_vid_acc["equip"].isin(equips_vid)].groupby(["quart","equip","tipus"]).size().reset_index(name="n")
            cistelles_q = df_q[df_q["tipus"]=="cistella"]
            if not cistelles_q.empty:
                pal = {equips_vid[0]: COLOR_A, equips_vid[1]: COLOR_B} if len(equips_vid) >= 2 else {equips_vid[0]: COLOR_A}
                fig_q = px.bar(cistelles_q, x="quart", y="n", color="equip", barmode="group",
                    color_discrete_map=pal,
                    labels={"quart":"Quart","n":"Cistelles","equip":"Equip"})
                fig_q.update_layout(paper_bgcolor="#ffffff",plot_bgcolor="#ffffff",
                    font=dict(color="#374151",family="Inter"),
                    legend=dict(bgcolor="#ffffff",bordercolor="#e2e4e8",borderwidth=1,title=""),
                    margin=dict(l=0,r=0,t=30,b=0),height=260)
                st.plotly_chart(fig_q, use_container_width=True)

        # ── Evolució del marcador (des de les accions del vídeo) ────────
        st.markdown(sec("Evolució del marcador"), unsafe_allow_html=True)
        if "marcador" in df_vid_acc.columns:
            df_marc = df_vid_acc[df_vid_acc["marcador"].str.contains("-", na=False)].copy()
            if not df_marc.empty:
                try:
                    df_marc["scoreA"] = df_marc["marcador"].str.split("-").str[0].astype(int)
                    df_marc["scoreB"] = df_marc["marcador"].str.split("-").str[1].astype(int)
                    df_marc = df_marc.sort_values("temps_joc")
                    fig_m = go.Figure()
                    fig_m.add_trace(go.Scatter(x=df_marc["temps_joc"]/60, y=df_marc["scoreA"],
                        name=equips_vid[0] if equips_vid else "Local",
                        line=dict(color=COLOR_A,width=2.5),mode="lines"))
                    fig_m.add_trace(go.Scatter(x=df_marc["temps_joc"]/60, y=df_marc["scoreB"],
                        name=equips_vid[1] if len(equips_vid)>1 else "Visitant",
                        line=dict(color=COLOR_B,width=2.5),mode="lines"))
                    fig_m.update_layout(paper_bgcolor="#ffffff",plot_bgcolor="#ffffff",
                        font=dict(color="#374151",family="Inter"),
                        xaxis=dict(title="Minut de joc",showgrid=False,color="#9ca3af"),
                        yaxis=dict(title="Punts",showgrid=True,gridcolor="#f3f4f6",color="#9ca3af"),
                        legend=dict(bgcolor="#ffffff",bordercolor="#e2e4e8",borderwidth=1,title="",
                                    orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1),
                        margin=dict(l=0,r=0,t=40,b=0),height=280)
                    st.plotly_chart(fig_m, use_container_width=True)
                except:
                    st.info("No s'ha pogut generar l'evolució del marcador.")

        # ── Taula d'accions ─────────────────────────────────────────────
        st.markdown(sec("Totes les accions"), unsafe_allow_html=True)
        col_eq_f, col_tip_f = st.columns(2)
        with col_eq_f:
            eq_filter = st.selectbox("Equip", ["Tots"] + equips_vid, key="vid_eq_f")
        with col_tip_f:
            tip_filter = st.selectbox("Tipus", ["Tots","cistella","tir_fallat","falta","rebot","altre"], key="vid_tip_f")

        df_show = df_vid_acc.copy()
        if eq_filter != "Tots":
            df_show = df_show[df_show["equip"] == eq_filter]
        if tip_filter != "Tots":
            df_show = df_show[df_show["tipus"] == tip_filter]

        st.caption(f"{len(df_show)} accions")
        st.dataframe(df_show[["quart","jugadora","equip","accio","tipus","marcador"]].rename(
            columns={"quart":"Q","jugadora":"Jugadora","equip":"Equip",
                     "accio":"Acció","tipus":"Tipus","marcador":"Marc"}),
            use_container_width=True, hide_index=True, height=350)

        # ── Resum de tracking (si disponible) ───────────────────────────
        if not df_vid_res.empty:
            st.markdown(sec("Presència en pantalla per jugadora"), unsafe_allow_html=True)
            st.caption("Basada en el tracking del vídeo — quant temps apareix cada ID a càmera.")
            st.dataframe(df_vid_res.sort_values("minuts_visibles", ascending=False).rename(
                columns={"track_id":"ID","equip":"Equip","aparicions":"Frames",
                         "minuts_visibles":"Minuts visibles"}),
                use_container_width=True, hide_index=True)

        # ── Descàrrega combinada ────────────────────────────────────────
        st.markdown("---")
        csv_exp = df_vid_acc.to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Descarregar accions CSV", csv_exp,
            "accions_video.csv", "text/csv")
